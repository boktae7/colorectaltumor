import os
from re import L
from tkinter import E
from turtle import pd

import numpy as np

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as patches
import pandas as pd

import cv2
import openpyxl
import torch
from func import *

def cal_score(gt_landmark, pd_landmark, shape):
    TP = 0
    FP = 0
    TN = 0
    FN = 0

    tp_x1 = gt_landmark[0][0]
    if pd_landmark[0][0] > tp_x1:
        tp_x1 = pd_landmark[0][0]

    tp_x2 = gt_landmark[1][0]
    if pd_landmark[1][0] < tp_x2:
        tp_x2 = pd_landmark[1][0]

    tp_y1 = gt_landmark[0][1]
    if pd_landmark[0][1] > tp_y1:
        tp_y1 = pd_landmark[0][1]
        
    tp_y2 = gt_landmark[1][1]
    if pd_landmark[1][1] < tp_y2:
        tp_y2 = pd_landmark[1][1]

    ## check intersection
    is_inter = False
    if (tp_x2 - tp_x1) >= 0:
        if (tp_y2 - tp_y1) >= 0:
            is_inter = True
    
    if is_inter:
        TP = (tp_x2 - tp_x1 + 1) * (tp_y2 - tp_y1 + 1)
    else:
        TP = 0
    FP = (pd_landmark[1][0] - pd_landmark[0][0] + 1) * (pd_landmark[1][1] - pd_landmark[0][1] + 1) - TP
    FN = (gt_landmark[1][0] - gt_landmark[0][0] + 1) * (gt_landmark[1][1] - gt_landmark[0][1] + 1) - TP
    TN = int(shape[0] * shape[1]) - TP - FP - FN - TN
    return TP, FP, TN, FN

def Cal_Result(TP, FP, TN, FN):
    ## Return dice, specitificity, sensitivity, precision
    # TP, FP, TN, FN = perf_measure(gt_img, pred_img)

    dice_value = 2 * TP / ((TP + FP) + (TP + FN))
    sensitivity = TP / (TP + FN)
    specitificity = TN / (FP + TN)
    iou_value = TP / (TP + FP + FN)

    if (TP + FP) != 0:
        precision = TP / (TP + FP)
    else : 
        precision = 0

    return iou_value, dice_value, specitificity, sensitivity, precision

def patient_score(data_path):
    is_first = True
    patient_list = []
    for file_name in os.listdir(data_path):
        if "~" not in file_name and "Slice_" in file_name and "Patient" not in file_name:
            print(f"{file_name} Patient analysis....")
            excel = pd.read_excel(f"{data_path}{file_name}")
            target_list = list(excel.columns[8:])
            target_num = len(target_list)
            target_list.insert(0, 'Patient')
    
            if is_first:
                # is_first = False
                patient_list = []
                for num in range(len(excel["Patient"])):
                    patient_list.append(excel["Patient"][num].split("_")[1])
                patient_list = sorted(set(patient_list))
            
            name = file_name.split('_')

            path_excel_result = data_path + 'Patient_'
            for n_idx in range(0, len(name)):
                path_excel_result += name[n_idx] + '_'
            path_excel_result = path_excel_result[:-1]
            wb_result = openpyxl.Workbook()
            worksheet = wb_result.active
            worksheet.append(target_list)
            wb_result.save(path_excel_result)
             
            patient_result = []
            for patient in patient_list:
                score_list = [] 
                for num in range(len(excel["Patient"])):
                    if patient in excel["Patient"][num]:
                        for t_num in range(target_num):
                            score_list.append(excel[target_list[t_num+1]][num])
                avg_list = []
                for i in range(target_num):
                    score = []
                    for j in range(len(score_list)):
                        if (j % target_num) == i:
                            score.append(score_list[j])
                    avg_list.append(round(np.mean(score), 4))
                avg_list.insert(0, patient)
                patient_result.append(avg_list)

            wb_result = openpyxl.load_workbook(path_excel_result)
            ws = wb_result.active
            for p in range(len(patient_result)):
                ws.append(patient_result[p])
            wb_result.save(path_excel_result)                        
    print("Patient Analysis Finish")

def patient_score_boxinfo(data_path):
    is_first = True
    patient_list = []
    for file_name in os.listdir(data_path):
        if "~" not in file_name and "BoxInfo_" in file_name and "Patient" not in file_name:
            print(f"{file_name} Patient Box analysis....")
            excel = pd.read_excel(f"{data_path}{file_name}")
            target_list = []
            for col in range(len(list(excel.columns))):
                if "_S" in excel.columns[col]:
                    target_list.append(excel.columns[col])
            target_num = len(target_list)
            target_list.insert(0, 'Patient')
    
            if is_first:
                # is_first = False
                patient_list = []
                for num in range(len(excel["Patient"])):
                    patient_list.append(excel["Patient"][num].split("_")[1])
                patient_list = sorted(set(patient_list))
            
            name = file_name.split('_')

            path_excel_result = data_path + 'Patient_'
            for n_idx in range(0, len(name)):
                path_excel_result += name[n_idx] + '_'
            path_excel_result = path_excel_result[:-1]
            wb_result = openpyxl.Workbook()
            worksheet = wb_result.active
            worksheet.append(target_list)
            wb_result.save(path_excel_result)
             
            patient_result = []
            for patient in patient_list:
                score_list = [] 
                for num in range(len(excel["Patient"])):
                    if patient in excel["Patient"][num]:
                        for t_num in range(target_num):
                            score_list.append(excel[target_list[t_num+1]][num])
                avg_list = []
                for i in range(target_num):
                    score = []
                    for j in range(len(score_list)):
                        if (j % target_num) == i:
                            score.append(score_list[j])
                    avg_list.append(round(np.mean(score), 4))
                avg_list.insert(0, patient)
                patient_result.append(avg_list)

            wb_result = openpyxl.load_workbook(path_excel_result)
            ws = wb_result.active
            for p in range(len(patient_result)):
                ws.append(patient_result[p])
            wb_result.save(path_excel_result)                        
    print("Patient Analysis Finish")

def cal_cutgt(ori_gt, cut_land):
    ratio_x = 128 / (cut_land[2] - cut_land[0])
    ratio_y = 128 / (cut_land[3] - cut_land[1])
    cut_gt = []
    if int(ori_gt[0][0]) > 0:
        cut_x1 = ori_gt[0][0]-cut_land[0]
        cut_y1 = ori_gt[0][1]-cut_land[1]
        cut_x2 = ori_gt[1][0]-cut_land[0]
        cut_y2 = ori_gt[1][1]-cut_land[1]
        cut_gt = np.array([[cut_x1 * ratio_x, cut_y1 * ratio_y], [cut_x2 * ratio_x, cut_y2 * ratio_y]], dtype=np.int64)
    else:
        cut_gt = np.array([[0, 0], [1, 1]], dtype=np.int64)

    return cut_gt, [ratio_x, ratio_y]

def s2_results(dir_excel):
    print(f"{dir_excel} results analysis....")

    results_excel = f"{dir_excel[:-5]}_results.xlsx"
 
    patient_list = []
    slice_list = []
    is_nomral_list_pd = []
    is_normal_list_gt = []
    score_list = []
    excel = pd.read_excel(f"{dir_excel}")
    cur_slice = excel["Slice"][0][:-2]
    pre_slice = 'pre'
    max_score = 0
    pd_marks = []
    results_list = []
    Is_last = False
    Is_Frist = False
    iou_list = []
    dice_list = []
    recall_list = []
    preci_list = []
    iou_list_con = []
    dice_list_con = []
    recall_list_con = []
    preci_list_con = []
    for num in range(len(excel["Patient"])): 
        slice_idx = excel["Slice"][num]
        slice_num = slice_idx[:-2]
        if excel["Slice"][0][:-2] != excel["Slice"][1][:-2]:
            Is_Frist = True
        # if cur_slice == 'start':
        #     slice_list.append(cur_slice)
        #     patient_list.append(patient)
        pd_marks = float(excel["Pred 2"][num].split(',')[0])
       
        if slice_num != cur_slice: 
            slice_list.append(cur_slice)
            patient_list.append(patient)
            is_normal_list_gt.append(excel["Is_Normal"][num-1])
            score_list.append(max_score)
            max_score = 0
            is_normal = "O"
            for r_num in range(len(results_list)):
                if results_list[r_num] > 1:
                    is_normal = "X"
            
            results_list = []
            is_nomral_list_pd.append(is_normal)

            iou_list.append(np.mean(iou_list_con))
            dice_list.append(np.mean(dice_list_con))
            recall_list.append(np.mean(recall_list_con))
            preci_list.append(np.mean(preci_list_con))
            iou_list_con = []
            dice_list_con = []
            recall_list_con = []
            preci_list_con = []
            
            cur_slice = slice_num
            if num == (len(excel["Patient"])-1):
                Is_last = True
                slice_list.append(cur_slice)
                patient_list.append(patient)
                results_list.append(pd_marks)
                is_normal_list_gt.append(excel["Is_Normal"][num])
                # score = float(excel["Score_1st"][num])
                score_list.append(float(excel["Score_1st"][num]))
                iou_list.append(float(excel["IoU"][num]))
                dice_list.append(float(excel["Dice"][num]))
                recall_list.append(float(excel["Recall"][num]))
                preci_list.append(float(excel["Precision"][num]))
                # max_score = 0
                is_normal = 'O'
                for r_num in range(len(results_list)):
                    if results_list[r_num] > 1:
                        is_normal = 'X'
                results_list = []
                is_nomral_list_pd.append(is_normal)
                cur_slice = slice_num
        results_list.append(pd_marks)
        score = float(excel["Score_1st"][num])
        patient = excel["Patient"][num] 
        iou_list_con.append(float(excel["IoU"][num]))
        dice_list_con.append(float(excel["Dice"][num]))
        recall_list_con.append(float(excel["Recall"][num]))
        preci_list_con.append(float(excel["Precision"][num]))
        # print(f"{patient} // {slice_num} / {cur_slice} // {max_score} / {score}")
        # if num == 20:
        #     exit()
        if max_score < score:
            max_score = score
        if num == (len(excel["Patient"])-1) and not Is_last:
            slice_list.append(cur_slice)
            patient_list.append(patient)
            is_normal_list_gt.append(excel["Is_Normal"][num])
            score_list.append(max_score)
            max_score = 0
            iou_list.append(np.mean(iou_list_con))
            dice_list.append(np.mean(dice_list_con))
            recall_list.append(np.mean(recall_list_con))
            preci_list.append(np.mean(preci_list_con))
            iou_list_con = []
            dice_list_con = []
            recall_list_con = []
            preci_list_con = []
            is_normal = 'O'
            for r_num in range(len(results_list)):
                if results_list[r_num] > 1:
                    is_normal = 'X'
                    # print(cur_slice)
            results_list = []
            is_nomral_list_pd.append(is_normal)
            cur_slice = slice_num
        
    # patient_list = set(patient_list)
    # print(slice_list)
    # print(is_nomral_list)
    # print(patient_list)
    # print(len(patient_list), len(slice_list), len(is_normal_list_gt), len(is_nomral_list_pd))
    wb_result = openpyxl.Workbook()
    worksheet = wb_result.active
    worksheet.append(['Patient', 'Slice', 'GT_Nomral', 'PD_Normal', 'IoU', 'Dice', 'Recall', 'Precision', 'Score'])
    wb_result.save(results_excel)

    s2_num = len(patient_list)
    s2_t_num = 0
    s2_n_num = 0
    s2_t_num_acc = 0
    s2_n_num_acc = 0
    wb_eval = openpyxl.load_workbook(results_excel)
    ws = wb_eval.active
    for s in range(len(patient_list)):
        reulst_list = [patient_list[s], slice_list[s], is_normal_list_gt[s], is_nomral_list_pd[s], iou_list[s], dice_list[s], recall_list[s], preci_list[s], score_list[s]]
        ws.append(reulst_list)
        if is_normal_list_gt[s] == "O":
            s2_n_num += 1
            if is_nomral_list_pd[s] == "O":
                s2_n_num_acc += 1
        else:
            s2_t_num += 1
            if is_nomral_list_pd[s] == "X":
                s2_t_num_acc += 1
    wb_eval.save(results_excel)
    # print(s2_num, s2_t_num, s2_n_num)
    # print(s2_t_num_acc, s2_n_num_acc)

    print("Patient Analysis Finish")
    return s2_num, s2_t_num, s2_n_num, s2_t_num_acc, s2_n_num_acc

def s2_results_normal(dir_excel):
    results_excel = f"{dir_excel[:-5]}_results.xlsx"
 
    excel = pd.read_excel(f"{dir_excel}")
    patient_list = []
    slice_list = []
    score_list = []
    gt_1_list = []
    gt_2_list = []
    pd_1_list = []
    pd_2_list = []

    iou_list = []
    dice_list = []
    recall_list = []
    preci_list = []
    speci_list = []
    for num in range(len(excel["Patient"])): 
        patient_list.append(excel["Patient"][num])
        slice_list.append(excel["Slice"][num])
        gt_1_list.append(excel["GT 1"][num])
        gt_2_list.append(excel["GT 2"][num])
        if int(excel['Pred 2'][num][0]) == 1:
            pd_1_list.append(excel['Pred 1'][num])
            pd_2_list.append(excel['Pred 2'][num])
        else:
            pd_1_x = int(excel['Pred 1'][num].split(',')[0]) * 2
            pd_1_y = int(excel['Pred 1'][num].split(',')[1:]) * 2
            pd_2_x = int(excel['Pred 2'][num].split(',')[0]) * 2
            pd_2_y = int(excel['Pred 2'][num].split(',')[1:]) * 2

            gt = [[0,0],[1,1]]
            pd = [[pd_1_x, pd_1_y],[pd_2_x, pd_2_y]]
            TP, FP, TN, FN = cal_score(gt, pd, (512, 512))
            iou, dice, specitificity, sensitivity, precision = Cal_Result(TP, FP, TN, FN)
            iou_list.append(iou)
            dice_list.append(dice)
            speci_list.append(specitificity)
            recall_list.append(sensitivity)
            preci_list.append(precision)
            pd_1_list.append(f"{pd_1_x}, {pd_1_y}")
            pd_2_list.append(f"{pd_2_x}, {pd_2_y}")

        
    wb_result = openpyxl.Workbook()
    worksheet = wb_result.active
    worksheet.append(['Patient', 'Slice', 'GT_Nomral', 'PD_Normal', 'IoU', 'Dice', 'Recall', 'Precision', 'Score'])
    wb_result.save(results_excel)

    s2_num = len(patient_list)
    s2_t_num = 0
    s2_n_num = 0
    s2_t_num_acc = 0
    s2_n_num_acc = 0
    wb_eval = openpyxl.load_workbook(results_excel)
    ws = wb_eval.active
    for s in range(len(patient_list)):
        reulst_list = [patient_list[s], slice_list[s], is_normal_list_gt[s], is_nomral_list_pd[s], iou_list[s], dice_list[s], recall_list[s], preci_list[s], score_list[s]]
        ws.append(reulst_list)
        if is_normal_list_gt[s] == "O":
            s2_n_num += 1
            if is_nomral_list_pd[s] == "O":
                s2_n_num_acc += 1
        else:
            s2_t_num += 1
            if is_nomral_list_pd[s] == "X":
                s2_t_num_acc += 1
    wb_eval.save(results_excel)
def detradd_net(save_dir, name_dir, network, device, input_tensor_cpu, gt_landmarks_cpu, detr_out, select_boxes, box_score_choice, cut_landmarks, number_list, pred_shape, is_save, is_o2, is_multi):   
    batch_size = input_tensor_cpu.shape[0]
    
    ## Hourglass Network
    new_pd_landmarks = []
    pd_heatmap_l1 = []
    pd_heatmap_l2 = []
    ori_pd_landmarks = []
    gt_shape = 512
    in_shape = 128
    ratio_gt_pd = int(gt_shape / pred_shape)
    ratio_in_pd = int(in_shape / pred_shape)
    for num in range(batch_size):
        if not is_multi:
            ori_img = input_tensor_cpu[num][0]
        else:
            ori_img = input_tensor_cpu[num][1]
        cut = cut_landmarks[num] 
        cut_img = ori_img[cut[1]:cut[1]+(cut[3]-cut[1]), cut[0]:cut[0]+(cut[2]-cut[0])]
        cut_img = cv2.resize(cut_img, dsize=(in_shape,in_shape), interpolation=cv2.INTER_AREA)
        # new_land = np.reshape(cut, (2,2))
        if cut[2] > 50:
            _, ratio = cal_cutgt(gt_landmarks_cpu[num], cut)
            input_tensor = torch.tensor(cut_img, dtype=torch.float64, device=device)
            input_tensor = input_tensor.view(1, 1, input_tensor.shape[0], input_tensor.shape[1]).float()
            pred_tensor = network(input_tensor)
            pred_slices = pred_tensor.detach().cpu().numpy()[0]
            pd_heatmap_l1.append(pred_slices[0])
            if not is_o2:
                pd_heatmap_l2.append(pred_slices[1])
            pred_land = []
            
            if not is_o2:
                for i in range(2):
                    landmark = convert_heatmap_to_landmark(pred_slices[i, :, :]) * ratio_in_pd 
                    pred_land.append(landmark)
            else:
                pred_land = convert_heatmap_to_landmark_2output(pred_slices[0, :, :]) * ratio_in_pd 

            pred_land = np.array(pred_land)
            pred_land = landmark_exception_cut(pred_land)
            ori_pd_landmarks.append(pred_land)
            if pred_land[1][0] > 80 and pred_land[1][1] > 80:
                tmp_land = np.array([[pred_land[0][0] * (1/ratio[0]), pred_land[0][1] * (1/ratio[1])], [pred_land[1][0] * (1/ratio[0]), pred_land[1][1] * (1/ratio[1])]])
                new_land = np.array([[tmp_land[0][0]+cut[0], tmp_land[0][1]+cut[1]],[tmp_land[1][0]+cut[0], tmp_land[1][1]+cut[1]]], dtype=np.int64)
                new_land = landmark_exception(new_land)
            else:
                new_land = np.array([[0, 0],[0, 0]])
        else:
            new_land = np.array([[0, 0],[0, 0]])
            pd_heatmap_l1.append(np.zeros((pred_shape, pred_shape)))
            if not is_o2:
                pd_heatmap_l2.append(np.zeros((pred_shape, pred_shape)))
            ori_pd_landmarks.append(new_land)
        new_pd_landmarks.append(new_land)
    ## Save Img 
    if is_save:
        if batch_size > 8:
            batch_size = 8
        fig_col = int(batch_size/2)

        ## detection result
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            if not is_multi:
                ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ax.imshow(input_tensor_cpu[num][1], cmap='gray')
            gt = gt_landmarks_cpu[num] 
            pds = detr_out[num]['boxes']
            for i in range(len(pds)):
                for j in range(4):
                    if pds[i][j] < 0:
                        pds[i][j] = 0
                ax.add_patch(patches.Rectangle((int(pds[i][0]), int(pds[i][1])), int(pds[i][2]) - int(pds[i][0]), int(pds[i][3]) - int(pds[i][1]), edgecolor = 'blue', fill=False))
            ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
            score = 'BG'
            if len(pds) > 0:
                best_idx = torch.argmax(detr_out[num]['scores'])
                score = round(float(detr_out[num]['scores'][best_idx]), 3) ## best score 
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}_{score}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f"{save_dir}DT/", exist_ok=True)
        plt.savefig(f"{save_dir}DT/DT_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_0.jpg")
        plt.close()
        
        ## dt choice box 
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            if not is_multi:
                ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ax.imshow(input_tensor_cpu[num][1], cmap='gray')
            gt = gt_landmarks_cpu[num] 
            pds = select_boxes[num]
            for i in range(len(pds)):
                pd = pds[i].reshape(2,2)
                ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
                # ax.add_patch(patches.Rectangle((int(pds[i][0]), int(pds[i][1])), int(pds[i][2]) - int(pds[i][0]), int(pds[i][3]) - int(pds[i][1]), edgecolor = 'blue', fill=False))
            ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
            score = 'BG'
            if len(pds) > 0:
                best_idx = torch.argmax(detr_out[num]['scores'])
                score = round(float(detr_out[num]['scores'][best_idx]), 3) ## best score 
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}_{score}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f"{save_dir}DT_Choice/", exist_ok=True)
        plt.savefig(f"{save_dir}DT_Choice/Choice_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_1.jpg")
        plt.close()

        ## dt output box 
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            if not is_multi:
                ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ax.imshow(input_tensor_cpu[num][1], cmap='gray')
            gt = gt_landmarks_cpu[num] 
            cut = cut_landmarks[num] 
            ax.add_patch(patches.Rectangle((cut[0], cut[1]), cut[2] - cut[0], cut[3] - cut[1], edgecolor = 'blue', fill=False))
            ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
            score = box_score_choice[num]
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}_{score}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f"{save_dir}DT_Result/", exist_ok=True)
        plt.savefig(f"{save_dir}DT_Result/Ori_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_2.jpg")
        plt.close()

        ## Cut resize img
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)   
        for num in range(batch_size):
            if not is_multi:
                ori_img = input_tensor_cpu[num][0]
            else:
                ori_img = input_tensor_cpu[num][1]
            cut = cut_landmarks[num] 
            cut_img = ori_img[cut[1]:cut[1]+(cut[3]-cut[1]), cut[0]:cut[0]+(cut[2]-cut[0])]
            cut_img = cv2.resize(cut_img, dsize=(128,128), interpolation=cv2.INTER_AREA)
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            ax.imshow(cut_img, cmap='gray')
            if cut[2] > 50:
                cut_gt, _ = cal_cutgt(gt_landmarks_cpu[num], cut)
                ax.add_patch(patches.Rectangle(cut_gt[0], cut_gt[1][0] - cut_gt[0][0], cut_gt[1][1] - cut_gt[0][1], edgecolor = 'red', fill=False))        
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f"{save_dir}Cut_Img/", exist_ok=True)
        plt.savefig(f"{save_dir}Cut_Img/Cut_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_3.jpg")
        plt.close()

        ## land 1 heat map
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            ax.imshow(pd_heatmap_l1[num])
            cut = cut_landmarks[num] 
            if cut[2] > 50:
                gt, _ = cal_cutgt(gt_landmarks_cpu[num], cut)
            else:
                gt = gt_landmarks_cpu[num] / ratio_gt_pd
            pd = ori_pd_landmarks[num] / ratio_in_pd
            ax.scatter(int(gt[0][0]), int(gt[0][1]), c='r', s=10)
            ax.scatter(pd[0][0], pd[0][1], c='blue', s=10)
            if is_o2:
                ax.scatter(int(gt[1][0]), int(gt[1][1]), c='r', s=10)
                ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f'{save_dir}/HMap1/', exist_ok=True)
        plt.savefig(f"{save_dir}/HMap1/HMap1_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_4.jpg")
        plt.close()

        if not is_o2:
            ## land 2 heat map
            fig = plt.figure() 
            gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
            for num in range(batch_size):
                ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
                ax.imshow(pd_heatmap_l2[num])
                cut = cut_landmarks[num] 
                if cut[2] > 50:
                    gt, _ = cal_cutgt(gt_landmarks_cpu[num], cut)
                else:
                    gt = gt_landmarks_cpu[num] / ratio_gt_pd
                pd = ori_pd_landmarks[num] / ratio_in_pd
                ax.scatter(int(gt[1][0]), int(gt[1][1]), c='r', s=10)
                ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)
                title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
                ax.set_title(title)
                ax.axis('off')
            plt.tight_layout()
            plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
            plt.margins(0,0)
            os.makedirs(f'{save_dir}/HMap2/', exist_ok=True)
            plt.savefig(f"{save_dir}/HMap2/HMap2_{name_dir}.jpg")
            plt.savefig(f"{save_dir}{name_dir}_5.jpg")
            plt.close()
        
        ## final output
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            if not is_multi:
                ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ax.imshow(input_tensor_cpu[num][1], cmap='gray')
            gt = gt_landmarks_cpu[num]
            pd = new_pd_landmarks[num]
            ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
            ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        os.makedirs(f'{save_dir}/Result/', exist_ok=True)
        plt.savefig(f"{save_dir}/Result/Result_{name_dir}.jpg")
        plt.savefig(f"{save_dir}{name_dir}_6.jpg")
        plt.close()  

    return new_pd_landmarks

def save_result_image(save_dir, name_dir, input_tensor_cpu, gt_landmarks_cpu, pred_tensor_cpu, gt_tensor_cpu, number_list):   
    batch_size = input_tensor_cpu.shape[0]
    if batch_size > 8:
        batch_size = 8
    ratio_gt_pd = int(input_tensor_cpu.shape[2] / pred_tensor_cpu.shape[2])
    fig_col = int(batch_size/2)
    total_input = []
    total_gt_landmark = []
    total_pd_landmark = []
    total_pd_heatmap_l1 = []
    total_pd_heatmap_l2 = []
    total_gt_heatmap_l1 = []
    total_gt_heatmap_l2 = []
    for num in range(batch_size):
        input_slice = input_tensor_cpu[num][0]
        gt_landmark = gt_landmarks_cpu[num]
        pred_slices = pred_tensor_cpu[num]
        number_slice = number_list[num]
        gt_slices = gt_tensor_cpu[num]

        gt_landmarks = np.array(gt_landmark, dtype=np.int64)

        pred_landmarks = []
        for i in range(pred_slices.shape[0]): ## pred_slices.shape[0] == num landmark 
            landmark = convert_heatmap_to_landmark(pred_slices[i, :, :]) 
            # landmark = convert_heatmap_to_landmark_2output(pred_slices[i, :, :]) 
            pred_landmarks.append(landmark)
        pred_landmarks = np.array(pred_landmarks)

        total_input.append(input_slice)
        total_gt_landmark.append(gt_landmarks)
        total_pd_landmark.append(pred_landmarks)
        total_pd_heatmap_l1.append(pred_slices[0])
        total_pd_heatmap_l2.append(pred_slices[1])
        total_gt_heatmap_l1.append(gt_slices[0])
        total_gt_heatmap_l2.append(gt_slices[1])

    ## input + box
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_input[num], cmap='gray')
        gt = total_gt_landmark[num] 
        pd = total_pd_landmark[num] * ratio_gt_pd 
        # pd = total_pd_landmark[num][0] * ratio_gt_pd
        ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
        ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    os.makedirs(f'{save_dir}/Result/', exist_ok=True)
    plt.savefig(f"{save_dir}/Result/Result_{name_dir}.jpg")
    plt.close()

    ## land heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l1[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        # pd = total_pd_landmark[num][0]
        ax.scatter(int(gt[0][0]/ratio_gt_pd), int(gt[0][1]/ratio_gt_pd), c='r', s=10)
        ax.scatter(pd[0][0], pd[0][1], c='blue', s=10)

        # ax.scatter(int(gt[1][0]/ratio_gt_pd), int(gt[1][1]/ratio_gt_pd), c='r', s=10)
        # ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)

        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    os.makedirs(f'{save_dir}/HMap1/', exist_ok=True)
    plt.savefig(f"{save_dir}/HMap1/HMap1_{name_dir}.jpg")
    plt.close()

    ## land 2 heat map
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_pd_heatmap_l2[num])
        gt = total_gt_landmark[num]
        pd = total_pd_landmark[num]
        ax.scatter(int(gt[1][0]/ratio_gt_pd), int(gt[1][1]/ratio_gt_pd), c='r', s=10)
        ax.scatter(pd[1][0], pd[1][1], c='blue', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    os.makedirs(f'{save_dir}/HMap2/', exist_ok=True)
    plt.savefig(f"{save_dir}/HMap2/HMap2_{name_dir}.jpg")
    plt.close()

    ## gt heat map 1
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_gt_heatmap_l1[num])
        # ax.scatter(int(gt[0][0]/ratio_gt_pd), int(gt[0][1]/ratio_gt_pd), c='r', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    os.makedirs(f'{save_dir}/HMap1_G/', exist_ok=True)
    plt.savefig(f"{save_dir}/HMap1_G/HMap1_G_{name_dir}.jpg")
    plt.close()

    ##  gt heat map 2
    fig = plt.figure() 
    gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
    for num in range(len(total_input)):
        ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
        ax.imshow(total_gt_heatmap_l2[num])
        # ax.scatter(int(gt[1][0]/ratio_gt_pd), int(gt[1][1]/ratio_gt_pd), c='r', s=10)
        title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
        ax.set_title(title)
        ax.axis('off')
    plt.tight_layout()
    plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
    plt.margins(0,0)
    os.makedirs(f'{save_dir}/HMap2_G/', exist_ok=True)
    plt.savefig(f"{save_dir}/HMap2_G/HMap2_G_{name_dir}.jpg")
    plt.close()

def save_result_image_cut(save_npz_dir, save_img_dir, name_dir, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list, is_save, multislice, epoch, save_npz=True):   
    batch_size = input_tensor_cpu.shape[0]
    
    ## Get & Save Cut Data
    cut_imgs = []
    cut_nums = []
    cut_gts = []
    ori_pds = []
 
    if save_npz:
        for num in range(batch_size):
            if multislice:
                ori_img = input_tensor_cpu[num][1]
            else:
                ori_img = input_tensor_cpu[num][0]
            cut = cut_landmarks[num] 
            cut_img = ori_img[cut[1]:cut[1]+(cut[3]-cut[1]), cut[0]:cut[0]+(cut[2]-cut[0])]
            if cut_img is not None:
                if cut[2] > 50:
                    cut_img = cv2.resize(cut_img, dsize=(128,128), interpolation=cv2.INTER_AREA)
                    cut_imgs.append(cut_img)
                    cut_nums.append(number_list[num])
                    cut_gt, _= cal_cutgt(gt_landmarks_cpu[num], cut)
                    cut_gt = landmark_exception_cut(cut_gt)
                    cut_gts.append(cut_gt)
                    ori_pds.append(cut)
                    
                    if cut_gt[0][0] < 0:
                        print(number_list[num])
                        print(cut_gt)
                        exit()
                # npz_name = f"{save_npz_dir}{number_list[num]}.npz"
                # np.savez_compressed(npz_name, cut_img = cut_img, number = number_list[num], cut_gt=cut_gt, ori_img = ori_img, ori_gt = gt_landmarks_cpu[num], cut_mark = cut)
    ## IMG Save
    if is_save and batch_size >= 8:
        if batch_size > 8:
            batch_size = 8
        fig_col = int(batch_size/2)

        ## input + box
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            if save_npz:
                if multislice:
                    ax.imshow(input_tensor_cpu[num][1], cmap='gray')
                else:
                    ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ax.imshow(input_tensor_cpu[num], cmap='gray')
            gt = gt_landmarks_cpu[num] 
            cut = cut_landmarks[num] 
            ax.add_patch(patches.Rectangle((cut[0], cut[1]), cut[2] - cut[0], cut[3] - cut[1], edgecolor = 'blue', fill=False))
            ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        plt.savefig(f"{save_img_dir}Ori_{name_dir}.jpg")
        plt.close()

        ## Cut resize img
        fig = plt.figure() 
        gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
        for num in range(batch_size):
            if save_npz:
                if multislice:
                    ax.imshow(input_tensor_cpu[num][1], cmap='gray')
                else:
                    ax.imshow(input_tensor_cpu[num][0], cmap='gray')
            else:
                ori_img = input_tensor_cpu[num]
            cut = cut_landmarks[num] 
            cut_img = ori_img[cut[1]:cut[1]+(cut[3]-cut[1]), cut[0]:cut[0]+(cut[2]-cut[0])]
            cut_img = cv2.resize(cut_img, dsize=(128,128), interpolation=cv2.INTER_AREA)
            ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
            ax.imshow(cut_img, cmap='gray')
            if cut[2] > 50:
                cut_gt, _ = cal_cutgt(gt_landmarks_cpu[num], cut)
                ax.add_patch(patches.Rectangle(cut_gt[0], cut_gt[1][0] - cut_gt[0][0], cut_gt[1][1] - cut_gt[0][1], edgecolor = 'red', fill=False))
            title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
            ax.set_title(title)
            ax.axis('off')
        plt.tight_layout()
        plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
        plt.margins(0,0)
        plt.savefig(f"{save_img_dir}Cut_E{epoch}_{name_dir}.jpg")
        plt.close()

    return cut_imgs, cut_nums, cut_gts, ori_pds

def create_box(boxes_cpu, scores_cpu, th, box_usenum, shape_x, shape_y):
    zero_range = 80
    if int(shape_x) == 256:
        zero_range = 0
    min_x = shape_x
    min_y = shape_y
    max_x = 0
    max_y = 0
    select_boxes = []
    zero_img = np.zeros((shape_x, shape_y))
    add_input_channel = len(boxes_cpu)
    if add_input_channel >= (box_usenum+1):
        add_input_channel = box_usenum
    zero_box_count = 0
    for bbnum in range(add_input_channel):
        if boxes_cpu[bbnum][2] < zero_range or boxes_cpu[bbnum][2] < zero_range:
            zero_box_count += 1 
    for bbnum in range(add_input_channel):
        select_boxes.append(landmark_exception(boxes_cpu[bbnum]))
        for x in range(shape_x):
            if x >= boxes_cpu[bbnum][0] and x <= boxes_cpu[bbnum][2]:
                for y in range(shape_y):
                    if y >= boxes_cpu[bbnum][1] and y <= boxes_cpu[bbnum][3]:
                        if bbnum == 0:
                            zero_img[x][y] += scores_cpu[bbnum]
                        else:
                            if x >= zero_range and y >= zero_range:
                                zero_img[x][y] += scores_cpu[bbnum]
                            else:
                                if zero_box_count > (add_input_channel/2):
                                    zero_img[x][y] += scores_cpu[bbnum]
        
    for x in range(shape_x):
        for y in range(shape_y):
            if zero_img[x][y] <= th:
                zero_img[x][y] = 0.0
            else:
                if x <= min_x:
                    min_x = x
                if y <= min_y:
                    min_y = y
                if x >= max_x:
                    max_x = x
                if y >= max_y:
                    max_y = y

    return zero_img, [min_x, min_y, max_x, max_y], select_boxes

def calculate_box_score(box_usenum, create_box, pred_boxes, pred_scores, number_slice=""):
    box_list = []
    score_list = []
    score_weight_list = []
    select_score_list = []
    select_score_weight_list = []
    min_x = create_box[0]
    min_y = create_box[1]
    max_x = create_box[2]
    max_y = create_box[3]
    c_width = max_x - min_x + 1
    c_height = max_y - min_y + 1
    for i in range(box_usenum):
        pred_box = pred_boxes[i]
        pred_score = pred_scores[i]
        if min_x < pred_box[0]:
            min_x = pred_box[0]
        if min_y < pred_box[1]:
            min_y = pred_box[1]
        if max_x > pred_box[2]:
            max_x = pred_box[2]
        if max_y > pred_box[3]:
            max_y = pred_box[3]

        ## check intersection
        is_inter = False
        if (max_x - min_x) >= 0:
            if (max_y - min_y) >= 0:
                is_inter = True
        if is_inter:
            width = max_x - min_x + 1
            height = max_y - min_y + 1
            weight_score = ((width * height) / (c_width * c_height)) * pred_score 
            pred_box = pred_box.reshape(2, 2)
            box_list.append(pred_box)
            score_list.append(pred_score)
            score_weight_list.append(weight_score)
            select_score_list.append(pred_score)
            select_score_weight_list.append(weight_score)
        else:
            tmp_result = [" No ", " No "]
            box_list.append(tmp_result)
            score_list.append(pred_score)
            score_weight_list.append(0)
    return box_list, score_list, score_weight_list, select_score_list, select_score_weight_list

def get_include_landmarks(dir_data, patient, target_slice_num, pd_landmarks):
    patient_txt = f"{patient}.txt"
    target_slice = target_slice_num
    min_x = pd_landmarks[0][0]
    min_y = pd_landmarks[0][1]
    max_x = pd_landmarks[1][0]
    max_y = pd_landmarks[1][1]
    
    with open(f"{dir_data}{patient_txt}", 'r') as txt_colon:
        lines = txt_colon.readlines()
        is_first = False
        start_idx = 0
        end_idx = 0
        is_include = False
        contour_c_num_pre = 0
        for line in lines:
            slice_idx = int(line.split(',')[0])
            if slice_idx == target_slice:
                if not is_first:
                    is_first = True
                    start_idx = lines.index(line)
                    contour_t_num = int(line.split(',')[1][1:])
                    is_over_contour = [False for i in range(contour_t_num)]
               
                contour_c_num = int(line.split(',')[2][1:])
                if contour_c_num_pre != contour_c_num:
                    contour_c_num_pre = contour_c_num
                    over_count = 0
                contour_x = int(line.split(',')[3][1:])
                contour_y = int(line.split(',')[4][:-2])
                
                for p_x in range(min_x, max_x):
                    for p_y in range(min_y, max_y):
                        if contour_x == p_x and contour_y == p_y:
                            is_over_contour[contour_c_num-1] = True
                            break
                        # if contour_x == p_x and contour_y == p_y:
                        #     over_count += 1     
                        #     break         
                        # if over_count == 10:
                        #     is_over_contour[contour_c_num-1] = True
                        #     break
            elif slice_idx != target_slice and is_first:
                end_idx = lines.index(line)
                break
        if end_idx == 0:
            end_idx = start_idx + 1
        zero_img = np.zeros(shape=(512, 512))
        try:
            for t_idx in range(start_idx, end_idx):
                contour_c_num = int(lines[t_idx].split(',')[2][1:])
                contour_x = int(lines[t_idx].split(',')[3][1:])
                contour_y = int(lines[t_idx].split(',')[4][:-2])
                is_over = is_over_contour[contour_c_num-1]
                if is_over:
                    zero_img[contour_y][contour_x] = 1
            
            min_x = 511
            min_y = 511
            max_x = 0
            max_y = 0     
            for j in range(zero_img.shape[0]):
                for k in range(zero_img.shape[1]):
                    if zero_img[j][k] > 0:
                        if j < min_y:
                            min_y = j
                        if k < min_x:
                            min_x = k
                        if j > max_y:
                            max_y = j
                        if k > max_x:
                            max_x = k

            if min_x == 511:
                min_x = 0
                min_y = 0
                max_x = 1
                max_y = 1
        except:
            print('')
            print(patient)
            print(target_slice)
            print(start_idx)
            print(int(lines[start_idx].split(',')[0]))
            print(end_idx)
    
    return [[[min_x, min_y], [max_x, max_y]]], 1

def get_noraml_landmarks(dir_data, patient, target_slice_num):
    patient_txt = f"{patient}.txt"
    target_slice = target_slice_num
    min_x = 0
    min_y = 0
    max_x = 1
    max_y = 1
    contour_t_num = 0
    with open(f"{dir_data}{patient_txt}", 'r') as txt_colon:
        lines = txt_colon.readlines()
        is_first = False
        start_idx = 0
        end_idx = 1
        over_count = 0
        contour_c_num_pre = 0
        
        for line in lines:
            slice_idx = int(line.split(',')[0])
            if slice_idx == target_slice:
                if not is_first:
                    is_first = True
                    start_idx = lines.index(line)
                    contour_t_num = int(line.split(',')[1][1:])
                    # if "2010_0396" in patient_txt and 87 == target_slice:
                    #     print(contour_t_num)
                    #     print(line)
                    #     exit()
                    is_over_contour = [False for i in range(contour_t_num)]
               
                contour_c_num = int(line.split(',')[2][1:])
                if contour_c_num_pre != contour_c_num:
                    contour_c_num_pre = contour_c_num
                    over_count = 0
                contour_x = int(line.split(',')[3][1:])
                contour_y = int(line.split(',')[4][:-2])
                
                # for p_x in range(min_x, max_x):
                #     for p_y in range(min_y, max_y):
                #         if contour_x == p_x and contour_y == p_y:
                #             over_count += 1     
                #             break         
                #         if over_count == 10:
                #             is_over_contour[contour_c_num-1] = True
                #             break
            elif slice_idx != target_slice and is_first:
                end_idx = lines.index(line)
                break
       
        # contour_t_num = int(lines[start_idx].split(',')[1][1:])
        contour_c_num_pre = 1
        is_first = False
        min_x = 511
        min_y = 511
        max_x = 0
        max_y = 0  
        contour_list = []
        try:
            if end_idx > 1:
                for t_idx in range(start_idx, end_idx):
                    contour_c_num = int(lines[t_idx].split(',')[2][1:])
                    # if contour_c_num_pre != contour_c_num and not is_first:
                    #     contour_c_num_pre = contour_c_num
                    #     is_first = True
                    if contour_c_num_pre != contour_c_num:
                        contour_c_num_pre = contour_c_num
                        contour_list.append([[min_x, min_y],[max_x, max_y]])
                        min_x = 511
                        min_y = 511
                        max_x = 0
                        max_y = 0  
                    contour_x = int(lines[t_idx].split(',')[3][1:])
                    contour_y = int(lines[t_idx].split(',')[4][:-2])
                    if contour_y < min_y:
                        min_y = contour_y
                    if contour_x < min_x:
                        min_x = contour_x
                    if contour_y > max_y:
                        max_y = contour_y
                    if contour_x > max_x:
                        max_x = contour_x
                    
                    if t_idx == (end_idx-1):
                        contour_list.append([[min_x, min_y],[max_x, max_y]])
            else:
                contour_list.append([[0, 0],[1, 1]])
        except:
            print("")
            print(patient)
            print(target_slice_num)
            print(int(lines[start_idx].split(',')[0]))
            print(start_idx)
            print(end_idx)
    
    return contour_list, contour_t_num

if __name__ == "__main__":
    dir_data = "Y:/yskim/BoundingBox/result/DT1014_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_Com/Val_S/Epoch_22/"
    s2_results(f"{dir_data}Slice_s2_Val_DT1014_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_22.xlsx")
    # s2_results_normal(f"{dir_data}Slice_Val_G_DT1014_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_29_s2_copy.xlsx")
    # dir_data = "Y:/yskim\BoundingBox/result/240314_14h14m_Val_G_DT1011_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_Demo/Epoch_17/"
    # s2_results(f"{dir_data}Slice_Val_G_DT1011_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_17_s2.xlsx")
    # gt_segmark = [[166,132],[232,211]]
    # pred_landmarks = [[382,223],[442,277]]
    # shape_x = 512
    # shape_y = 512
    # TP ,_, _, _ = cal_score(gt_segmark, pred_landmarks, (shape_x, shape_y))
    # print(TP)
