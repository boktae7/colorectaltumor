# Copyright (c) Facebook, Inc. and its affiliates. All Rights Reserved
"""
Train and eval functions used in main.py
"""
from tkinter.messagebox import NO
import openpyxl
# from DETR.calculate import patient_score, patient_score_boxinfo
import math
import os
import sys
from typing import Iterable
from numpy.core.fromnumeric import shape

import torch

import util.misc as utils
from models import conv

import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import matplotlib.patches as patches

import cv2
import numpy as np
import time
from calculate import *
from func import *

def train_one_epoch(args, postprocessors, model: torch.nn.Module, criterion: torch.nn.Module, data_loader: Iterable, optimizer: torch.optim.Optimizer,  device: torch.device, epoch: int, max_norm: float = 0, ):
    model.train()
    criterion.train()
    metric_logger = utils.MetricLogger(delimiter="  ")
    metric_logger.add_meter('lr', utils.SmoothedValue(window_size=1, fmt='{value:.6f}'))
    metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
    header = 'Epoch: [{}]'.format(epoch)
    iou_types = tuple(k for k in ('segm', 'bbox') if k in postprocessors.keys())
    print_freq = 10
    idx_data = 0
    img_save_count = 0
    img_max_num = 10
    epoch_score = 0
    # multiconv = conv.Multi_Conv(5).to(device)
    for batch in metric_logger.log_every(data_loader, print_freq, header):
        input_tensor = batch['ct'].to(device)
        number_list = batch['number']
        if args.data_5slice is not None:
            input_tensor = make_multislice_train(args.data_5slice, input_tensor.shape[0], number_list)
            input_tensor = input_tensor.to(device)  
            input_tensor_rgb = input_tensor
            # input_tensor_rgb = multiconv(input_tensor) # torch.Size([batch, 1, 512, 512])
            # input_tensor_rgb = torch.tensor(input_tensor_rgb.clone(), requires_grad=False)
            # print(input_tensor.shape)
            # print(input_tensor_rgb.shape)
            # exit()
            
            # print(input_tensor_rgb.shape)
            # input_tensor_rgb = input_tensor
            
        # input_slices = input_tensor.cpu().numpy()
        input_tensor_cpu = input_tensor.detach().cpu().numpy()

        if args.data_5slice is None:
            input_slices_rgb = []
            for num in range(input_tensor.shape[0]):
                tmp_img = input_tensor_cpu[num]  ## (1, 512, 512)
                tmp_img = tmp_img.squeeze(0)
                # tmp_img = cv2.resize(tmp_img, (224, 224), interpolation=cv2.INTER_CUBIC)
                tmp_img = cv2.cvtColor(tmp_img, cv2.COLOR_GRAY2BGR)
                input_slices_rgb.append(tmp_img)
            input_tensor_rgb = torch.tensor(input_slices_rgb).to(device= device)
            input_tensor_rgb = input_tensor_rgb.transpose(1,3)
            input_tensor_rgb = input_tensor_rgb.transpose(2,3)
            input_tensor_rgb = input_tensor_rgb.to(device)

        cur_bath_size = input_tensor.shape[0]
        
        if idx_data == 0:
            Batch_Size = cur_bath_size
        idx_data += cur_bath_size
        shape_x = input_tensor.shape[2]
        shape_y = input_tensor.shape[3]
        gt_landmarks = batch['landmarks'].to(device) 
        
        # gt_landmarks = gt_landmarks.reshape(cur_bath_size, -1)
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()
        targets = []
        d = {}
        ## label =1 tumor, =2 zero
        for s_num in range(len(number_list)): 
            bbox = []   
            ## ValueError: All bounding boxes should have positive height and width
            if gt_landmarks_cpu[s_num][0][0] == gt_landmarks_cpu[s_num][1][1]:
                gt_landmarks_cpu[s_num][1][1] += 1
            if gt_landmarks_cpu[s_num][0][1] == gt_landmarks_cpu[s_num][1][1]:
                gt_landmarks_cpu[s_num][1][1] += 1
            center_x = ((gt_landmarks_cpu[s_num][0][0] + gt_landmarks_cpu[s_num][1][0]) / 2) / shape_x
            center_y = ((gt_landmarks_cpu[s_num][0][1] + gt_landmarks_cpu[s_num][1][1]) / 2) / shape_y
            bbox_w = (gt_landmarks_cpu[s_num][1][0] - gt_landmarks_cpu[s_num][0][0]) / shape_x
            bbox_y = (gt_landmarks_cpu[s_num][1][1] - gt_landmarks_cpu[s_num][0][1]) / shape_y
            bbox = [center_x, center_y, bbox_w, bbox_y]
            if "ZERO" not in number_list[s_num]:
                label = [1]
            else:
                label = [args.zero_label]
            gt_bbox = torch.tensor([bbox], dtype=torch.float64, device=device)
            d['boxes'] = gt_bbox
            d['labels'] = torch.tensor(label, dtype=torch.int64, device=device) 
            targets.append(d)     
            d = {}   
        outputs = model(input_tensor_rgb) ## return out, hs[-1].squeeze()  # return out, [100, 256]
        # print(outputs[0])
        # print(targets)
        loss_dict = criterion(outputs[0], targets) ## return losses, indices

        weight_dict = criterion.weight_dict
        losses = sum(loss_dict[0][k] * weight_dict[k] for k in loss_dict[0].keys() if k in weight_dict)
        
        print(f"Train {idx_data} / {len(data_loader)*Batch_Size} Bbox_Loss = {round(loss_dict[0]['loss_bbox'].item(),4)} CIoU_Loss = {round(loss_dict[0]['loss_giou'].item(),4)}", end="\r")
        # reduce losses over all GPUs for logging purposes
        loss_dict_reduced = utils.reduce_dict(loss_dict[0])
        loss_dict_reduced_unscaled = {f'{k}_unscaled': v for k, v in loss_dict_reduced.items()}
        loss_dict_reduced_scaled = {k: v * weight_dict[k] for k, v in loss_dict_reduced.items() if k in weight_dict}
        losses_reduced_scaled = sum(loss_dict_reduced_scaled.values())

        loss_value = losses_reduced_scaled.item()

        if not math.isfinite(loss_value):
            print("Loss is {}, stopping training".format(loss_value))
            print(loss_dict_reduced)
            sys.exit(1)

        optimizer.zero_grad()
        losses.backward()
        
        if max_norm > 0:
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm)
        optimizer.step()

        metric_logger.update(loss=loss_value, **loss_dict_reduced_scaled, **loss_dict_reduced_unscaled)
        metric_logger.update(class_error=loss_dict_reduced['class_error'])
        metric_logger.update(lr=optimizer.param_groups[0]["lr"])

        t_sizes = []
        for t_num in range(cur_bath_size):
            t_sizes.append([512, 512])
        
        orig_target_sizes = torch.tensor(t_sizes, dtype=torch.int64, device=device)
        
        results = postprocessors['bbox'](outputs[0], orig_target_sizes)

        total_pd_landmark = []
        for idx in range(len(results)):
            best_idx = int(torch.argmax(results[idx]['scores']))
            epoch_score += float(results[idx]['scores'][best_idx])
            # best_idx_list.append(best_idx_list)
            if img_save_count <= img_max_num:
                pred_landmarks = results[idx]['boxes'][best_idx].cpu().numpy().reshape(2, 2)
                total_pd_landmark.append(pred_landmarks)
        if img_save_count < img_max_num:
            img_save_count += 1
            for idx in range(len(results)):
                best_idx = int(torch.argmax(results[idx]['scores']))                 
                pred_landmarks = results[idx]['boxes'][best_idx].cpu().numpy().reshape(2, 2)
                total_pd_landmark.append(pred_landmarks)

            os.makedirs(f'{args.output_dir}/Train_IMG/Train_IMG{img_save_count}/', exist_ok=True)
            # title = f'{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            path_save = f'{args.output_dir}/Train_IMG/Train_IMG{img_save_count}/Epoch_{epoch}.jpg'
            fig_num = cur_bath_size
            if cur_bath_size > 8:
                fig_num = 8
            fig_col = int(fig_num/2)
            total_input = []
            total_gt_landmark = []
            
            for num in range(fig_num):
                if args.data_5slice is not None:
                    intput_img_idx = int(len(input_tensor_cpu[num]) / 2)                
                    input_slice = input_tensor_cpu[num][intput_img_idx]
                else:
                    input_slice = input_tensor_cpu[num][0]
                gt_landmark = gt_landmarks_cpu[num]
                gt_landmarks = np.array(gt_landmark, dtype=np.int64)
                
                total_input.append(input_slice)
                total_gt_landmark.append(gt_landmarks)
                
            ## input + box
            fig = plt.figure() 
            gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
            for num in range(len(total_input)):
                ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
                ax.imshow(total_input[num], cmap='gray')
                gt = total_gt_landmark[num]
                pd = total_pd_landmark[num]  
                ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
                ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
                title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
                ax.set_title(title)
                ax.axis('off')
            plt.tight_layout()
            plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
            plt.margins(0,0)
            plt.savefig(path_save)
            plt.close()
                 
    # gather the stats from all processes
    metric_logger.synchronize_between_processes()
    # print("Averaged stats:", metric_logger)
    epoch_score /= idx_data
    return {k: meter.global_avg for k, meter in metric_logger.meters.items()}, epoch_score

@torch.no_grad()
def evaluate(args, model, criterion, postprocessors, data_loader, device, epoch):
    # data_loader에는 coco val 이미지 5000장에 대한 정보가 들어가 있다. type( util.misc.NestedTensor, dict ) 
    model.eval()
    criterion.eval()
    """
    category_query = torch.load('category_query.pth', map_location=torch.device('cpu'))
    mask = torch.isnan(category_query)
    category_query[mask] = 0
    category_query = category_query.cuda(device)

    weight = torch.nn.Parameter( category_query )
    model.query_embed.weight = weight
    """
    metric_logger = utils.MetricLogger(delimiter="  ")
    metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
    header = 'Test:'

    iou_types = tuple(k for k in ('segm', 'bbox') if k in postprocessors.keys())
    # coco_evaluator = CocoEvaluator(base_ds, iou_types)
    # coco_evaluator.coco_eval[iou_types[0]].params.iouThrs = [0, 0.1, 0.5, 0.75]

    # panoptic_evaluator = None
    # if 'panoptic' in postprocessors.keys():
    #     panoptic_evaluator = PanopticEvaluator(
    #         data_loader.dataset.ann_file,
    #         data_loader.dataset.ann_folder,
    #         output_dir=os.path.join(output_dir, "panoptic_eval"),
    #     )
    epoch_score = 0
    idx_data = 0
    img_save_count = 0
    img_max_num = 10
    # multiconv = conv.Multi_Conv(5).to(device)
    for batch  in metric_logger.log_every(data_loader, 10, header):
        input_tensor = batch['ct'].to(device)
        number_list = batch['number']
        if args.data_5slice is not None:
            input_tensor = make_multislice_train(args.data_5slice, input_tensor.shape[0], number_list)
            input_tensor = input_tensor.to(device)    
            input_tensor_rgb = input_tensor
            # input_tensor_rgb = multiconv(input_tensor) # torch.Size([batch, 1, 512, 512])
            # input_tensor_rgb = torch.tensor(input_tensor_rgb.clone(), requires_grad=False)

        input_tensor_cpu = input_tensor.detach().cpu().numpy()
        
        if args.data_5slice is None:
            input_slices_rgb = []
            for num in range(input_tensor.shape[0]):
                tmp_img = input_tensor_cpu[num]
                tmp_img = tmp_img.squeeze(0)
                # tmp_img = cv2.resize(tmp_img, (224, 224), interpolation=cv2.INTER_CUBIC)
                tmp_img = cv2.cvtColor(tmp_img, cv2.COLOR_GRAY2BGR)
                input_slices_rgb.append(tmp_img)
            input_tensor_rgb = torch.tensor(input_slices_rgb).to(device= device)
            input_tensor_rgb = input_tensor_rgb.transpose(1,3)
            input_tensor_rgb = input_tensor_rgb.transpose(2,3)
            input_tensor_rgb = input_tensor_rgb.to(device)

        cur_bath_size = input_tensor.shape[0]
        shape_x = input_tensor.shape[2]
        shape_y = input_tensor.shape[3]
        if idx_data == 0:
            Batch_Size = cur_bath_size
        idx_data += cur_bath_size
        gt_landmarks = batch['landmarks'].to(device) 
        gt_landmarks_cpu = gt_landmarks.cpu().numpy()
        
        gt_landmarks_reshape = gt_landmarks.reshape(cur_bath_size, -1)
        gt_landmarks_cpu_reshape = gt_landmarks_reshape.cpu().numpy()

        print(f"Eval iteration {idx_data} / {len(data_loader)*Batch_Size}", end="\r")
        
        targets = []
        d = {}
        ## label =1 tumor, =0 zero
        for s_num in range(len(number_list)): 
            bbox = []   
            ## ValueError: All bounding boxes should have positive height and width
            if gt_landmarks_cpu_reshape[s_num][0] == gt_landmarks_cpu_reshape[s_num][2]:
                gt_landmarks_cpu_reshape[s_num][2] += 1
            if gt_landmarks_cpu_reshape[s_num][1] == gt_landmarks_cpu_reshape[s_num][3]:
                gt_landmarks_cpu_reshape[s_num][3] += 1
            center_x = ((gt_landmarks_cpu_reshape[s_num][0] + gt_landmarks_cpu_reshape[s_num][2]) / 2) / shape_x
            center_y = ((gt_landmarks_cpu_reshape[s_num][1] + gt_landmarks_cpu_reshape[s_num][3]) / 2) / shape_y
            bbox_w = (gt_landmarks_cpu_reshape[s_num][2] - gt_landmarks_cpu_reshape[s_num][0]) / shape_x
            bbox_y = (gt_landmarks_cpu_reshape[s_num][3] - gt_landmarks_cpu_reshape[s_num][1]) / shape_y
            bbox = [center_x, center_y, bbox_w, bbox_y]
            if "ZERO" not in number_list[s_num]:
                label = [1]
            else:
                label = [args.zero_label]
            gt_bbox = torch.tensor([bbox], dtype=torch.float64, device=device)
            d['boxes'] = gt_bbox
            d['labels'] = torch.tensor(label, dtype=torch.int64, device=device) 
            targets.append(d)     
            d = {}   
        outputs , decoder_out = model(input_tensor_rgb) # ['pred_logits'] ([2, 100, 92]),['pred_boxes'] ([2, 100, 4]) / [100, 256]      
        loss_dict , indices= criterion(outputs, targets) 
        weight_dict = criterion.weight_dict

        del input_tensor, input_tensor_rgb, gt_landmarks
        """
        for q_position, label_postion in zip(indices[0][0], indices[0][1]):
            catagory_number = int(targets[0]['labels'][label_postion])
            category_query[catagory_number] += decoder_out[q_position].cpu() # class number == targets[0]['labels'][label_postion]
            class_count[catagory_number] += 1
        """
        # reduce losses over all GPUs for logging purposes
        loss_dict_reduced = utils.reduce_dict(loss_dict)
        loss_dict_reduced_scaled = {k: v * weight_dict[k]
                                    for k, v in loss_dict_reduced.items() if k in weight_dict}
        loss_dict_reduced_unscaled = {f'{k}_unscaled': v
                                    for k, v in loss_dict_reduced.items()}
        metric_logger.update(loss=sum(loss_dict_reduced_scaled.values()),
                            **loss_dict_reduced_scaled,
                            **loss_dict_reduced_unscaled)
        metric_logger.update(class_error=loss_dict_reduced['class_error'])


        # orig_target_sizes = torch.stack([t["orig_size"] for t in targets], dim=0)
        t_sizes = []
        for t_num in range(cur_bath_size):
            t_sizes.append([512, 512])
        
        orig_target_sizes = torch.tensor(t_sizes, dtype=torch.int64, device=device)
        
        results = postprocessors['bbox'](outputs, orig_target_sizes)
        
        total_pd_landmark = []
        # for idx in range(len(results)):
        #     best_idx = int(torch.argmax(results[idx]['scores']))
        #     epoch_score += float(results[idx]['scores'][best_idx])
        #     # best_idx_list.append(best_idx_list)
        #     if img_save_count <= img_max_num:
        #         img_save_count += 1
        #         pred_landmarks = results[idx]['boxes'][best_idx].cpu().numpy().reshape(2, 2)
        #         total_pd_landmark.append(pred_landmarks)
        if img_save_count < img_max_num:
            img_save_count += 1
            for idx in range(len(results)):
                best_idx = int(torch.argmax(results[idx]['scores']))
                epoch_score += float(results[idx]['scores'][best_idx])
                pred_landmarks = results[idx]['boxes'][best_idx].cpu().numpy().reshape(2, 2)
                total_pd_landmark.append(pred_landmarks)

            os.makedirs(f'{args.output_dir}/Test_IMG/Test_IMG{img_save_count}/', exist_ok=True)
            # title = f'{int(cur_bath_size/idx_data)}/Epoch_{epoch}'
            path_save = f'{args.output_dir}/Test_IMG/Test_IMG{img_save_count}/Epoch_{epoch}.jpg'
            fig_num = cur_bath_size
            if cur_bath_size > 8:
                fig_num = 8
            fig_col = int(fig_num/2)
            total_input = []
            total_gt_landmark = []
            
            for num in range(fig_num):
                if args.data_5slice is not None:
                    intput_img_idx = int(len(input_tensor_cpu[num]) / 2)             
                    input_slice = input_tensor_cpu[num][intput_img_idx]
                else:
                    input_slice = input_tensor_cpu[num][0]
                gt_landmark = gt_landmarks_cpu[num]
                gt_landmarks = np.array(gt_landmark, dtype=np.int64)
                
                total_input.append(input_slice)
                total_gt_landmark.append(gt_landmarks)
                
            ## input + box
            fig = plt.figure() 
            gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
            for num in range(len(total_input)):
                ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
                ax.imshow(total_input[num], cmap='gray')
                gt = total_gt_landmark[num]
                pd = total_pd_landmark[num]  
                ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
                ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
                title = f'{number_list[num].split("_")[1]}_{number_list[num].split("_")[2]}'
                ax.set_title(title)
                ax.axis('off')
            plt.tight_layout()
            plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
            plt.margins(0,0)
            plt.savefig(path_save)
            plt.close()
        # print(results[0]['scores'])
        # print(results[0]['scores'].shape)
        # print(torch.max(results[0]['scores']))
        # print(torch.argmax(results[0]['scores']))
        # print(int(torch.argmax(results[0]['scores'])))



        # if 'segm' in postprocessors.keys():
        #     target_sizes = torch.stack([t["size"] for t in targets], dim=0)
        #     results = postprocessors['segm'](results, outputs, orig_target_sizes, target_sizes)
        # res = {target['image_id'].item(): output for target, output in zip(targets, results)}
        # if coco_evaluator is not None:
        #     coco_evaluator.update(res)

        # if panoptic_evaluator is not None:
        #     res_pano = postprocessors["panoptic"](outputs, target_sizes, orig_target_sizes)
        #     for i, target in enumerate(targets):
        #         image_id = target["image_id"].item()
        #         file_name = f"{image_id:012d}.png"
        #         res_pano[i]["image_id"] = image_id
        #         res_pano[i]["file_name"] = file_name

        #     panoptic_evaluator.update(res_pano)

    """
    for i in range(100):
        category_query[i] = category_query[i] / (class_count[i])
    torch.save(category_query, 'category_query.pth')
    """


    # 2500개의 validataion 값이 모든 계산된 이후에 아래의 작업이 수행된다.
    # gather the stats from all processes
    ####
    # metric_logger.synchronize_between_processes()
    # print("Averaged stats:", metric_logger)
    ####
    # if coco_evaluator is not None:
    #     coco_evaluator.synchronize_between_processes()
    # if panoptic_evaluator is not None:
    #     print("PANOPTIFDFDFSDF")
    #     panoptic_evaluator.synchronize_between_processes()

    # accumulate predictions from all images
    # if coco_evaluator is not None:
    #     coco_evaluator.accumulate()
    #     coco_evaluator.summarize()
    # panoptic_res = None
    # if panoptic_evaluator is not None:
    #     panoptic_res = panoptic_evaluator.summarize()
    ###
    stats = {k: meter.global_avg for k, meter in metric_logger.meters.items()}
    ####
    # if coco_evaluator is not None:
    #     if 'bbox' in postprocessors.keys():
    #         stats['coco_eval_bbox'] = coco_evaluator.coco_eval['bbox'].stats.tolist()
    #     if 'segm' in postprocessors.keys():
    #         stats['coco_eval_masks'] = coco_evaluator.coco_eval['segm'].stats.tolist()
    # if panoptic_res is not None:
    #     stats['PQ_all'] = panoptic_res["All"]
    #     stats['PQ_th'] = panoptic_res["Things"]
    #     stats['PQ_st'] = panoptic_res["Stuff"]
    
    # return stats, coco_evaluator
    epoch_score /= idx_data
    return stats, epoch_score

@torch.no_grad()
def evaluate_val(args, model, criterion, postprocessors, data_loader, device, output_dir):
    # data_loader에는 coco val 이미지 5000장에 대한 정보가 들어가 있다. type( util.misc.NestedTensor, dict ) 
    model.eval()
    criterion.eval()
    """
    category_query = torch.load('category_query.pth', map_location=torch.device('cpu'))
    mask = torch.isnan(category_query)
    category_query[mask] = 0
    category_query = category_query.cuda(device)

    weight = torch.nn.Parameter( category_query )
    model.query_embed.weight = weight
    """
    metric_logger = utils.MetricLogger(delimiter="  ")
    metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
    header = 'Val:'

    idx_data = 0
    iou_epoch = []
    dice_epoch = []
    specificity_epoch = []
    sensitivity_epoch = []
    precision_epoch = []
    gt_epoch = []
    pd_epoch = []
    patient_epoch = []
    slice_epoch = []
    iszero_epoch = []
    tumor_result = []
    zero_result = []
    img_save_count = 0
    score_result = []
    score_result_2 = []
    tumor_num = 0
    zero_num = 0
    correct_tumor_num = 0
    correct_zero_num_1 = 0
    correct_zero_num_2 = 0
    remain_box_list = []
    remain_score_list = []
    remain_score_weight_list = []
    select_tumor_score = []
    select_tumor_score_weight = []
    select_zero_score = []
    select_zero_score_weight = []
    select_use_score_list = []
    select_use_score_weight_list = []
    # multiconv = conv.Multi_Conv(5).to(device)
    with torch.no_grad():
        for batch  in metric_logger.log_every(data_loader, 10, header):
            input_tensor = batch['ct'].to(device)
            number_list = batch['number']
            shape_x = input_tensor.shape[2]
            shape_y = input_tensor.shape[3]
            input_tensor_cpu = input_tensor.cpu().numpy()
           
            input_slices_rgb = []
            for num in range(input_tensor.shape[0]):
                tmp_img = input_tensor_cpu[num]
                tmp_img = tmp_img.squeeze(0)
                
                tmp_img = cv2.cvtColor(tmp_img, cv2.COLOR_GRAY2BGR)
                input_slices_rgb.append(tmp_img)
            input_tensor_rgb = torch.tensor(input_slices_rgb).to(device= device)
            input_tensor_rgb = input_tensor_rgb.transpose(1,3)
            input_tensor_rgb = input_tensor_rgb.transpose(2,3)
            input_tensor_rgb = input_tensor_rgb.to(device)

            cur_bath_size = input_tensor.shape[0]
            if idx_data == 0:
                Batch_Size = cur_bath_size
            idx_data += cur_bath_size
            gt_landmarks = batch['landmarks'].to(device) 
            gt_landmarks_cpu = gt_landmarks.cpu().numpy()
            
            print(f"Eval iteration {idx_data} / {len(data_loader)*Batch_Size}", end="\r")

            gt_landmarks_reshape = gt_landmarks.reshape(cur_bath_size, -1)
            gt_landmarks_cpu_reshape = gt_landmarks_reshape.cpu().numpy()
            targets = []
            d = {}
            labels = []
            ## label =1 tumor, =2 zero
            for s_num in range(len(number_list)): 
                bbox = []   
                ## ValueError: All bounding boxes should have positive height and width
                if gt_landmarks_cpu_reshape[s_num][0] == gt_landmarks_cpu_reshape[s_num][2]:
                    gt_landmarks_cpu_reshape[s_num][2] += 1
                if gt_landmarks_cpu_reshape[s_num][1] == gt_landmarks_cpu_reshape[s_num][3]:
                    gt_landmarks_cpu_reshape[s_num][3] += 1
                center_x = ((gt_landmarks_cpu_reshape[s_num][0] + gt_landmarks_cpu_reshape[s_num][2]) / 2) / shape_x
                center_y = ((gt_landmarks_cpu_reshape[s_num][1] + gt_landmarks_cpu_reshape[s_num][3]) / 2) / shape_y
                bbox_w = (gt_landmarks_cpu_reshape[s_num][2] - gt_landmarks_cpu_reshape[s_num][0]) / shape_x
                bbox_y = (gt_landmarks_cpu_reshape[s_num][3] - gt_landmarks_cpu_reshape[s_num][1]) / shape_y
                bbox = [center_x, center_y, bbox_w, bbox_y]
                if "normal" not in number_list[s_num]:
                    label = [1]
                else:
                    label = [args.zero_label]
                gt_bbox = torch.tensor([bbox], dtype=torch.float64, device=device)
                d['boxes'] = gt_bbox
                d['labels'] = torch.tensor(label, dtype=torch.int64, device=device) 
                targets.append(d)     
                d = {} 
            outputs , decoder_out = model(input_tensor_rgb) # ['pred_logits'] ([2, 100, 92]),['pred_boxes'] ([2, 100, 4]) / [100, 256]      
            loss_dict , indices= criterion(outputs, targets) 
            weight_dict = criterion.weight_dict

            del input_tensor, input_tensor_rgb, gt_landmarks
            """
            for q_position, label_postion in zip(indices[0][0], indices[0][1]):
                catagory_number = int(targets[0]['labels'][label_postion])
                category_query[catagory_number] += decoder_out[q_position].cpu() # class number == targets[0]['labels'][label_postion]
                class_count[catagory_number] += 1
            """
            # reduce losses over all GPUs for logging purposes
            loss_dict_reduced = utils.reduce_dict(loss_dict)
            loss_dict_reduced_scaled = {k: v * weight_dict[k]
                                        for k, v in loss_dict_reduced.items() if k in weight_dict}
            loss_dict_reduced_unscaled = {f'{k}_unscaled': v
                                        for k, v in loss_dict_reduced.items()}
            metric_logger.update(loss=sum(loss_dict_reduced_scaled.values()),
                                **loss_dict_reduced_scaled,
                                **loss_dict_reduced_unscaled)
            metric_logger.update(class_error=loss_dict_reduced['class_error'])


            # orig_target_sizes = torch.stack([t["orig_size"] for t in targets], dim=0)
            t_sizes = []
            for t_num in range(cur_bath_size):
                t_sizes.append([512, 512])
            
            orig_target_sizes = torch.tensor(t_sizes, dtype=torch.int64, device=device)
            
            results = postprocessors['bbox'](outputs, orig_target_sizes)
            
            # best_idx_list = []
            best_score_list = []
            output_score_list = []
            pd_list = []
            pd_list_all = []
            pd_list_select = []
            for idx in range(len(results)):
                best_idx = int(torch.argmax(results[idx]['scores']))
                # best_idx_list.append(best_idx)
                score = round(float(results[idx]['scores'][best_idx]), 4)
                score_2 = 0
                best_score_list.append(round(score,3))
                
                score_result.append(score)
                gt_landmark = gt_landmarks_cpu[idx]
                gt_land = np.array(gt_landmark, dtype=np.int64)
                gt_epoch.append(gt_land)

                pd_land = np.array([[0,0],[0,0]], dtype=np.int64)
                if args.box_mode == "SB":
                    boxes_cpu_ori = results[idx]['boxes'].cpu().numpy().astype(np.int64)
                    scores_cpu_ori = results[idx]['scores'].cpu().numpy()
                    boxes_cpu, scores_cpu = arrage_result(boxes_cpu_ori, scores_cpu_ori)

                    output_score_list.append(round(score,3))
                    if score >= args.box_th:
                        # pd_list_select.append([results[idx]['boxes'][best_idx].cpu().numpy().astype(np.int64)])
                        # pd_land = results[idx]['boxes'][best_idx].cpu().numpy().reshape(2, 2)
                        pd_list_select.append([boxes_cpu[0]])
                        pd_land = boxes_cpu[0].reshape(2, 2)
                        pd_land = landmark_exception(pd_land)
                    # scores_cpu_ori = results[idx]['scores'].cpu().numpy()
                    if args.num_queries > 1:
                        sort_score_list = np.sort(scores_cpu_ori)[::-1]
                        score_2 = sort_score_list[1]
                        score_result_2.append(scores_cpu[1])    
                    else:         
                        score_result_2.append(scores_cpu[1])
               
                    box_list = boxes_cpu[0:4]
                    box_list = []
                    for i in range(args.box_usenum):
                        box_list.append(boxes_cpu[i].reshape(2,2))
                   
                    score_list = scores_cpu[0:4]
                    score_weight_list = [0, 0, 0, 0]
                    select_score_list = score_list
                    select_score_weihgt_list = score_list

                    remain_box_list.append(box_list)
                    remain_score_list.append(score_list)
                    remain_score_weight_list.append(score_weight_list)
                    if len(select_score_list) > 0:
                        select_use_score_list.extend(select_score_list)
                        select_use_score_weight_list.extend(select_score_weihgt_list)

                pd_epoch.append(pd_land)
                pd_list.append(pd_land)
                pd_list_all.append(results[idx]['boxes'].cpu().numpy())

                TP, FP, TN, FN = cal_score(gt_land, pd_land, [shape_x, shape_y])
                iou_value, dice_value, specitificity, sensitivity, precision = Cal_Result(TP, FP, TN, FN)
                iou_epoch.append(iou_value)
                dice_epoch.append(dice_value)
                precision_epoch.append(precision)
                sensitivity_epoch.append(sensitivity)
                specificity_epoch.append(specitificity) 

                number_slice = number_list[idx]               
                is_zero = 'X'
                if 'normal' in number_slice:
                    is_zero = "O"
                    zero_result.append([iou_value, dice_value, precision, sensitivity, specitificity, score, score_2])
                    select_zero_score.extend(select_score_list)
                    select_zero_score_weight.extend(select_score_weihgt_list)
                    zero_num += 1
                    if int(pd_land[0][0]) == 0 and int(pd_land[0][1]) == 0:
                        correct_zero_num_1 += 1
                        if int(pd_land[1][1]) == 0:
                            correct_zero_num_2 += 1
                else:
                    tumor_result.append([iou_value, dice_value, precision, sensitivity, specitificity, score, score_2])
                    select_tumor_score.extend(select_score_list)
                    select_tumor_score_weight.extend(select_score_weihgt_list)
                    tumor_num += 1
                    if int(pd_land[1][1]) != 0:
                        correct_tumor_num += 1

                slice_name = number_slice.split("_")[0] 
                patient = number_slice.split("_")[1]        
                       
                patient_epoch.append(patient)
                slice_epoch.append(slice_name)
                iszero_epoch.append(is_zero)

              
            fig_num = cur_bath_size
            if cur_bath_size > 8:
                fig_num = 8
            fig_col = int(fig_num/2)
            total_input = []
            total_gt_landmark = []
            total_pd_landmark = []
            for num in range(fig_num):            
                input_slice = input_tensor_cpu[num][0]
                gt_landmark = gt_landmarks_cpu[num]
                gt_landmarks = np.array(gt_landmark, dtype=np.int64)
                pred_landmarks = pd_list[num]

                total_input.append(input_slice)
                total_gt_landmark.append(gt_landmarks)
                total_pd_landmark.append(pred_landmarks)

            ## Batch image
            fig = plt.figure() 
            gs = gridspec.GridSpec(nrows=2, ncols=fig_col)
            for num in range(len(total_input)):
                ax = fig.add_subplot(gs[int(num/fig_col), num%fig_col])
                ax.imshow(total_input[num], cmap='gray')
                gt = total_gt_landmark[num]
                pd = total_pd_landmark[num]  
                ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
                ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
                title_score = output_score_list[num]
                if not args.box_mode == "TG":
                    if title_score != 'BG':
                        if title_score <= args.box_th:
                            title_score = 'BG'
                title = f'{number_list[num].split("_")[0]}_{number_list[num].split("_")[1]}_{title_score}'
                ax.set_title(title)
                ax.axis('off')
            plt.tight_layout()
            plt.subplots_adjust(left = 0, bottom = 0, right = 1, top = 1, hspace = 0.1, wspace = 0.1)
            plt.margins(0,0)
            plt.savefig(f'{output_dir}/Output.tiff')
            plt.close()

            ## Each image
            for num in range(len(total_input)):
                fig = plt.figure() 
                gs = gridspec.GridSpec(nrows=1, ncols=1)
                ax = fig.add_subplot(gs[0, 0])
                ax.imshow(total_input[num], cmap='gray')
                gt = total_gt_landmark[num]
                pd = total_pd_landmark[num]  
                ax.add_patch(patches.Rectangle(gt[0], gt[1][0] - gt[0][0], gt[1][1] - gt[0][1], edgecolor = 'red', fill=False))
                ax.add_patch(patches.Rectangle(pd[0], pd[1][0] - pd[0][0], pd[1][1] - pd[0][1], edgecolor = 'blue', fill=False))
                title_score = output_score_list[num]
                if not args.box_mode == "TG":
                    if title_score != 'BG':
                        if title_score <= args.box_th:
                            title_score = 'BG'
                title = f'{number_list[num].split("_")[0]}_{number_list[num].split("_")[1]}_{title_score}'
                ax.set_title(title)
                ax.axis('off')
                plt.tight_layout()
                plt.margins(0,0)
                plt.savefig(f'{output_dir}/Output_{number_list[num].split("_")[1]}_{number_list[num].split("_")[0]}.tiff')
                plt.close()

    cal_result = [iou_epoch, dice_epoch, precision_epoch, sensitivity_epoch, specificity_epoch]
    acc_result = [correct_tumor_num, correct_zero_num_1, correct_zero_num_2, round(correct_tumor_num/tumor_num, 4), round(correct_zero_num_1/zero_num, 4), round(correct_zero_num_2/zero_num, 4)]
    
    remain_result = [remain_box_list, remain_score_list, remain_score_weight_list, 
                    select_tumor_score, select_tumor_score_weight, select_zero_score, select_zero_score_weight, select_use_score_list, select_use_score_weight_list]
    return [cal_result, patient_epoch, slice_epoch, iszero_epoch, gt_epoch, pd_epoch, tumor_result, zero_result, score_result, score_result_2, acc_result, remain_result] 

def get_detr_result(args, model, criterion, postprocessors, train_data_loader, test_data_loader, device, output_dir):
    model.eval()
    criterion.eval()
    is_multi_slice = False
    if args.data_5slice is not None:
        is_multi_slice = True
    model_info = (output_dir.split('/')[-2]).split('_')
    DT_idx = 0
    for model_idx in range(len(model_info)):
        if "DT" in model_info[model_idx]:
            DT_idx = model_idx           
    DT_info = ''
    for info in range(4):
        DT_info += '_' + model_info[DT_idx+info]
    DT_info = DT_info[1:]  
    print(f"DETR model info = {DT_info}")
    DIR_Train_Cut = f"{output_dir}IMG_Result/Train_Cut/"            
    DIR_Cut_Npz_Train = f"{args.model_path}/Cut_Data/{DT_info}/Cut_Train/"
    DIR_Test_Cut = f"{output_dir}IMG_Result/Test_Cut/"
    DIR_Cut_Npz_Test = f"{args.model_path}/Cut_Data/{DT_info}/Cut_Test/"
    
    os.makedirs(DIR_Train_Cut, exist_ok=True)
    os.makedirs(DIR_Test_Cut, exist_ok=True)
    
    metric_logger = utils.MetricLogger(delimiter="  ")
    metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
    header = 'Train:'

    Train_Cut_Imgs = []
    Train_Cut_Nums = []
    Train_Cut_Gts = []
    Train_Ori_Pds = []
    Test_Cut_Imgs = []
    Test_Cut_Nums = []
    Test_Cut_Gts = []
    Test_Ori_Pds = []
    idx_data = 0
    img_save_count = 0
    is_save = True
    if os.path.isdir(DIR_Cut_Npz_Test):
        print("DETR Train file exist")
        train_file_num = len(os.listdir(DIR_Cut_Npz_Train))
        test_file_num = len(os.listdir(DIR_Cut_Npz_Test))
        batch_ori_img = []
        batch_ori_gt = []    
        batch_number = []
        batch_cutmark = []
        for file_name in os.listdir(DIR_Cut_Npz_Train):
            cut_data = np.load(DIR_Cut_Npz_Train + file_name, allow_pickle=True)
            Train_Cut_Imgs.append(cut_data['cut_img'])
            Train_Cut_Nums.append(str(cut_data['number']))
            Train_Cut_Gts.append(cut_data['cut_gt'])
            Train_Ori_Pds.append(cut_data['cut_mark'])
            idx_data += 1
            
            batch_ori_img.append(cut_data['ori_img'])
            batch_ori_gt.append(cut_data['ori_gt'])
            batch_number.append(str(cut_data['number']))
            batch_cutmark.append(cut_data['cut_mark'])
            if len(batch_ori_img) == args.batch_size or idx_data == train_file_num:
                if img_save_count < 20:
                    img_save_count += 1
                else:
                    is_save = False
                _, _, _, _ = save_result_image_cut(DIR_Cut_Npz_Train, DIR_Train_Cut, idx_data, np.array(batch_ori_img), batch_ori_gt, batch_cutmark, batch_number, is_save, multislice=is_multi_slice, save_npz=False)
                batch_ori_img = []
                batch_ori_gt = []    
                batch_number = []
                batch_cutmark = []

            print(f"DETR Train Progress {idx_data} / {train_file_num}", end="\r")
        print("")
        idx_data = 0
        img_save_count = 0
        batch_ori_img = []
        batch_ori_gt = []    
        batch_number = []
        batch_cutmark = []
        for file_name in os.listdir(DIR_Cut_Npz_Test):
            cut_data = np.load(DIR_Cut_Npz_Test + file_name, allow_pickle=True)
            Test_Cut_Imgs.append(cut_data['cut_img'])
            Test_Cut_Nums.append(str(cut_data['number']))
            Test_Cut_Gts.append(cut_data['cut_gt'])
            Test_Ori_Pds.append(cut_data['cut_mark'])
            idx_data += 1

            batch_ori_img.append(cut_data['ori_img'])
            batch_ori_gt.append(cut_data['ori_gt'])
            batch_number.append(str(cut_data['number']))
            batch_cutmark.append(cut_data['cut_mark'])
            if len(batch_ori_img) == args.batch_size or idx_data == test_file_num:
                if img_save_count < 20:
                    img_save_count += 1
                else:
                    is_save = False
                _, _, _, _ = save_result_image_cut(DIR_Cut_Npz_Test, DIR_Test_Cut, idx_data, np.array(batch_ori_img), batch_ori_gt, batch_cutmark, batch_number, is_save, multislice=is_multi_slice, save_npz=False)
                batch_ori_img = []
                batch_ori_gt = []    
                batch_number = []
                batch_cutmark = []

            print(f"DETR Test Progress {idx_data} / {test_file_num}", end="\r")
        print("")
    else:
        os.makedirs(DIR_Cut_Npz_Test, exist_ok=True)
        os.makedirs(DIR_Cut_Npz_Train, exist_ok=True)
        with torch.no_grad():
            for batch in metric_logger.log_every(train_data_loader, 10, header):
                idx_data, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list = get_cut_data(args, model, criterion, postprocessors, metric_logger, batch, idx_data, device)
                is_save = True
                if img_save_count >= 20:
                    is_save = False
                cut_img, cut_num, cut_gt, ori_pd = save_result_image_cut(DIR_Cut_Npz_Train, DIR_Train_Cut, idx_data, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list, is_save, multislice=is_multi_slice)
                Train_Cut_Imgs.extend(cut_img)
                Train_Cut_Nums.extend(cut_num)
                Train_Cut_Gts.extend(cut_gt)
                Train_Ori_Pds.extend(ori_pd)
                img_save_count += 1
                print(f"DETR Train Progress {idx_data} / {len(train_data_loader)*args.batch_size}", end="\r")
            print("")
            
            metric_logger = utils.MetricLogger(delimiter="  ")
            metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
            header = 'Test:'
            idx_data = 0
            img_save_count = 0
            for batch in metric_logger.log_every(test_data_loader, 10, header): 
                idx_data, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list = get_cut_data(args, model, criterion, postprocessors, metric_logger, batch, idx_data, device)
                is_save = True
                if img_save_count >= 20:
                    is_save = False
                cut_img, cut_num, cut_gt, ori_pd = save_result_image_cut(DIR_Cut_Npz_Test, DIR_Test_Cut, idx_data, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list, is_save, multislice=is_multi_slice)
                Test_Cut_Imgs.extend(cut_img)
                Test_Cut_Nums.extend(cut_num)
                Test_Cut_Gts.extend(cut_gt)
                Test_Ori_Pds.extend(ori_pd)
                img_save_count += 1
                print(f"DETR Test Progress {idx_data} / {len(test_data_loader)*args.batch_size}", end="\r")
            print("")
    return [Train_Cut_Imgs, Train_Cut_Gts, Train_Cut_Nums, Train_Ori_Pds], [Test_Cut_Imgs, Test_Cut_Gts, Test_Cut_Nums, Train_Ori_Pds]

def train_hg(args, network, opti, loss_f, train_result, test_result, device, output_dir, path_excel_result, epoch, best_dist):
    network.train()
    time_start = time.time()
    idx_data = 0
    idx_batch = 0
    train_idx_step = 0
    train_loss_epoch = 0
    train_dist_epoch = 0
    img_save_count = 0
    max_img_num = 20
    cut_inputs = []
    gt_landmarks = []
    number_list = []
    pd_landmarks = []
    for num in range(len(train_result[0])):
        cut_inputs.append(train_result[0][num])
        gt_landmarks.append(train_result[1][num])
        number_list.append(train_result[2][num])
        pd_landmarks.append(train_result[3][num])
        if len(cut_inputs) == args.batch_size or num == (len(train_result[0])-1):
            pred_tensor_cpu_list = []
            input_tensor = torch.tensor(cut_inputs, dtype=torch.float64, device=device)
            input_tensor = input_tensor.view(len(cut_inputs), 1, input_tensor.shape[1], input_tensor.shape[2]).float()
            # if args.data_5slice is not None:
            #     input_tensor, multi_gts = make_multislice_input(args.data_5slice, input_tensor.shape[0], number_list, pd_landmarks)
            #     input_tensor = input_tensor.to(device) ## torch.Size([batch, slice_num, 128, 128])
            pred_tensor = network(input_tensor) ## torch.Size([batch, slice_num * 2, 128, 128])
            # pred_tensor = network(input_tensor[:, 0:1, :, :]) 
            pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
            input_shape = input_tensor.shape[2]
            pred_shape = pred_tensor.shape[2]
            gt_hmaps = []
            for idx in range(len(input_tensor)):
                # gt_hmap = generate_hm_new_2point(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                gt_hmaps.append(gt_hmap)
            # if args.data_5slice is None:
            #     for idx in range(len(input_tensor)):
            #         # gt_hmap = generate_hm_new_2point(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
            #         gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
            #         gt_hmaps.append(gt_hmap)
            # else:
            #     for i in range(len(input_tensor)):
            #         gt_hmap_multi = []
            #         for j in range(input_tensor.shape[1]):
            #             gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[multi_gts[i][j][0][0],multi_gts[i][j][0][1]],[multi_gts[i][j][1][0], multi_gts[i][j][1][1]]]))
            #             # gt_hmap_multi.append(gt_hmap)
            #             gt_hmap_multi.extend(gt_hmap)
            #         gt_hmaps.append(gt_hmap_multi)
            opti.zero_grad()
            gt_tensor = torch.tensor(gt_hmaps, device=device).float()

            loss = loss_f(pred_tensor, gt_tensor)

            # if args.data_5slice is None:
            #     loss = loss_f(pred_tensor, gt_tensor)
            # else:
            #     # loss_1 = loss_f(pred_tensor[:, 0:2, :, :], gt_tensor[:, 0:2, :, :]) * 1.25
            #     # loss_2 = loss_f(pred_tensor[:, 2:4, :, :], gt_tensor[:, 2:4, :, :]) * 1.5
            #     # loss_3 = loss_f(pred_tensor[:, 4:6, :, :], gt_tensor[:, 4:6, :, :]) * 2.0
            #     # loss_4 = loss_f(pred_tensor[:, 6:8, :, :], gt_tensor[:, 6:8, :, :]) * 1.5
            #     # loss_5 = loss_f(pred_tensor[:, 8:10, :, :], gt_tensor[:, 8:10, :, :]) * 1.25
            #     # loss = loss_1 + loss_2 + loss_3 + loss_4 + loss_5
            #     # loss = loss_f(pred_tensor[:, 4:6, :, :], gt_tensor[:, 4:6, :, :]) 
            #     # loss = loss_f(pred_tensor[:, 2:4, :, :], gt_tensor[:, 2:4, :, :]) 
            #     # loss = loss_f(pred_tensor, gt_tensor) 
            #     loss_1 = loss_f(pred_tensor, gt_tensor[:, 0:2, :, :]) * 1.25
            #     pred_tensor = network(input_tensor[:, 1:2, :, :]) 
            #     pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
            #     loss_2 = loss_f(pred_tensor, gt_tensor[:, 2:4, :, :]) * 1.5
            #     pred_tensor = network(input_tensor[:, 2:3, :, :]) 
            #     pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
            #     loss_3 = loss_f(pred_tensor, gt_tensor[:, 4:6, :, :]) * 1.25
            #     loss = loss_1 + loss_2 + loss_3
            
            loss.backward()
            opti.step()

            idx_batch += 1
            train_idx_step += 1
            idx_data += input_tensor.shape[0]
            cur_bath_size = input_tensor.shape[0]
            train_loss_epoch += loss.item()

            print(f'HG Train Progress ------- idx_data={idx_data}/{len(train_result[0])}, curr loss={round(loss.item(),4)}', end='\r')
            input_tensor_cpu = input_tensor.cpu().numpy()
            pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
            gt_tensor_cpu = gt_tensor.detach().cpu().numpy()

            # if args.data_5slice is None:
            #     input_tensor_cpu = input_tensor.cpu().numpy()
            #     pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
            #     gt_tensor_cpu = gt_tensor.detach().cpu().numpy()
            # else:
            #     # input_tensor_cpu = input_tensor[:, 2:3, :, :].cpu().numpy()
            #     # pred_tensor_cpu = pred_tensor[:, 4:6, :, :].detach().cpu().numpy()
            #     input_tensor_cpu = input_tensor[:, 1:2, :, :].cpu().numpy()
            #     # pred_tensor_cpu = pred_tensor[:, 2:4, :, :].detach().cpu().numpy()
            #     pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
            

            multi_gts_reshape = []
            for i in range(cur_bath_size):
                gt_landmark = gt_landmarks[i]
                pred_slices = pred_tensor_cpu[i]
                landmark_num = pred_slices.shape[0]
                # if args.data_5slice is not None:
                #     multi_gts_reshape.append(multi_gts[i])
                # calculate score 
                for j in range(landmark_num):
                    pred_landmark = convert_heatmap_to_landmark(pred_slices[j, :, :])  * int(input_tensor_cpu[i][0].shape[0] / pred_slices.shape[1]) 
                    dist = np.linalg.norm(gt_landmark[j]-pred_landmark)
                    # pred_landmark = convert_heatmap_to_landmark_2output(pred_slices[j, :, :])  * int(input_tensor_cpu[i][0].shape[0] / pred_slices.shape[1])
                    # dist = 0
                    # for k in range(len(pred_landmark)):
                    #     dist += np.linalg.norm(gt_landmark[k]-pred_landmark[k])
                    # dist /= 2
                    train_dist_epoch += dist

                    if dist > 300:
                        print("")
                        print(number_list[i])
                        print(gt_landmark[j])
                        print(pred_landmark)
                        print(dist)
                        print("Dist too large value")
                        print("")
                    
            if img_save_count < max_img_num:      
                save_result_image(f"{output_dir}/IMG_Result/Train/", f"{img_save_count}_{epoch}", input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                # if args.data_5slice is None:
                #     save_result_image(f"{output_dir}/IMG_Result/Train/", f"{img_save_count}_{epoch}", input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                # else:
                #     multi_gts_reshape = np.array(multi_gts_reshape, dtype=np.int64)
                #     for k in range(3):
                #         input_tensor_cpu = input_tensor[:, k:k+1, :, :].cpu().numpy()
                #         # pred_tensor_cpu = pred_tensor[:, (k*2):(k*2)+2, :, :].detach().cpu().numpy()
                #         pred_tensor_cpu = pred_tensor_cpu_list[k]
                #         gt_landmark = multi_gts_reshape[:, k]
                #         gt_tensor_cpu = gt_tensor[:, (k*2):(k*2)+2, :, :].detach().cpu().numpy()
                #         save_result_image(f"{output_dir}/IMG_Result/Train/", f"IMG{img_save_count}_{k+1}_E{epoch}", input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                img_save_count += 1

         
            cut_inputs = []
            gt_landmarks = []
            number_list = []
            
    train_loss_epoch /= idx_batch
    train_dist_epoch /= idx_data * landmark_num

    time_end = time.time()
    time_epoch = (time_end - time_start) / 60
    time_epoch_h = int(time_epoch // 60)
    time_epoch_m = time_epoch % 60

    # train results
    print(f'HG {epoch} / {args.epochs} Epoch training finished. Loss={train_loss_epoch}, Dist={train_dist_epoch}, Time={time_epoch_h} H {round(time_epoch_m, 2)} M')                
        
    ## Test
    network.eval()
    with torch.no_grad():
        time_start = time.time()
        idx_data = 0
        idx_batch = 0
        test_idx_step = 0
        test_loss_epoch = 0
        test_dist_epoch = 0
        img_save_count = 0
        cut_inputs = []
        gt_landmarks = []
        number_list = []
        pd_landmarks = []
        for num in range(len(test_result[0])):
            cut_inputs.append(test_result[0][num])
            gt_landmarks.append(test_result[1][num])
            number_list.append(test_result[2][num])
            pd_landmarks.append(test_result[3][num])
            if len(cut_inputs) == args.batch_size or num == (len(test_result[0])-1):
                pred_tensor_cpu_list = []
                input_tensor = torch.tensor(cut_inputs, dtype=torch.float64, device=device)
                input_tensor = input_tensor.view(len(cut_inputs), 1, input_tensor.shape[1], input_tensor.shape[2]).float()
                pred_tensor = network(input_tensor)
                # if args.data_5slice is not None:
                #     input_tensor, multi_gts = make_multislice_input(args.data_5slice, input_tensor.shape[0], number_list, pd_landmarks)
                #     input_tensor = input_tensor.to(device)
                # pred_tensor = network(input_tensor[:, 0:1, :, :]) 
                pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
                input_shape = input_tensor.shape[2]
                pred_shape = pred_tensor.shape[2]
                gt_hmaps = []
                for idx in range(len(input_tensor)):
                    # gt_hmap = generate_hm_new_2point(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                    gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                    gt_hmaps.append(gt_hmap)
                # if args.data_5slice is None:
                #     for idx in range(len(input_tensor)):
                #         # gt_hmap = generate_hm_new_2point(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                #         gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[gt_landmarks[idx][0][0],gt_landmarks[idx][0][1]],[gt_landmarks[idx][1][0], gt_landmarks[idx][1][1]]]))
                #         gt_hmaps.append(gt_hmap)
                # else:
                #     for i in range(len(input_tensor)):
                #         gt_hmap_multi = []
                #         for j in range(input_tensor.shape[1]):
                #             gt_hmap = generate_hm_new(input_shape, pred_shape, pred_shape, np.array([[multi_gts[i][j][0][0],multi_gts[i][j][0][1]],[multi_gts[i][j][1][0], multi_gts[i][j][1][1]]]))
                #             # gt_hmap_multi.append(gt_hmap)
                #             gt_hmap_multi.extend(gt_hmap)
                #         gt_hmaps.append(gt_hmap_multi)
                gt_tensor = torch.tensor(gt_hmaps, device=device).float()
                
                loss = loss_f(pred_tensor, gt_tensor)
                # if args.data_5slice is None:
                #     loss = loss_f(pred_tensor, gt_tensor)
                # else:
                #     # loss_1 = loss_f(pred_tensor[:, 0:2, :, :], gt_tensor[:, 0:2, :, :]) * 1.25
                #     # loss_2 = loss_f(pred_tensor[:, 2:4, :, :], gt_tensor[:, 2:4, :, :]) * 1.5
                #     # loss_3 = loss_f(pred_tensor[:, 4:6, :, :], gt_tensor[:, 4:6, :, :]) * 2.0
                #     # loss_4 = loss_f(pred_tensor[:, 6:8, :, :], gt_tensor[:, 6:8, :, :]) * 1.5
                #     # loss_5 = loss_f(pred_tensor[:, 8:10, :, :], gt_tensor[:, 8:10, :, :]) * 1.25
                #     # loss = loss_1 + loss_2 + loss_3 + loss_4 + loss_5
                #     # loss = loss_f(pred_tensor[:, 4:6, :, :], gt_tensor[:, 4:6, :, :]) 
                #     # loss = loss_f(pred_tensor, gt_tensor) 
                #     # loss = loss_f(pred_tensor[:, 2:4, :, :], gt_tensor[:, 2:4, :, :]) 
                #     loss_1 = loss_f(pred_tensor, gt_tensor[:, 0:2, :, :]) * 1.25
                #     pred_tensor = network(input_tensor[:, 1:2, :, :]) 
                #     pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
                #     loss_2 = loss_f(pred_tensor, gt_tensor[:, 2:4, :, :]) * 1.5
                #     pred_tensor = network(input_tensor[:, 2:3, :, :]) 
                #     pred_tensor_cpu_list.append(pred_tensor.detach().cpu().numpy())
                #     loss_3 = loss_f(pred_tensor, gt_tensor[:, 4:6, :, :]) * 1.25
                #     loss = loss_1 + loss_2 + loss_3
                
                idx_batch += 1
                test_idx_step += 1
                idx_data += input_tensor.shape[0]
                cur_bath_size = input_tensor.shape[0]
                test_loss_epoch += loss.item()

                print(f'HG Test Progress ------- idx_data={idx_data}/{len(test_result[0])}, curr loss={round(loss.item(),4)}', end='\r')
                
                input_tensor_cpu = input_tensor.cpu().numpy()
                pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
                gt_tensor_cpu = gt_tensor.detach().cpu().numpy()

                # if args.data_5slice is None:
                #     input_tensor_cpu = input_tensor.cpu().numpy()
                #     pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
                #     gt_tensor_cpu = gt_tensor.detach().cpu().numpy()
                # else:
                #     # input_tensor_cpu = input_tensor[:, 2:3, :, :].cpu().numpy()
                #     input_tensor_cpu = input_tensor[:, 1:2, :, :].cpu().numpy()
                #     # pred_tensor_cpu = pred_tensor[:, 2:4, :, :].detach().cpu().numpy()
                #     pred_tensor_cpu = pred_tensor.detach().cpu().numpy()
               

                multi_gts_reshape = []
                for i in range(cur_bath_size):
                    gt_landmark = gt_landmarks[i]
                    pred_slices = pred_tensor_cpu[i]
                    landmark_num = pred_slices.shape[0]
                    # if args.data_5slice is not None:
                    #     multi_gts_reshape.append(multi_gts[i])
                    # calculate score 
                    for j in range(landmark_num):
                        pred_landmark = convert_heatmap_to_landmark(pred_slices[j, :, :])  * int(input_tensor_cpu[i][0].shape[0] / pred_slices.shape[1])
                        dist = np.linalg.norm(gt_landmark[j]-pred_landmark)
                        # pred_landmark = convert_heatmap_to_landmark_2output(pred_slices[j, :, :])  * int(input_tensor_cpu[i][0].shape[0] / pred_slices.shape[1])
                        # dist = 0
                        # for k in range(len(pred_landmark)):
                        #     dist += np.linalg.norm(gt_landmark[k]-pred_landmark[k])
                        # dist /= 2
                        test_dist_epoch += dist
                        if dist > 300:
                            print("")
                            print(number_list[i])
                            print(gt_landmark[j])
                            print(pred_landmark)
                            print(dist)
                            print("dist value too large")
                            print("")

                if img_save_count < max_img_num:      
                    save_result_image(f"{output_dir}/IMG_Result/Test/", f"{img_save_count}_{epoch}",input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                    # if args.data_5slice is None:
                    #     save_result_image(f"{output_dir}/IMG_Result/Test/", f"{img_save_count}_{epoch}",input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                    # else:
                    #     multi_gts_reshape = np.array(multi_gts_reshape, dtype=np.int64)
                    #     for k in range(3):
                    #         input_tensor_cpu = input_tensor[:, k:k+1, :, :].cpu().numpy()
                    #         # pred_tensor_cpu = pred_tensor[:, (k*2):(k*2)+2, :, :].detach().cpu().numpy()
                    #         pred_tensor_cpu = pred_tensor_cpu_list[k]
                    #         gt_landmark = multi_gts_reshape[:, k]
                    #         gt_tensor_cpu = gt_tensor[:, (k*2):(k*2)+2, :, :].detach().cpu().numpy()
                    #         save_result_image(f"{output_dir}/IMG_Result/Test/", f"IMG{img_save_count}_{k+1}_E{epoch}", input_tensor_cpu, gt_landmarks, pred_tensor_cpu, gt_tensor_cpu, number_list)
                    img_save_count += 1

                torch.cuda.empty_cache()
                cut_inputs = []
                gt_landmarks = []
                number_list = []
                
        test_loss_epoch /= idx_batch
        test_dist_epoch /= idx_data * landmark_num

        time_end = time.time()
        time_epoch = (time_end - time_start) / 60
        time_epoch_h = int(time_epoch // 60)
        time_epoch_m = time_epoch % 60

        # test results
        print(f'HG {epoch} / {args.epochs} Epoch test finished. Loss={test_loss_epoch}, Dist={test_dist_epoch}, Time={time_epoch_h} H {round(time_epoch_m, 2)} M')            

        ########## Save oeverall results ##########
        # save model     
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        new_epoch = f"{z}{epoch}"
        os.makedirs(f'{output_dir}/CheckPoint/', exist_ok=True)
        path_model = f"{output_dir}/CheckPoint/Check_epoch_{new_epoch}.pth"
        torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)

        if test_dist_epoch < best_dist:
            # best_dist = test_dist_epoch
            path_model = f"{output_dir}/CheckPoint/Check_epoch_0999.pth"
            torch.save({'epoch':epoch, 'train_idx_step':train_idx_step, 'test_idx_step':test_idx_step, 'model_state_dict':network.state_dict()}, path_model)
            print(f"Current Best model is {epoch} epoch model")

        wb = openpyxl.load_workbook(path_excel_result)
        ws = wb.active
        result_list = [epoch, train_loss_epoch, train_dist_epoch, test_loss_epoch, test_dist_epoch]
        ws.append(result_list)
        wb.save(path_excel_result)
    
    return [train_loss_epoch, test_loss_epoch, train_dist_epoch, test_dist_epoch], best_dist

@torch.no_grad()
def evaluate_add_hg(args, model, criterion, postprocessors, model_hg, data_loader, device, output_dir, output_excel_result, output_excel_dir, output_excel_slices, output_excel_box):
    model.eval()
    criterion.eval()

    DIR_OUTPUT_Val = f"{output_dir}IMG_Result/"
    os.makedirs(DIR_OUTPUT_Val, exist_ok=True)

    metric_logger = utils.MetricLogger(delimiter="  ")
    metric_logger.add_meter('class_error', utils.SmoothedValue(window_size=1, fmt='{value:.2f}'))
    header = 'Train:'

    is_multi_slice = False
    if args.data_5slice is not None:
        is_multi_slice = True

    gt_epoch = []
    pd_epoch = []
    number_epoch = []
    idx_data = 0
    img_save_count = 0
    img_max_count = 20
    pred_shape = 0
    create_box_list = []
    remain_box_list = []
    remain_score_list = []
    remain_score_weight_list = []
    select_tumor_score = []
    select_tumor_score_weight = []
    select_zero_score = []
    select_zero_score_weight = []
    select_use_score_list = []
    select_use_score_weight_list = []
    time_start = time.time()
    with torch.no_grad():
        for batch  in metric_logger.log_every(data_loader, 10, header):
            input_tensor = batch['ct'].to(device)
            number_list = batch['number']
            if args.data_5slice is not None:
                input_tensor = make_multislice_train(args.data_5slice, input_tensor.shape[0], number_list)
                input_tensor = input_tensor.to(device)    
                input_tensor_rgb = input_tensor
            
            shape_x = input_tensor.shape[2]
            shape_y = input_tensor.shape[3]
            input_tensor_cpu = input_tensor.cpu().numpy()

            if args.data_5slice is None:
                input_slices_rgb = []
                for num in range(input_tensor.shape[0]):
                    tmp_img = input_tensor_cpu[num]
                    tmp_img = tmp_img.squeeze(0)
                    # tmp_img = cv2.resize(tmp_img, (224, 224), interpolation=cv2.INTER_CUBIC)
                    tmp_img = cv2.cvtColor(tmp_img, cv2.COLOR_GRAY2BGR)
                    input_slices_rgb.append(tmp_img)
                input_tensor_rgb = torch.tensor(input_slices_rgb).to(device= device)
                input_tensor_rgb = input_tensor_rgb.transpose(1,3)
                input_tensor_rgb = input_tensor_rgb.transpose(2,3)
                input_tensor_rgb = input_tensor_rgb.to(device)

            cur_bath_size = input_tensor.shape[0]
            idx_data += cur_bath_size
            
            gt_landmarks = batch['landmarks'].to(device) 
            gt_landmarks_cpu = gt_landmarks.cpu().numpy()
            gt_landmarks_reshape = gt_landmarks.reshape(cur_bath_size, -1)
            gt_landmarks_cpu_reshape = gt_landmarks_reshape.cpu().numpy()

            targets = []
            d = {}
            labels = []
            ## label = 1 tumor, = 2 zero
            for s_num in range(len(number_list)): 
                bbox = []   
                ## ValueError: All bounding boxes should have positive height and width
                if gt_landmarks_cpu_reshape[s_num][0] == gt_landmarks_cpu_reshape[s_num][2]:
                    gt_landmarks_cpu_reshape[s_num][2] += 1
                if gt_landmarks_cpu_reshape[s_num][1] == gt_landmarks_cpu_reshape[s_num][3]:
                    gt_landmarks_cpu_reshape[s_num][3] += 1
                center_x = ((gt_landmarks_cpu_reshape[s_num][0] + gt_landmarks_cpu_reshape[s_num][2]) / 2) / shape_x
                center_y = ((gt_landmarks_cpu_reshape[s_num][1] + gt_landmarks_cpu_reshape[s_num][3]) / 2) / shape_y
                bbox_w = (gt_landmarks_cpu_reshape[s_num][2] - gt_landmarks_cpu_reshape[s_num][0]) / shape_x
                bbox_y = (gt_landmarks_cpu_reshape[s_num][3] - gt_landmarks_cpu_reshape[s_num][1]) / shape_y
                bbox = [center_x, center_y, bbox_w, bbox_y]
                if "ZERO" not in number_list[s_num]:
                    label = [1]
                else:
                    label = [args.zero_label]
                gt_bbox = torch.tensor([bbox], dtype=torch.float64, device=device)
                d['boxes'] = gt_bbox
                d['labels'] = torch.tensor(label, dtype=torch.int64, device=device) 
                targets.append(d)     
                d = {} 
            outputs , decoder_out = model(input_tensor_rgb) # ['pred_logits'] ([2, 100, 92]),['pred_boxes'] ([2, 100, 4]) / [100, 256]      
            loss_dict , indices= criterion(outputs, targets) 
            weight_dict = criterion.weight_dict

            del input_tensor, input_tensor_rgb, gt_landmarks
            """
            for q_position, label_postion in zip(indices[0][0], indices[0][1]):
                catagory_number = int(targets[0]['labels'][label_postion])
                category_query[catagory_number] += decoder_out[q_position].cpu() # class number == targets[0]['labels'][label_postion]
                class_count[catagory_number] += 1
            """
            # reduce losses over all GPUs for logging purposes
            loss_dict_reduced = utils.reduce_dict(loss_dict)
            loss_dict_reduced_scaled = {k: v * weight_dict[k]
                                        for k, v in loss_dict_reduced.items() if k in weight_dict}
            loss_dict_reduced_unscaled = {f'{k}_unscaled': v
                                        for k, v in loss_dict_reduced.items()}
            metric_logger.update(loss=sum(loss_dict_reduced_scaled.values()),
                                **loss_dict_reduced_scaled,
                                **loss_dict_reduced_unscaled)
            metric_logger.update(class_error=loss_dict_reduced['class_error'])

            # orig_target_sizes = torch.stack([t["orig_size"] for t in targets], dim=0)
            t_sizes = []
            for t_num in range(cur_bath_size):
                t_sizes.append([512, 512])
            
            orig_target_sizes = torch.tensor(t_sizes, dtype=torch.int64, device=device)
            
            results = postprocessors['bbox'](outputs, orig_target_sizes)
            cut_landmarks = []
            pd_list_select = []
            box_score_choice = []
            for bnum in range(cur_bath_size):
                gt_landmark = gt_landmarks_cpu[bnum]
                gt_land = np.array(gt_landmark, dtype=np.int64)
                number_slice = number_list[bnum]
                boxes = results[bnum]['boxes']
                scores = results[bnum]['scores']
                boxes_cpu_ori = boxes.cpu().numpy().astype(np.int64)
                scores_cpu_ori = scores.cpu().numpy()
                boxes_cpu, scores_cpu = arrage_result(boxes_cpu_ori, scores_cpu_ori)

                min_x = 512
                min_y = 512
                max_x = 0
                max_y = 0
                land_margin_min = 0 
                land_margin_max = 0
                if args.box_mode == "SB":
                    if scores_cpu[0] >= args.box_th:
                        min_x = boxes_cpu[0][0]
                        min_y = boxes_cpu[0][1]
                        max_x = boxes_cpu[0][2]
                        max_y = boxes_cpu[0][3]
                        if min_x == max_x:
                            max_x += 1
                        if max_y == min_y:
                            max_y += 1
                        if min_x > 50:
                            land_margin_min = 20
                        if max_x > 80:
                            land_margin_max = 20
                    else:
                        min_x = 0
                        min_y = 0
                        max_x = 1
                        max_y = 1
                else:
                    zero_img, create_box_points, select_boxes = create_box(boxes_cpu, scores_cpu, args.box_th, args.box_usenum, shape_x, shape_y)                    
                    pd_list_select.append(select_boxes)
                    box_score = np.max(zero_img)
                    
                    if box_score > args.box_th:
                        box_score_choice.append(round(box_score,3))
                        min_x = create_box_points[0]
                        min_y = create_box_points[1]
                        max_x = create_box_points[2]
                        max_y = create_box_points[3]
                        # for x in range(create_box_points[0], create_box_points[2] + 1):
                        #     for y in range(create_box_points[1], create_box_points[3] + 1):
                        #         if zero_img[x][y] != 0.0:
                        #             if x <= min_x:
                        #                 min_x = x
                        #             if y <= min_y:
                        #                 min_y = y
                        #             if x >= max_x:
                        #                 max_x = x
                        #             if y >= max_y:
                        #                 max_y = y
                    else:
                        box_score_choice.append("BG")
                        min_x = 0
                        min_y = 0
                        max_x = 1
                        max_y = 1  
                    create_box_points_final = [min_x, min_y, max_x, max_y]
                    create_box_list.append(np.array(create_box_points_final).reshape(2,2))
                    box_list, score_list, score_weight_list, select_score_list, select_score_weihgt_list = calculate_box_score(args.box_usenum, create_box_points_final, boxes_cpu, scores_cpu)
                    remain_box_list.append(box_list)
                    remain_score_list.append(score_list)
                    remain_score_weight_list.append(score_weight_list)
                    if len(select_score_list) > 0:
                        select_use_score_list.extend(select_score_list)
                        select_use_score_weight_list.extend(select_score_weihgt_list)

                    if min_x > 50:
                        land_margin_min = 20
                    if max_x > 80:
                        land_margin_max = 20

                    if min_x == max_x:
                        max_x += 1
                    if max_y == min_y:
                        max_y += 1
                    
                cut_landmarks.append([int(min_x)-land_margin_min, int(min_y)-land_margin_min, int(max_x)+land_margin_max, int(max_y)+land_margin_max])
                
                gt_epoch.append(gt_land)
                number_epoch.append(number_slice)
                if 'ZERO' in number_slice:
                    select_zero_score.extend(select_score_list)
                    select_zero_score_weight.extend(select_score_weihgt_list)            
                else:     
                    select_tumor_score.extend(select_score_list)
                    select_tumor_score_weight.extend(select_score_weihgt_list)
            is_save = False
            is_tumor = False
            for name in number_list:
                if number_list.index(name) < 8:
                    if img_save_count >= 17:
                        is_tumor = True
                    if "ZERO" not in name:
                        is_tumor = True
            if img_save_count < img_max_count and is_tumor:
                is_save = True
                img_save_count += 1       
            
            if idx_data == args.batch_size:
                ori_img = input_tensor_cpu[0][0]
                cut = cut_landmarks[0] 
                cut_img = ori_img[cut[1]:cut[1]+(cut[3]-cut[1]), cut[0]:cut[0]+(cut[2]-cut[0])]
                cut_img = cv2.resize(cut_img, dsize=(128,128), interpolation=cv2.INTER_AREA)
                tmp_tensor = torch.tensor(cut_img, dtype=torch.float64, device=device)
                tmp_tensor = tmp_tensor.view(1, 1, tmp_tensor.shape[0], tmp_tensor.shape[1]).float()
                pred_tensor = model_hg(tmp_tensor)
                pred_shape = pred_tensor.shape[2]
                del tmp_tensor
            
            pred_land = detradd_net(DIR_OUTPUT_Val, idx_data, model_hg, device, input_tensor_cpu, gt_landmarks_cpu, results, 
                                        pd_list_select, box_score_choice, cut_landmarks, number_list, pred_shape, is_save, args.is_o2, is_multi_slice)  
            pd_epoch.extend(pred_land)
            
            torch.cuda.empty_cache()
            
            print(f"Network Eval Progress ------ {idx_data} / {len(data_loader)*args.batch_size} ", end='\r')
    time_net_end = time.time()
    time_net = (time_net_end - time_start) / 60
    time_net_H = int(time_net // 60)
    time_net_M = time_net % 60
    print(f"Network Progress Finish Time = About {time_net_H} h {round(time_net_M,2)} m, Calculate Score start")

    ## score
    iou_epoch = []
    dice_epoch = []
    specificity_epoch = []
    sensitivity_epoch = []
    precision_epoch = []
    dist_epoch = []
    patient_epoch = []
    slice_epoch = []
    iszero_epoch = []
    tumor_result = []
    zero_result = []
    tumor_num = 0
    zero_num = 0
    correct_tumor_num = 0
    correct_zero_num_1 = 0
    correct_zero_num_2 = 0
    
    for idx in range(len(gt_epoch)):
        print(f"Calculate Score {idx+1} / {len(gt_epoch)}", end="\r")
        TP, FP, TN, FN = cal_score(gt_epoch[idx], pd_epoch[idx], [shape_x, shape_y])
        
        number_slice = number_epoch[idx]
        
        # gt_img, pred_img = fill_box(gt_land, pred_land, [input_slice.shape[0], input_slice.shape[1]])
        iou_value, dice_value, specitificity, sensitivity, precision = Cal_Result(TP, FP, TN, FN)
        iou_epoch.append(iou_value)
        dice_epoch.append(dice_value)
        specificity_epoch.append(specitificity)
        sensitivity_epoch.append(sensitivity)
        precision_epoch.append(precision)
        dist = 0
        for land_mark in range(2):
            dist += np.linalg.norm(gt_epoch[idx][land_mark]-pd_epoch[idx][land_mark])
        dist /= 2
        dist_epoch.append(dist)
        
        is_zero = 'X'
        if 'ZERO' in number_slice:
            is_zero = "O"
            zero_result.append([iou_value, dice_value, precision, sensitivity, specitificity, dist])
            # select_zero_score.extend(select_score_list)
            # select_zero_score_weight.extend(select_score_weihgt_list)
            zero_num += 1
            if int(pd_epoch[idx][0][0]) == 0 and int(pd_epoch[idx][0][1]) == 0:
                correct_zero_num_1 += 1
                if int(pd_epoch[idx][1][1]) == 0:
                    correct_zero_num_2 += 1
        else:
            tumor_result.append([iou_value, dice_value, precision, sensitivity, specitificity, dist])
            # select_tumor_score.extend(select_score_list)
            # select_tumor_score_weight.extend(select_score_weihgt_list)
            tumor_num += 1
            if int(pd_epoch[idx][1][1]) != 0:
                correct_tumor_num += 1

        patient = number_slice.split("_")[0] + "_" + number_slice.split("_")[1]
        slice_name_list = number_slice.split("_")[2:]
        slice_name = ''
        for s in slice_name_list:
            if s == "ZERO":
                slice_name += '_' + s
            else:
                slice_name += s
        patient_epoch.append(patient)
        slice_epoch.append(slice_name)
        iszero_epoch.append(is_zero)
    acc_result = [correct_tumor_num, correct_zero_num_1, correct_zero_num_2, round(correct_tumor_num/tumor_num, 4), round(correct_zero_num_1/zero_num, 4), round(correct_zero_num_2/zero_num, 4)]
    remain_result = [remain_box_list, remain_score_list, remain_score_weight_list, 
                    select_tumor_score, select_tumor_score_weight, select_zero_score, select_zero_score_weight, select_use_score_list, select_use_score_weight_list]

    ## Init excel 
    wb_result = openpyxl.load_workbook(output_excel_result)
    ws = wb_result.active
    reulst_list = [f"All", round(np.mean(iou_epoch), 4), round(np.mean(dice_epoch), 4), round(np.mean(precision_epoch), 4), round(np.mean(sensitivity_epoch), 4), round(np.mean(specificity_epoch), 4), 
                    round(np.mean(dist_epoch), 4), acc_result[3], acc_result[4], acc_result[5], round(np.mean(remain_result[7]), 4), round(np.mean(remain_result[8]), 4)]
    ws.append(reulst_list)
    iou = []
    dice = []
    precision = []
    sensitiy = []
    speci = []
    dist = []
    for i in range(len(tumor_result)):
        iou.append(tumor_result[i][0])
        dice.append(tumor_result[i][1])
        precision.append(tumor_result[i][2])
        sensitiy.append(tumor_result[i][3])
        speci.append(tumor_result[i][4])
        dist.append(tumor_result[i][5])
    reulst_list = ["Tumor", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(precision), 4), round(np.mean(sensitiy), 4), round(np.mean(speci), 4), round(np.mean(dist), 4), 
                    acc_result[0], "", "", round(np.mean(remain_result[3]), 4), round(np.mean(remain_result[4]), 4)]
    ws.append(reulst_list)
    iou = []
    dice = []
    precision = []
    sensitiy = []
    speci = []
    dist = []
    for i in range(len(zero_result)):
        iou.append(zero_result[i][0])
        dice.append(zero_result[i][1])
        precision.append(zero_result[i][2])
        sensitiy.append(zero_result[i][3])
        speci.append(zero_result[i][4])
        dist.append(zero_result[i][5])
    reulst_list = ["Zero", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(precision), 4), round(np.mean(sensitiy), 4), round(np.mean(speci), 4), round(np.mean(dist), 4), 
                    "", acc_result[1], acc_result[2], round(np.mean(remain_result[5]), 4), round(np.mean(remain_result[6]), 4)]
    ws.append(reulst_list)
    wb_result.save(output_excel_result)

    wb_eval = openpyxl.load_workbook(output_excel_slices)
    ws = wb_eval.active
    for s in range(len(dice_epoch)):
        print(f"Write Slice Excel {s+1} / {len(dice_epoch)}", end="\r")
        reulst_list = [patient_epoch[s], slice_epoch[s], iszero_epoch[s], str(gt_epoch[s][0])[1:-1], str(pd_epoch[s][0])[1:-1], str(gt_epoch[s][1])[1:-1], str(pd_epoch[s][1])[1:-1], 
                        np.round(iou_epoch[s],4), np.round(dice_epoch[s],4), np.round(precision_epoch[s],4), np.round(sensitivity_epoch[s],4), np.round(specificity_epoch[s],4), np.round(dist_epoch[s],4)]
        ws.append(reulst_list)
    wb_eval.save(output_excel_slices)

    wb_box = openpyxl.load_workbook(output_excel_box)
    ws = wb_box.active
    for s in range(len(dice_epoch)):
        print(f"Write Slice Box info Excel {s+1} / {len(dice_epoch)}", end="\r")
        reulst_list = [patient_epoch[s], slice_epoch[s], iszero_epoch[s], str(gt_epoch[s][0])[1:-1], str(pd_epoch[s][0])[1:-1], str(gt_epoch[s][1])[1:-1], str(pd_epoch[s][1])[1:-1], 
                        str(create_box_list[s][0])[1:-1], str(create_box_list[s][1])[1:-1],
                        str(remain_result[0][s][0][0])[1:-1], str(remain_result[0][s][0][1])[1:-1], round(remain_result[1][s][0],4), round(remain_result[2][s][0],4),
                        str(remain_result[0][s][1][0])[1:-1], str(remain_result[0][s][1][1])[1:-1], round(remain_result[1][s][1],4), round(remain_result[2][s][1],4),
                        str(remain_result[0][s][2][0])[1:-1], str(remain_result[0][s][2][1])[1:-1], round(remain_result[1][s][2],4), round(remain_result[2][s][2],4),
                        str(remain_result[0][s][3][0])[1:-1], str(remain_result[0][s][3][1])[1:-1], round(remain_result[1][s][3],4), round(remain_result[2][s][3],4)]
        ws.append(reulst_list)
    wb_box.save(output_excel_box)

    time_end = time.time()
    time_total = (time_end - time_start) / 60
    time_total_H = int(time_total // 60)
    time_total_M = time_total % 60
    print(f"Validation Finish Time = About {time_total_H} h {round(time_total_M,2)} m", end='\r')
    
    patient_score(output_excel_dir)
    patient_score_boxinfo(output_excel_dir)
    # dt_dir = f"{args.model_path}Checkpoint/"  
    print(f"Used device = {device} - {torch.cuda.get_device_name(device)}")

def get_cut_data(args, model, criterion, postprocessors, metric_logger, batch, idx_data, device, mode='train'):
    input_tensor = batch['ct'].to(device)
    number_list = batch['number']

    if args.data_5slice is not None:
        input_tensor = make_multislice_train(args.data_5slice, input_tensor.shape[0], number_list)
        input_tensor = input_tensor.to(device)    
        input_tensor_rgb = input_tensor
    shape_x = input_tensor.shape[2]
    shape_y = input_tensor.shape[3]
    input_tensor_cpu = input_tensor.cpu().numpy()
    if args.data_5slice is None:
        input_slices_rgb = []
        for num in range(input_tensor.shape[0]):
            tmp_img = input_tensor_cpu[num]
            tmp_img = tmp_img.squeeze(0)
            # tmp_img = cv2.resize(tmp_img, (224, 224), interpolation=cv2.INTER_CUBIC)
            tmp_img = cv2.cvtColor(tmp_img, cv2.COLOR_GRAY2BGR)
            input_slices_rgb.append(tmp_img)
        input_tensor_rgb = torch.tensor(input_slices_rgb).to(device= device)
        input_tensor_rgb = input_tensor_rgb.transpose(1,3)
        input_tensor_rgb = input_tensor_rgb.transpose(2,3)
        input_tensor_rgb = input_tensor_rgb.to(device)

    cur_bath_size = input_tensor.shape[0]
    idx_data += cur_bath_size
    gt_landmarks = batch['landmarks'].to(device) 
    gt_landmarks_cpu = gt_landmarks.cpu().numpy()
    
    gt_landmarks_reshape = gt_landmarks.reshape(cur_bath_size, -1)
    gt_landmarks_cpu_reshape = gt_landmarks_reshape.cpu().numpy()

    targets = []
    d = {}
    labels = []
    ## label = 1 tumor, = 2 zero
    for s_num in range(len(number_list)): 
        bbox = []   
        ## ValueError: All bounding boxes should have positive height and width
        if gt_landmarks_cpu_reshape[s_num][0] == gt_landmarks_cpu_reshape[s_num][2]:
            gt_landmarks_cpu_reshape[s_num][2] += 1
        if gt_landmarks_cpu_reshape[s_num][1] == gt_landmarks_cpu_reshape[s_num][3]:
            gt_landmarks_cpu_reshape[s_num][3] += 1
        center_x = ((gt_landmarks_cpu_reshape[s_num][0] + gt_landmarks_cpu_reshape[s_num][2]) / 2) / shape_x
        center_y = ((gt_landmarks_cpu_reshape[s_num][1] + gt_landmarks_cpu_reshape[s_num][3]) / 2) / shape_y
        bbox_w = (gt_landmarks_cpu_reshape[s_num][2] - gt_landmarks_cpu_reshape[s_num][0]) / shape_x
        bbox_y = (gt_landmarks_cpu_reshape[s_num][3] - gt_landmarks_cpu_reshape[s_num][1]) / shape_y
        bbox = [center_x, center_y, bbox_w, bbox_y]
        if "ZERO" not in number_list[s_num]:
            label = [1]
        else:
            label = [args.zero_label]
        gt_bbox = torch.tensor([bbox], dtype=torch.float64, device=device)
        d['boxes'] = gt_bbox
        d['labels'] = torch.tensor(label, dtype=torch.int64, device=device) 
        targets.append(d)     
        d = {} 
    outputs , decoder_out = model(input_tensor_rgb) # ['pred_logits'] ([2, 100, 92]),['pred_boxes'] ([2, 100, 4]) / [100, 256]      
    loss_dict , indices = criterion(outputs, targets) 
    weight_dict = criterion.weight_dict

    del input_tensor, input_tensor_rgb, gt_landmarks
    """
    for q_position, label_postion in zip(indices[0][0], indices[0][1]):
        catagory_number = int(targets[0]['labels'][label_postion])
        category_query[catagory_number] += decoder_out[q_position].cpu() # class number == targets[0]['labels'][label_postion]
        class_count[catagory_number] += 1
    """
    # reduce losses over all GPUs for logging purposes
    loss_dict_reduced = utils.reduce_dict(loss_dict)
    loss_dict_reduced_scaled = {k: v * weight_dict[k]
                                for k, v in loss_dict_reduced.items() if k in weight_dict}
    loss_dict_reduced_unscaled = {f'{k}_unscaled': v
                                for k, v in loss_dict_reduced.items()}
    metric_logger.update(loss=sum(loss_dict_reduced_scaled.values()),
                        **loss_dict_reduced_scaled,
                        **loss_dict_reduced_unscaled)
    metric_logger.update(class_error=loss_dict_reduced['class_error'])

    # orig_target_sizes = torch.stack([t["orig_size"] for t in targets], dim=0)
    t_sizes = []
    for t_num in range(cur_bath_size):
        t_sizes.append([512, 512])
    
    orig_target_sizes = torch.tensor(t_sizes, dtype=torch.int64, device=device)
    
    results = postprocessors['bbox'](outputs, orig_target_sizes)
    cut_landmarks = []
    pd_list_select = []
    output_score_list = []
    for bnum in range(cur_bath_size):
        boxes = results[bnum]['boxes']
        scores = results[bnum]['scores']
        boxes_cpu_ori = boxes.cpu().numpy().astype(np.int64)
        scores_cpu_ori = scores.cpu().numpy()
        boxes_cpu, scores_cpu = arrage_result(boxes_cpu_ori, scores_cpu_ori)
        
        min_x = 512
        min_y = 512
        max_x = 0
        max_y = 0
        land_margin_min = 0 
        land_margin_max = 0
        if args.box_mode == "SB":
            if scores_cpu[0] >= args.box_th:
                min_x = boxes_cpu[0][0]
                min_y = boxes_cpu[0][1]
                max_x = boxes_cpu[0][2]
                max_y = boxes_cpu[0][3]
                if min_x == max_x:
                    max_x += 1
                if max_y == min_y:
                    max_y += 1
                if min_x > 50:
                    land_margin_min = 20
                if max_x > 80:
                    land_margin_max = 20
            else:
                min_x = 0
                min_y = 0
                max_x = 1
                max_y = 1
        else:
            zero_img, create_box_points, select_boxes = create_box(boxes_cpu, scores_cpu, args.box_th, args.box_usenum, shape_x, shape_y)
            
            pd_list_select.append(select_boxes)
            box_score = np.max(zero_img)
            output_score_list.append(round(box_score,3))
            min_x = 512
            min_y = 512
            max_x = 0
            max_y = 0

            if box_score > args.box_th:
                for x in range(create_box_points[0], create_box_points[2] + 1):
                    for y in range(create_box_points[1], create_box_points[3] + 1):
                        if zero_img[x][y] != 0.0:
                            if x <= min_x:
                                min_x = x
                            if y <= min_y:
                                min_y = y
                            if x >= max_x:
                                max_x = x
                            if y >= max_y:
                                max_y = y

            ## if detr output = 0 
            if min_x == 512:
                min_x = 0
                min_y = 0
                max_x = 1
                max_y = 1
            if min_x == max_x:
                max_x += 1
            if max_y == min_y:
                max_y += 1
            
            gt = gt_landmarks_cpu[bnum]
            ## if detr output = tumor and gt = tumor, but gt > output 
            if min_x > 50 and gt[0][0] > 50:
                if min_x >= gt[0][0]:
                    min_x = gt[0][0]
                if min_y >= gt[0][1]:
                    min_y = gt[0][1]
                if max_x <= gt[1][0]:
                    max_x = gt[1][0]
                if max_y <= gt[1][1]:
                    max_y = gt[1][1]
                land_margin_min = 20
                land_margin_max = 20
            ## if detr output = 0 but gt = tumor 
            elif min_x < 50 and gt[0][0] > 50:
                min_x = gt[0][0]
                min_y = gt[0][1]
                max_x = gt[1][0]
                max_y = gt[1][1]
                land_margin_min = 20
                land_margin_max = 20
            ## detr output = tumor and gt = 0, detr output = 0 and gt = 0 don't need to revise

            if min_x < 50 and max_x > 80:
                land_margin_min = 0
                land_margin_max = 20
        cut_landmarks.append([int(min_x)-land_margin_min, int(min_y)-land_margin_min, int(max_x)+land_margin_max, int(max_y)+land_margin_max])

    return idx_data, input_tensor_cpu, gt_landmarks_cpu, cut_landmarks, number_list

