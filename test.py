# Copyright (c) Facebook, Inc. and its affiliates. All Rights Reserved
"""
,
            "args" : ["--batch_size", "2", 
                "--no_aux_loss", 
                "--eval", 
                "--resume", "checkpoints/detr-r50-e632da11.pth", 
                "--num_workers", "4",
                "--world_size", "2",
                "--coco_path", "/dataset/coco",
                "--output_dir", "result"]
"""
# 아래 코드 주석에서 ## 으로 적어 놓은 주석은, 내가 적은 주석. 영어 말고 최대한 한글로 적어 놓기. 미래의 나를 위해서
import argparse
import datetime
import json
import random
import time
from pathlib import Path

import numpy as np
import torch
from torch.utils.data import DataLoader, DistributedSampler

# import datasets
import util.misc as utils
from validation import val, val_max_colon, val_s2OnlyT, val_s2OnlyT_NSq, val_contour

from models import build_model
import torch.optim as optim 
import torch.nn as nn
from via import via

import datetime
import random
import natsort

from BoundingBoxDataSet import BoundingBoxDataSet
import matplotlib.pyplot as plt
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
import openpyxl
from calculate import *
import learning_s1_colon

def get_args_parser():
    GPU_Name = "cuda:1"
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    try:
        gpu_info = torch.cuda.get_device_name(device)
    except:
        GPU_Name = "cuda:0"
        device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
        gpu_info = torch.cuda.get_device_name(device)
    
    Computer = ""
    Batch_Size = 32
    Nas_Mount = "Y:/" 
    Data_Mount = "D:/" 
    DIR_DATA_VAL = f"{Data_Mount}yskim/BoundingBox/data/processed/Validation_all/"
    DIR_DATA_VAL = f"{Data_Mount}yskim/BoundingBox/data/processed/Validation_all_G/"
    
    if "3060" in gpu_info:
        Computer = "Demo"
        Batch_Size = int(Batch_Size/2)
        Demo = "_tmp"
        Data_Mount = "Y:/" 
        DIR_DATA_VAL = f"{Data_Mount}yskim/BoundingBox/data/processed/Validation/tmp_val/"
        DIR_DATA_VAL = f"{Data_Mount}yskim/BoundingBox/data/processed/Validation_all_G_tmp/"
        # DIR_DATA_VAL = f"{Data_Mount}yskim/BoundingBox/data/processed/Validation_all_G/"
    elif "3090" in gpu_info:
        Computer = "Com"
    else:
        print(f"Check this PC GPU = {gpu_info}")
        exit()
    Model_Path = f"{Nas_Mount}yskim/BoundingBox/result/DT1014_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_Com/CheckPoint/"
    Model_Epoch_List = ['0022', '0024', '0026', '0028', '0030']
    # Model_Epoch_List = ['0027', '0028']
    Model_Epoch_List = ['0022']
   
    model_info = Model_Path.split('/')[-3].split('_')
    Task = ''
    task_idx = 0
    for info in model_info:
        if 'C2' in info:
            Class_Num = 2 ## 2= tumor+bg 3= tumor+zero+bg
            Zero_Label = 0
        if 'C3' in info:
            Class_Num = 2
            Zero_Label = 1
        if 'Q' in info:
            Query_Num = int(info[1:])
        if 'UB' in info:
            Box_UseNum = int(info[2:])
            task_idx = model_info.index(info)
        if 'SB' in info:
            Box_Mode = "SB"
            Box_TH = float(f"{info[2]}.{info[3]}")
        if 'NB' in info:
            Box_Mode = "NB"
            Box_TH = float(f"{info[2]}.{info[3]}")
        if 'DT' in info:
            Data_Year = info[2:]
        if "TNSame" in info:
            Data_Year += f"_{info}"
        if (task_idx > 0 and model_info.index(info) > task_idx) and info != model_info[-1]:
            Task += f'_{info}'
    th = str(Box_TH).split('.')
   
    Pretrain_model = "detr-r101-dc5.pth"
    # Pretrain_model = None
    Back_bone = 'resnet101'
    # Back_bone = 'resnet50'
    # DIR_RocoResnet = f"{Nas_Mount}yskim/BoundingBox/result/Roco_Resnet101_Pretrained_Com_3090/CheckPoint/net_93.pt"
    DIR_RocoResnet = f"{Nas_Mount}yskim/BoundingBox/result/Roco_Resnet101_DT_3CH_Com_3090/CheckPoint/net_93.pt"

    Name_Result = ''
    if "Validation_all_G" in DIR_DATA_VAL:
        Name_Result = "Val_G_"
        DIR_DATA_SEG = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/gang_fast_colon/" ## sin_fast_colon gang_fast_colon sin_fast_colon_tumor
        DIR_DATA_SEG_Contour = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/gang_Fast_colon_contour/"
    else: 
        Name_Result = "Val_"
        DIR_DATA_SEG = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/sin_fast_colon/" ## sin_fast_colon gang_fast_colon sin_fast_colon_tumor
        DIR_DATA_SEG_Contour = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/sin_Fast_colon_contour/"
    Name_Result += f"DT{Data_Year}_Q{Query_Num}_C{Class_Num}_BR101_R101_{Box_Mode}{th[0]}{th[1]}_UB{Box_UseNum}{Task}"    
    

    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))
    DIR_RESULT = f"{Nas_Mount}yskim/BoundingBox/result/{today}_{now}_{Name_Result}_{Computer}/"                


    parser = argparse.ArgumentParser('Set transformer detector', add_help=False)
    parser.add_argument('--lr', default=1e-4, type=float)
    parser.add_argument('--lr_backbone', default=1e-5, type=float)
    parser.add_argument('--batch_size', default=Batch_Size, type=int)
    parser.add_argument('--weight_decay', default=1e-4, type=float)
    parser.add_argument('--lr_drop', default=200, type=int)
    parser.add_argument('--clip_max_norm', default=0.1, type=float, help='gradient clipping max norm')

    # Model parameters
    parser.add_argument('--frozen_weights', type=str, default=None, help="Path to the pretrained model. If set, only the mask head will be trained")
    # * Backbone
    parser.add_argument('--backbone', default=Back_bone, type=str, help="Name of the convolutional backbone to use")
    parser.add_argument('--dilation', action='store_false',
                        help="If true, we replace stride with dilation in the last convolutional block (DC5)") # stride를 적용하지 않고, dilation을 추가해서 backbone output resolution을 높힌다.
    parser.add_argument('--position_embedding', default='sine', type=str, choices=('sine', 'learned'), help="Type of positional embedding to use on top of the image features")
    parser.add_argument('--pretraind_model', default=Pretrain_model, type=str, help="Name of the pretrain model")

    # * Transformer
    parser.add_argument('--enc_layers', default=6, type=int,
                        help="Number of encoding layers in the transformer")
    parser.add_argument('--dec_layers', default=6, type=int,
                        help="Number of decoding layers in the transformer")
    parser.add_argument('--dim_feedforward', default=2048, type=int,
                        help="Intermediate size of the feedforward layers in the transformer blocks")
    parser.add_argument('--hidden_dim', default=256, type=int,
                        help="Size of the embeddings (dimension of the transformer)")
    parser.add_argument('--dropout', default=0.1, type=float,
                        help="Dropout applied in the transformer")
    parser.add_argument('--nheads', default=8, type=int,
                        help="Number of attention heads inside the transformer's attentions")
    parser.add_argument('--num_queries', default=Query_Num, type=int, help="Number of query slots")
    parser.add_argument('--num_classes', default=Class_Num, type=int, help="Number of class")
    parser.add_argument('--zero_label', default=Zero_Label, type=int, help="zero label 0 or 2")
    parser.add_argument('--pre_norm', action='store_true')

    # * Segmentation
    parser.add_argument('--masks', action='store_true',
                        help="Train segmentation head if the flag is provided")

    # Loss
    parser.add_argument('--no_aux_loss', dest='aux_loss', action='store_false',
                        help="Disables auxiliary decoding losses (loss at each layer)")  # args.aux_loss == false
                        
    # * Matcher
    parser.add_argument('--set_cost_class', default=1, type=float, help="Class coefficient in the matching cost")
    parser.add_argument('--set_cost_bbox', default=5, type=float, help="L1 box coefficient in the matching cost")
    parser.add_argument('--set_cost_giou', default=2, type=float, help="giou box coefficient in the matching cost")

    # * Loss coefficients
    parser.add_argument('--mask_loss_coef', default=1, type=float)
    parser.add_argument('--dice_loss_coef', default=1, type=float)
    parser.add_argument('--bbox_loss_coef', default=5, type=float)
    parser.add_argument('--giou_loss_coef', default=2, type=float)
    parser.add_argument('--eos_coef', default=0.1, type=float, help="Relative classification weight of the no-object class")

    # dataset parameters
    parser.add_argument('--dataset_file', default='coco')
    parser.add_argument('--coco_path', type=str)
    parser.add_argument('--coco_panoptic_path', type=str)
    parser.add_argument('--remove_difficult', action='store_true')

    parser.add_argument('--output_dir', default=DIR_RESULT, help='path where to save, empty for no saving')
    parser.add_argument('--output_name', default=Name_Result, help='path where to save name')
    parser.add_argument('--device', default=GPU_Name, help='device to use for training / testing')
    parser.add_argument('--data_val', default=DIR_DATA_VAL, help="train data dir")
    parser.add_argument('--data_seg', default=DIR_DATA_SEG, help="test data dir")
    parser.add_argument('--data_seg_contour', default=DIR_DATA_SEG_Contour, help="contour data dir")
    parser.add_argument('--seed', default=42, type=int)
    parser.add_argument('--computer', default=Computer, type=str, help='this computer name')
    parser.add_argument('--model_path', default=f'{Model_Path}', help='path where to train model')
    parser.add_argument('--model_epoch_list', default=Model_Epoch_List, help="val model epoch") 
    parser.add_argument('--num_workers', default=0, type=int) # 

    # distributed training parameters
    parser.add_argument('--world_size', default=1, type=int, help='number of distributed processes')
    parser.add_argument('--dist_url', default='env://', help='url used to set up distributed training')

    ## parameters
    parser.add_argument('--box_th', default=Box_TH, type=float, help='box threshold')
    parser.add_argument('--box_th_s2', default=(Box_TH/2.0), type=float, help='box threshold')
    parser.add_argument('--box_mode', default=Box_Mode, type=str, help='box mode')
    parser.add_argument('--box_usenum', default=Box_UseNum, type=int, help='box usenum')
    parser.add_argument('--nas_mount', default=Nas_Mount, type=str, help='nas mount path')
    parser.add_argument('--rocoresnet_dir', default=DIR_RocoResnet, help='path where to roco resnet save')
    return parser


def main(args):
    utils.init_distributed_mode(args) # Multi-GPU 사용할 거라면, args.gpu / args.world_size / args.rank 가 여기서 정의 된다.
    # print("git:\n  {}\n".format(utils.get_sha()))

    if args.frozen_weights is not None:
        assert args.masks, "Frozen training is meant for segmentation only"
    # print(args)
    
    
    print("colon data loding ....", end='\r')
    tumor_num = 0
    tumor_list = []
    val_data_dir = "Y:/yskim/BoundingBox/data/processed/Validation_all_G/"
    val_data_dir = "Y:/yskim/BoundingBox/data/processed/Validation_all/"
    with open(f"{args.data_seg[:-1]}.txt", 'r') as txt_colon:
        lines = txt_colon.readlines()
        last = lines[-1]
        colon_info = {}
        slice_info = {}
        patient_cur = ''
        is_first = True
        for line in lines:
            patient = line.split(',')[0]
            
            if patient_cur != patient and is_first:
                patient_cur = patient
                is_first = False
            elif patient_cur != patient and not is_first:
                colon_info[patient_cur] = slice_info
                patient_cur = patient
                slice_info = {}
            s_idx = line.split(',')[1][1:]
            min_x = line.split(',')[2][1:]
            min_y = line.split(',')[3][1:]
            max_x = line.split(',')[4][1:]
            max_y = line.split(',')[5][:-2]
            slice_info[s_idx] = [min_x, min_y, max_x, max_y]
            # if int(max_y) < 50:
            #     str_num = len(str(s_idx))
            #     zero_need = 4 - str_num
            #     new_sidx = ''
            #     for i in range(zero_need):
            #         new_sidx += '0'
                
            #     ct_name = f"{patient}_{new_sidx}{s_idx}_ZERO.npz"
            #     if os.path.isfile(f"{val_data_dir}{ct_name}"):
            #         pass
            #     else:
            #         tumor_num += 1
            #         tumor_list.append(f"{patient}_{new_sidx}{s_idx}")
            if line is last:
                colon_info[patient_cur] = slice_info
    print("colon data loding finish")


    device = torch.device(args.device)
   
    # Multi-GPU 사용할 거라면, fix the seed for reproducibility 
    seed = args.seed + utils.get_rank() 
    torch.manual_seed(seed)
    np.random.seed(seed)
    random.seed(seed)

    ## 전까지는, args 설정 및 seed 설정이 전부이다.
    ## 여기가 진짜 시작 
    model, criter, ppro = build_model(args)
    model_2, criter_2, ppro_2 = build_model(args)
    ## The models are also available via torch hub, to load DETR R50 with pretrained weights simply do:
    if args.pretraind_model is not None:
        checkpoint = torch.load(f"{args.nas_mount}/yskim/BoundingBox/code/DETR/{args.pretraind_model}", map_location=device)
        del checkpoint["model"]["class_embed.weight"]
        del checkpoint["model"]["class_embed.bias"]
        del checkpoint["model"]["query_embed.weight"]
        model.load_state_dict(checkpoint['model'], strict=False)
        model_2.load_state_dict(checkpoint['model'], strict=False)
        print("Load Pretrained COCO Data model ")
    
    model.to(device)
    model_2.to(device)

    ####
    # dtpt = torch.load(args.rocoresnet_dir , map_location=device)
    # list_check = list(dtpt['model_state_dict'].items())
    # del_list = []
    # list_check_body = []
    # for i in range(len(list_check)):
    #     if "fc." not in list_check[i][0]:
    #         # del_list.append(f"0.body.{list_check[i]}") 
    #         # del_list.append(f"{list_check[i]}")
    #         list_check_body.append((f"0.body.{list_check[i][0]}",list_check[i][1])) 
    #     # else:
    #     #     list_check_body.append((f"0.body.{list_check[i][0]}",list_check[i][1]))
    # load_check = list_check_body
    # load_check = dict(load_check)
    # model.backbone.load_state_dict(load_check)
    # print("ROCO Resnet load")

    model_without_ddp = model
    # if args.distributed:
    #     model = torch.nn.parallel.DistributedDataParallel(model, device_ids=[args.gpu])
    #     model_without_ddp = model.module
    
    ## 총 파라메터 갯수 계산 하는 방법
    n_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad) 
    # print('number of params:', n_parameters)

    if "Validation_all_G" in args.data_val:
        max_colon = [[117,58],[422,349]] ## 335_108
    else:
        max_colon = [[69,131],[459,371]] ## 226_107

    print("Data Loading ....", end='\r')
    val_dataset = BoundingBoxDataSet(args.data_val, flag_shuffle=False)
    data_loader_val = DataLoader(val_dataset, batch_size=args.batch_size, shuffle=False, num_workers=0)
    del val_dataset
    print("Data Loaded !!!!!")

    os.makedirs(args.output_dir, exist_ok=True)
    Dir_Statics = f"{args.output_dir}/Results_{args.output_name}.xlsx"
    wb_result = openpyxl.Workbook()
    worksheet = wb_result.active
    worksheet.append(['Epoch', 'T_IoU', "T_Acc (%)", "N_Acc (%)", 'T_IoU', "T_Acc (%)", "N_Acc (%)"])
    wb_result.save(Dir_Statics)

    for epoch in args.model_epoch_list:
        output_dir = f"{args.output_dir}Epoch_{epoch[2:]}/"
    
        print("Model loading ....", end='\r')
        ckpt = torch.load(f"{args.model_path}DETR_{epoch}_S1.pth", map_location=device)  ## S1 S2
        model.load_state_dict(ckpt['model_state_dict'])
        ckpt_2 = torch.load(f"{args.model_path}DETR_{epoch}_S2.pth", map_location=device)
        model_2.load_state_dict(ckpt_2['model_state_dict'])
        print("Model loaded !!!!!")

        print(f"{output_dir} / {args.device} START")
        os.makedirs(output_dir, exist_ok=True)

        Dir_Statics_val = f"{output_dir}/Result_{args.output_name}_{epoch[2:]}.xlsx"
        Dir_Statics_slice_s1 = f"{output_dir}/Slice_s1_{args.output_name}_{epoch[2:]}.xlsx"
        Dir_Statics_box_s1 = f"{output_dir}/BoxInfo_s1_{args.output_name}_{epoch[2:]}.xlsx"
        Dir_Statics_slice_s2 = f"{output_dir}/Slice_s2_{args.output_name}_{epoch[2:]}.xlsx"
        Dir_Statics_box_s2 = f"{output_dir}/BoxInfo_s2_{args.output_name}_{epoch[2:]}.xlsx"

        Dir_Statics_val_s2_ori = f"{output_dir}/Result_S2_{args.output_name}_{epoch[2:]}.xlsx"
        Dir_Statics_slice_s2_ori = f"{output_dir}/SliceOri_s2_{args.output_name}_{epoch[2:]}.xlsx"
        
        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(['Type', 'IoU', 'Dice', 'Recall', 'Precision', 'Speci', "Data_Num", "Acc_Num", "Acc (%)", "Score", "Score_W"])
        wb_result.save(Dir_Statics_val)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(["Patient", "Slice", "Is_Normal", "Is_Next", "GT 1", "Pred 1", "GT 2", "Pred 2", "IoU", "Dice", 'Recall', 'Precision', 'Speci', "Score_1st", "Score_2nd"])
        wb_result.save(Dir_Statics_slice_s1)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(["Patient", "Slice", "Is_Normal", "Is_Next", "GT 1", "Pred 1", "GT 2", "Pred 2", "B1 1", "B1 2", "B1_S", "B1_S_W", "B2 1", "B2 2", "B2_S", "B2_S_W",
                                    "B3 1", "B3 2", "B3_S", "B3_S_W", "B4 1", "B4 2", "B4_S", "B4_S_W"])
        wb_result.save(Dir_Statics_box_s1)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(["Patient", "Slice", "Is_Normal", "GT 1", "Pred 1", "GT 2", "Pred 2", "IoU", "Dice", 'Recall', 'Precision', 'Speci', "Score_1st", "Score_2nd"])
        wb_result.save(Dir_Statics_slice_s2)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(["Patient", "Slice", "Is_Normal", "GT 1", "Pred 1", "GT 2", "Pred 2", "B1 1", "B1 2", "B1_S", "B1_S_W", "B2 1", "B2 2", "B2_S", "B2_S_W",
                                    "B3 1", "B3 2", "B3_S", "B3_S_W", "B4 1", "B4 2", "B4_S", "B4_S_W"])
        wb_result.save(Dir_Statics_box_s2)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(["Patient", "Slice", "Is_Normal", "Cut 1", "Cut 2", "GT 1", "Pred 1", "GT 2", "Pred 2", "IoU", "Dice", 'Recall', 'Precision', 'Speci'])
        wb_result.save(Dir_Statics_slice_s2_ori)
        
        # 학습이 잘 되고 있는지, 확인하기 위해서 위에서 사용하는 함수를 그대로 사용한다. 
        val_start = time.time()
        if "ColonMax_" in args.model_path:
            print(f"Colon Max Validation Start!!!!!! max colon = {max_colon}") 
            results_s1, data_num_s2, results_s2, results_slice, resul_s2_ori = val_max_colon(args, model, ppro, criter, data_loader_val, model_2, ppro_2, criter_2, epoch[2:], device, max_colon, colon_info)
        elif "S2OnlyT_"  in args.model_path:
            if "Square_" in args.model_path:
                print(f"S2 Only Tumor Square Start!!!!!! ") 
                results_s1, data_num_s2, results_s2, results_slice, resul_s2_ori = val_s2OnlyT(args, model, ppro, criter, data_loader_val, model_2, ppro_2, criter_2, epoch[2:], device, colon_info)
            else:
                print(f"S2 Only Tumor Start!!!!!! ")
                results_s1, data_num_s2, results_s2, results_slice, resul_s2_ori = val_s2OnlyT_NSq(args, model, ppro, criter, data_loader_val, model_2, ppro_2, criter_2, epoch[2:], device, colon_info)
        elif "Contour_" in args.model_path:
            print(f"Contour Start!!!!!! ") 
            results_s1, data_num_s2, results_s2, results_slice, resul_s2_ori = val_contour(args, model, ppro, criter, data_loader_val, model_2, ppro_2, criter_2, epoch[2:], device)
        else:
            print("Else Val Start")
            results_s1, data_num_s2, results_s2, results_slice, resul_s2_ori = val(args, model, ppro, criter, data_loader_val, model_2, ppro_2, criter_2, epoch[2:], device, colon_info)
        ## Dir_Statics_val
        score_s1 = results_s1[0]
        a_acc_s1 = round(((score_s1[6] + score_s1[8])/(score_s1[5] + score_s1[7])), 6) * 100.0
        t_acc_s1 = round((score_s1[6]/score_s1[5]), 6) * 100.0
        n_acc_s1 = round((score_s1[8]/score_s1[7]), 6) * 100.0
        box_s1 = results_s1[3]
        score_s2 = results_s2[0]
        is_s2_nor = True
        try:
            a_acc_s2 = round(((score_s2[6] + score_s2[8])/(score_s2[5] + score_s2[7])), 6) * 100.0
            t_acc_s2 = round((score_s2[6]/score_s2[5]), 6) * 100.0
            n_acc_s2 = round((score_s2[8]/score_s2[7]), 6) * 100.0
        except:
            a_acc_s2 = 0
            t_acc_s2 = 0
            n_acc_s2 = 0
            is_s2_nor = False
        box_s2 = results_s2[3]
        wb = openpyxl.load_workbook(Dir_Statics_val)
        ws = wb.active
        result_list = ["All_S1", round(np.mean(score_s1[0]), 4), round(np.mean(score_s1[1]), 4), round(np.mean(score_s1[2]), 4), round(np.mean(score_s1[3]), 4), round(np.mean(score_s1[4]), 4), 
                                len(data_loader_val.dataset), score_s1[6] + score_s1[8], a_acc_s1, round(np.mean(box_s1[9]), 4), round(np.mean(box_s1[10]), 4)]
        ws.append(result_list)
        iou = []
        dice = []
        speci = []
        sensitiy = []
        precision = []
        scores = []
        scores_2 = []
        for i in range(len(results_s1[4])):
            iou.append(results_s1[4][i][0])
            dice.append(results_s1[4][i][1])
            sensitiy.append(results_s1[4][i][2])
            precision.append(results_s1[4][i][3])
            speci.append(results_s1[4][i][4])
            scores.append(results_s1[4][i][5])
            scores_2.append(results_s1[4][i][6])
        result_list = ["Tumor", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                        score_s1[5], score_s1[6], t_acc_s1, round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
        ws.append(result_list)

        t_iou_s1 = round(np.mean(iou), 4)
        iou = []
        dice = []
        speci = []
        sensitiy = []
        precision = []
        scores = []
        scores_2 = []
        for i in range(len(results_s1[5])):
            iou.append(results_s1[5][i][0])
            dice.append(results_s1[5][i][1])
            sensitiy.append(results_s1[5][i][2])
            speci.append(results_s1[5][i][4])
            precision.append(results_s1[5][i][3])
            scores.append(results_s1[5][i][5])
            scores_2.append(results_s1[5][i][6])
        result_list = ["Normal", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                        score_s1[7], score_s1[8], n_acc_s1, round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
        ws.append(result_list)
        if "Contour_" in args.model_path:
            result_list = ["All_S2", round(np.mean(score_s2[0]), 4), round(np.mean(score_s2[1]), 4), round(np.mean(score_s2[2]), 4), round(np.mean(score_s2[3]), 4), round(np.mean(score_s2[4]), 4), 
                                    "", "", "", round(np.mean(box_s2[9]), 4), round(np.mean(box_s2[10]), 4)]
        else:
            result_list = ["All_S2", round(np.mean(score_s2[0]), 4), round(np.mean(score_s2[1]), 4), round(np.mean(score_s2[2]), 4), round(np.mean(score_s2[3]), 4), round(np.mean(score_s2[4]), 4), 
                                    data_num_s2, score_s2[6] + score_s2[8], a_acc_s2, round(np.mean(box_s2[9]), 4), round(np.mean(box_s2[10]), 4)]
        ws.append(result_list)
        iou = []
        dice = []
        speci = []
        sensitiy = []
        precision = []
        scores = []
        scores_2 = []
        for i in range(len(results_s2[4])):
            iou.append(results_s2[4][i][0])
            dice.append(results_s2[4][i][1])
            sensitiy.append(results_s2[4][i][2])
            speci.append(results_s2[4][i][4])
            precision.append(results_s2[4][i][3])
            scores.append(results_s2[4][i][5])
            scores_2.append(results_s2[4][i][6])
        if "Contour_" in args.model_path:
            result_list = ["Tumor", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                            "", "", "", round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
        else:
            result_list = ["Tumor", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                            score_s2[5], score_s2[6], t_acc_s2, round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
        ws.append(result_list)

        t_iou_s2 = round(np.mean(iou), 4)
        if is_s2_nor:
            iou = []
            dice = []
            speci = []
            sensitiy = []
            precision = []
            scores = []
            scores_2 = []
            for i in range(len(results_s2[5])):
                iou.append(results_s2[5][i][0])
                dice.append(results_s2[5][i][1])
                sensitiy.append(results_s2[5][i][2])
                speci.append(results_s2[5][i][4])
                precision.append(results_s2[5][i][3])
                scores.append(results_s2[5][i][5])
                scores_2.append(results_s2[5][i][6])
            if "Contour_" in args.model_path:
                result_list = ["Normal", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                                "", "", "", round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
            else:
                result_list = ["Normal", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(sensitiy), 4), round(np.mean(precision), 4), round(np.mean(speci), 4),
                                score_s2[7], score_s2[8], n_acc_s2, round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
            ws.append(result_list)
        result_list = ["Resize", round(np.mean(resul_s2_ori[5]), 4), round(np.mean(resul_s2_ori[6]), 4), 
                        round(np.mean(resul_s2_ori[7]), 4), round(np.mean(resul_s2_ori[8]), 4), round(np.mean(resul_s2_ori[9]), 4)]
        ws.append(result_list)
        wb.save(Dir_Statics_val)

        wb_eval = openpyxl.load_workbook(Dir_Statics_slice_s1)
        ws = wb_eval.active
        for s in range(len(results_slice[0])):
            print(f"Write Slice Excel S1 {s+1} / {len(results_slice[0])}", end="\r")
            reulst_list = [results_slice[0][s], results_slice[1][s], results_slice[2][s],  results_slice[3][s], str(results_s1[1][s][0])[1:-1], str(results_s1[2][s][0])[1:-1], 
                            str(results_s1[1][s][1])[1:-1], str(results_s1[2][s][1])[1:-1], 
                            np.round(score_s1[0][s],4), np.round(score_s1[1][s],4), np.round(score_s1[2][s],4), np.round(score_s1[3][s],4),  np.round(score_s1[4][s],4), 
                            np.round(results_s1[3][9][s],4), np.round(results_s1[3][10][s],4)]
            ws.append(reulst_list)
        wb_eval.save(Dir_Statics_slice_s1)

        wb_eval_box = openpyxl.load_workbook(Dir_Statics_box_s1)
        ws = wb_eval_box.active
        for s in range(len(results_slice[0])):
            print(f"Write BoxInfo Excel S1 {s+1} / {len(results_slice[0])}", end="\r")
            reulst_list = [results_slice[0][s], results_slice[1][s], results_slice[2][s],  results_slice[3][s], 
                            str(results_s1[1][s][0])[1:-1], str(results_s1[2][s][0])[1:-1], str(results_s1[1][s][1])[1:-1], str(results_s1[2][s][1])[1:-1], 
                            str(results_s1[3][0][s][0][0])[1:-1], str(results_s1[3][0][s][0][1])[1:-1], round(results_s1[3][1][s][0],4), round(results_s1[3][2][s][0],4),
                            str(results_s1[3][0][s][1][0])[1:-1], str(results_s1[3][0][s][1][1])[1:-1], round(results_s1[3][1][s][1],4), round(results_s1[3][2][s][1],4),
                            str(results_s1[3][0][s][2][0])[1:-1], str(results_s1[3][0][s][2][1])[1:-1], round(results_s1[3][1][s][2],4), round(results_s1[3][2][s][2],4),
                            str(results_s1[3][0][s][3][0])[1:-1], str(results_s1[3][0][s][3][1])[1:-1], round(results_s1[3][1][s][3],4), round(results_s1[3][2][s][3],4)]
            ws.append(reulst_list)
        wb_eval_box.save(Dir_Statics_box_s1)
        ##[Iou_s1, Dice_s1, Sensitivity_s1, Precision_s1, Speciticity_s1, T_Num_Total_s1, T_Num_Detect_s1, N_Num_Total_s1, N_Num_Detect_s1]
        wb_eval = openpyxl.load_workbook(Dir_Statics_slice_s2)
        ws = wb_eval.active
        for s in range(len(results_slice[4])):
            print(f"Write Slice Excel S2 {s+1} / {len(results_slice[4])}", end="\r")
            reulst_list = [results_slice[4][s], results_slice[5][s], results_slice[6][s], 
                            str(results_s2[1][s][0])[1:-1], str(results_s2[2][s][0])[1:-1], str(results_s2[1][s][1])[1:-1], str(results_s2[2][s][1])[1:-1], 
                            np.round(score_s2[0][s],4), np.round(score_s2[1][s],4), np.round(score_s2[2][s],4), np.round(score_s2[3][s],4),  np.round(score_s2[4][s],4), 
                            np.round(results_s2[3][9][s],4), np.round(results_s2[3][10][s],4)]
            ws.append(reulst_list)
        wb_eval.save(Dir_Statics_slice_s2)

        wb_eval_box = openpyxl.load_workbook(Dir_Statics_box_s2)
        ws = wb_eval_box.active
        for s in range(len(results_slice[4])):
            print(f"Write BoxInfo Excel S2 {s+1} / {len(results_slice[4])}", end="\r")
            reulst_list = [results_slice[4][s], results_slice[5][s], results_slice[6][s], 
                            str(results_s2[1][s][0])[1:-1], str(results_s2[2][s][0])[1:-1], str(results_s2[1][s][1])[1:-1], str(results_s2[2][s][1])[1:-1], 
                            str(results_s2[3][0][s][0][0])[1:-1], str(results_s2[3][0][s][0][1])[1:-1], round(results_s2[3][1][s][0],4), round(results_s2[3][2][s][0],4),
                            str(results_s2[3][0][s][1][0])[1:-1], str(results_s2[3][0][s][1][1])[1:-1], round(results_s2[3][1][s][1],4), round(results_s2[3][2][s][1],4),
                            str(results_s2[3][0][s][2][0])[1:-1], str(results_s2[3][0][s][2][1])[1:-1], round(results_s2[3][1][s][2],4), round(results_s2[3][2][s][2],4),
                            str(results_s2[3][0][s][3][0])[1:-1], str(results_s2[3][0][s][3][1])[1:-1], round(results_s2[3][1][s][3],4), round(results_s2[3][2][s][3],4)]
            ws.append(reulst_list)
        wb_eval_box.save(Dir_Statics_box_s2)
    
        # results_s2_ori = [Patient_S2, Slice_S2, Iszero_S2, S2_GT_marks_ori, S2_PD_marks_ori, Iou_s2_ori, Dice_s2_ori, Speciticity_s2_ori, Sensitivity_s2_ori, Precision_s2_ori]
        # worksheet.append(["Patient", "Slice", "Is_Normal", "GT 1", "Pred 1", "GT 2", "Pred 2", "IoU", "Dice", 'Speci', 'Sensi', 'Precision'])
        # wb_result.save(Dir_Statics_slice_s2_ori)

        wb_eval = openpyxl.load_workbook(Dir_Statics_slice_s2_ori)
        ws = wb_eval.active
        for s in range(len(resul_s2_ori[0])):
            print(f"Write Slice Excel S2 Ori {s+1} / {len(resul_s2_ori[0])}", end="\r")
            reulst_list = [resul_s2_ori[0][s], resul_s2_ori[1][s], resul_s2_ori[2][s], str(resul_s2_ori[10][s][0])[1:-1], str(resul_s2_ori[10][s][1])[1:-1], 
                            str(resul_s2_ori[3][s][0])[1:-1], str(resul_s2_ori[4][s][0])[1:-1], str(resul_s2_ori[3][s][1])[1:-1], str(resul_s2_ori[4][s][1])[1:-1], 
                            np.round(resul_s2_ori[5][s],4), np.round(resul_s2_ori[6][s],4), np.round(resul_s2_ori[7][s],4), np.round(resul_s2_ori[8][s],4), np.round(resul_s2_ori[9][s],4)]
            ws.append(reulst_list)
        wb_eval.save(Dir_Statics_slice_s2_ori)

        epoch_time = (time.time() - val_start) / 60
        time_H = int(epoch_time // 60)
        time_M = epoch_time % 60
        
        print("")
        patient_score(output_dir)
        if "Contour_" in args.model_path:
            s2_num, s2_t_num, s2_n_num, s2_t_num_acc, s2_n_num_acc = s2_results(Dir_Statics_slice_s2)
            all_acc = ((s2_t_num_acc+s2_n_num_acc)/s2_num) * 100.0
            t_acc = s2_t_num_acc/s2_t_num * 100.0
            n_acc = s2_n_num_acc/s2_n_num * 100.0

            s2_t_num_2 = score_s1[5] 
            s2_n_num_2 = score_s1[7]
            s2_num_2 = s2_t_num_2 + s2_n_num_2
            s2_n_num_acc_2 = s2_n_num_acc + (score_s1[7] - s2_n_num)

            all_acc_2 = ((s2_t_num_acc+s2_n_num_acc_2)/s2_num_2) * 100.0
            t_acc_2 = s2_t_num_acc/s2_t_num_2 * 100.0
            n_acc_2 = s2_n_num_acc_2/s2_n_num_2 * 100.0
            
            wb = openpyxl.load_workbook(Dir_Statics_val)
            ws = wb.active
            result_list = ["All_S2", "", "", "", "", "", s2_num_2, s2_t_num_acc+s2_n_num_acc_2, round(all_acc_2, 4), "", s2_num, s2_t_num_acc+s2_n_num_acc, round(all_acc, 4)]
            ws.append(result_list)
            result_list = ["Tumor", "", "", "", "", "", s2_t_num_2, s2_t_num_acc, round(t_acc_2, 4), "", s2_t_num, s2_t_num_acc, round(t_acc, 4)]
            ws.append(result_list)
            result_list = ["Normal", "", "", "", "", "", s2_n_num_2, s2_n_num_acc_2, round(n_acc_2, 4), "", s2_n_num, s2_n_num_acc, round(n_acc, 4)]
            ws.append(result_list)
            wb.save(Dir_Statics_val)

                
            wb = openpyxl.load_workbook(Dir_Statics)
            ws = wb.active
            result_list = [epoch[2:], t_iou_s1, t_acc_s1, n_acc_s1, t_iou_s2, round(t_acc_2, 4), round(n_acc_2, 4)] 
            ws.append(result_list)
            wb.save(Dir_Statics)
        else:
            wb = openpyxl.load_workbook(Dir_Statics)
            ws = wb.active
            result_list = [epoch[2:], t_iou_s1, t_acc_s1, n_acc_s1, t_iou_s2, t_acc_s2, n_acc_s2] 
            ws.append(result_list)
            wb.save(Dir_Statics)
        print(f"Epoch {epoch[2:]} Val --- Time = {time_H} H {round(time_M,2)} M / {args.device}")

        # s1_cor_num = (score_s1[6] + score_s1[8])
        # s2_cor_num = (score_s2[6] + score_s2[8])
        # print(f"Epoch {epoch[2:]} Val --- 1st Data = {s1_cor_num}/{len(data_loader_val.dataset)}, 2nd Data = {s2_cor_num}/{data_num_s2}, Time = {time_H} H {round(time_M,2)} M")
    print(f"{args.output_dir} / {args.device} FINSH")
    

if __name__ == '__main__':
    parser = argparse.ArgumentParser('DETR training and evaluation script', parents=[get_args_parser()])
    args = parser.parse_args()
    print(f"Dilation Mode = {args.dilation}")
    
    # if not args.eval and not args.add_hg:
    #     if args.output_dir:
    #         Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    main(args)


