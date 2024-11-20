# Copyright (c) Facebook, Inc. and its affiliates. All Rights Reserved
"""
,
            "args" : ["--batch_size", "2", 
                "--no_aux_loss", 
                "--eval", 
                "--resume", "checkpoints/detr-r50-e632da11.pth", 
                "--num_workers", "4",
                "--world_size", "2",
                "--coco_path", "/dataset/coco",
                "--output_dir", "result"]
"""
# 아래 코드 주석에서 ## 으로 적어 놓은 주석은, 내가 적은 주석. 영어 말고 최대한 한글로 적어 놓기. 미래의 나를 위해서
import argparse
import datetime
from itertools import count
import json
import random
import time
from pathlib import Path

import numpy as np
import torch
from torch.utils.data import DataLoader, DistributedSampler

# import datasets
import util.misc as utils
from learning import train_epoch, evaluate

from models import build_model
from via import via

import datetime
import random

from BoundingBoxDataSet import BoundingBoxDataSet
import matplotlib.pyplot as plt
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
import openpyxl
from calculate import *

def get_args_parser():
    GPU_Name = "cuda:0"
    device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
    try:
        gpu_info = torch.cuda.get_device_name(device)
    except:
        GPU_Name = "cuda:0"
        device = torch.device(GPU_Name if torch.cuda.is_available() else "cpu")
        gpu_info = torch.cuda.get_device_name(device)
    
    Computer = ""
    Batch_Size = 16
    Num_Epoch = 30
    Nas_Mount = "Y:/" 
    Data_Mount = "D:/" 
    Demo = ''
    DIR_DATA = f"{Data_Mount}yskim/BoundingBox/data/processed/LN_201014_HU/"
    # DIR_DATA = f"{Data_Mount}yskim/BoundingBox/data/processed/LN_2014_HU/"
    if "3060" in gpu_info:
        Computer = "Test"
        # Computer = "Demo"
        Batch_Size = int(Batch_Size/2)
        if Computer == "Demo":
            Num_Epoch = int(Num_Epoch/5)
        # Val_Epoch = int(Val_Epoch/10) 
        # Val_Epoch_early = 3
        Demo = "_tmp"
        Data_Mount = "Y:/" 
        DIR_DATA = f"{Data_Mount}yskim/BoundingBox/data/processed/tmp/"
    elif "3090" in gpu_info:
        Computer = "Com"
    else:
        print(f"Check this PC GPU = {gpu_info}")
        exit()

    DIR_RESULT_PARENT = f"{Nas_Mount}yskim/BoundingBox/result/"    
    DIR_DATA_TRAIN = f"{DIR_DATA}train/" 
    DIR_DATA_TEST = f"{DIR_DATA}test/" 
    DIR_DATA_SEG = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/Fast_colon/"
    DIR_DATA_SEG_Contour = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/Fast_colon_contour/"
    # DIR_DATA_SEG = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/2014_Fast_colon/"
    # DIR_DATA_SEG_Contour = f"{Data_Mount}yskim/BoundingBox/data/raw/Colorectal/2014_Fast_colon_contour/"
    dir_data_split = DIR_DATA.split('_')

    DIR_DATA_5Slice = None
    ## train path
    # DIR_DATA_5Slice = f"{Data_Nas}yskim/BoundingBox/data/processed/LN2_20All_HU_5slice/"

    PATH_PRETRAINED = None
    check_path = "240612_20h40m_DT1014_Q9_C2_BR101_R101_NB08_UB4_Square_Contour_V3_Com/"
    Check_last = [x for x in os.listdir(f"{DIR_RESULT_PARENT}{check_path}Checkpoint/") if not "999" in x and "DETR" in x][-1]
    PATH_PRETRAINED = f"{DIR_RESULT_PARENT}{check_path}Checkpoint/{Check_last}"
    print(f"Pretrain model = {PATH_PRETRAINED}")
    
    Class_Num = 2 ## 2= tumor+bg 3= tumor+zero+bg
    Zero_Label = 0
    if Class_Num == 3:
        Zero_Label = 1
    Query_Num = 9
    Box_TH = 0.8
    th = str(Box_TH).split('.')
    Box_Mode = "NB" ## SB NB TG
    Box_UseNum = 4
    Data_Year = ''
    for year in dir_data_split:
        if "20" in year:
            Data_Year = year
    Data_Year = Data_Year[2:]  
    Pretrain_model = "detr-r101-dc5.pth"
    # Pretrain_model = None
    Back_bone = 'resnet101'
    # Back_bone = 'resnet50'
    # DIR_RocoResnet = f"{Nas_Mount}yskim/BoundingBox/result/Roco_Resnet101_Pretrained_Com_3090/CheckPoint/net_93.pt"
    DIR_RocoResnet = f"{Nas_Mount}yskim/BoundingBox/result/Roco_Resnet101_DT_3CH_Com_3090/CheckPoint/net_93.pt"

    multi_slice = ''
    if DIR_DATA_5Slice is not None:
        multi_slice = '5Slice_'
    NAME_RESULT = f"DT{Data_Year}_Q{Query_Num}_C{Class_Num}_BR101_R101_{multi_slice}{Box_Mode}{th[0]}{th[1]}_UB{Box_UseNum}_Square_Contour_V3" ## TNSame_ ColonMax
    today = str(datetime.date.today().strftime('%y%m%d'))
    now = str(datetime.datetime.now().strftime('%Hh%Mm'))
    DIR_RESULT = f"{DIR_RESULT_PARENT}{today}_{now}_{NAME_RESULT}_{Computer}/"         
   
    if PATH_PRETRAINED is not None:   
        DIR_RESULT = f"{DIR_RESULT_PARENT}{check_path}"
  
    if "3060" in gpu_info:
        Box_TH = 0.2
        
    parser = argparse.ArgumentParser('Set transformer detector', add_help=False)
    parser.add_argument('--lr', default=1e-4, type=float)
    parser.add_argument('--lr_backbone', default=1e-5, type=float)
    parser.add_argument('--batch_size', default=Batch_Size, type=int)
    parser.add_argument('--weight_decay', default=1e-4, type=float)
    parser.add_argument('--epochs', default=Num_Epoch, type=int)
    parser.add_argument('--lr_drop', default=200, type=int)
    parser.add_argument('--clip_max_norm', default=0.1, type=float, help='gradient clipping max norm')

    # Model parameters
    parser.add_argument('--frozen_weights', type=str, default=None, help="Path to the pretrained model. If set, only the mask head will be trained")
    # * Backbone
    parser.add_argument('--backbone', default=Back_bone, type=str, help="Name of the convolutional backbone to use")
    parser.add_argument('--dilation', action='store_false',
                        help="If true, we replace stride with dilation in the last convolutional block (DC5)") # stride를 적용하지 않고, dilation을 추가해서 backbone output resolution을 높힌다.
    parser.add_argument('--position_embedding', default='sine', type=str, choices=('sine', 'learned'), help="Type of positional embedding to use on top of the image features")
    parser.add_argument('--pretraind_model', default=Pretrain_model, type=str, help="Name of the pretrain model")

    # * Transformer
    parser.add_argument('--enc_layers', default=6, type=int,
                        help="Number of encoding layers in the transformer")
    parser.add_argument('--dec_layers', default=6, type=int,
                        help="Number of decoding layers in the transformer")
    parser.add_argument('--dim_feedforward', default=2048, type=int,
                        help="Intermediate size of the feedforward layers in the transformer blocks")
    parser.add_argument('--hidden_dim', default=256, type=int,
                        help="Size of the embeddings (dimension of the transformer)")
    parser.add_argument('--dropout', default=0.1, type=float,
                        help="Dropout applied in the transformer")
    parser.add_argument('--nheads', default=8, type=int,
                        help="Number of attention heads inside the transformer's attentions")
    parser.add_argument('--num_queries', default=Query_Num, type=int, help="Number of query slots")
    parser.add_argument('--num_classes', default=Class_Num, type=int, help="Number of class")
    parser.add_argument('--zero_label', default=Zero_Label, type=int, help="zero label 0 or 2")
    parser.add_argument('--pre_norm', action='store_true')

    # * Segmentation
    parser.add_argument('--masks', action='store_true',
                        help="Train segmentation head if the flag is provided")

    # Loss
    parser.add_argument('--no_aux_loss', dest='aux_loss', action='store_false',
                        help="Disables auxiliary decoding losses (loss at each layer)")  # args.aux_loss == false
                        
    # * Matcher
    parser.add_argument('--set_cost_class', default=1, type=float, help="Class coefficient in the matching cost")
    parser.add_argument('--set_cost_bbox', default=5, type=float, help="L1 box coefficient in the matching cost")
    parser.add_argument('--set_cost_giou', default=2, type=float, help="giou box coefficient in the matching cost")

    # * Loss coefficients
    parser.add_argument('--mask_loss_coef', default=1, type=float)
    parser.add_argument('--dice_loss_coef', default=1, type=float)
    parser.add_argument('--bbox_loss_coef', default=5, type=float)
    parser.add_argument('--giou_loss_coef', default=2, type=float)
    parser.add_argument('--eos_coef', default=0.1, type=float, help="Relative classification weight of the no-object class")

    # dataset parameters
    parser.add_argument('--dataset_file', default='coco')
    parser.add_argument('--coco_path', type=str)
    parser.add_argument('--coco_panoptic_path', type=str)
    parser.add_argument('--remove_difficult', action='store_true')

    parser.add_argument('--output_dir', default=DIR_RESULT, help='path where to save, empty for no saving')
    parser.add_argument('--output_name', default=NAME_RESULT, help='path where to save name')
    parser.add_argument('--device', default=GPU_Name, help='device to use for training / testing')
    parser.add_argument('--data_train', default=DIR_DATA_TRAIN, help="train data dir")
    parser.add_argument('--data_test', default=DIR_DATA_TEST, help="test data dir")
    parser.add_argument('--data_seg', default=DIR_DATA_SEG, help="colon seg data dir")
    parser.add_argument('--data_seg_contour', default=DIR_DATA_SEG_Contour, help="contour data dir")
    parser.add_argument('--data_5slice', default=DIR_DATA_5Slice, help="5silce data dir")
    parser.add_argument('--seed', default=42, type=int)
    parser.add_argument('--resume', default='', help='resume from checkpoint')
    parser.add_argument('--start_epoch', default=1, type=int, metavar='N', help='start epoch')
    parser.add_argument('--computer', default=Computer, type=str, help='this computer name')
    parser.add_argument('--model_pretrain', default=PATH_PRETRAINED, help='path where to pretrain model')
    parser.add_argument('--num_workers', default=0, type=int) # 

    # distributed training parameters
    parser.add_argument('--world_size', default=1, type=int, help='number of distributed processes')
    parser.add_argument('--dist_url', default='env://', help='url used to set up distributed training')

    ## parameters
    parser.add_argument('--box_th', default=Box_TH, type=float, help='box threshold')
    parser.add_argument('--box_mode', default=Box_Mode, type=str, help='box mode')
    parser.add_argument('--box_usenum', default=Box_UseNum, type=int, help='box usenum')
    parser.add_argument('--nas_mount', default=Nas_Mount, type=str, help='nas mount path')
    parser.add_argument('--rocoresnet_dir', default=DIR_RocoResnet, help='path where to roco resnet save')
    return parser


def main(args):
    utils.init_distributed_mode(args) # Multi-GPU 사용할 거라면, args.gpu / args.world_size / args.rank 가 여기서 정의 된다.
    # print("git:\n  {}\n".format(utils.get_sha()))

    if args.frozen_weights is not None:
        assert args.masks, "Frozen training is meant for segmentation only"
    # print(args)
    
    device = torch.device(args.device)
    print(f"Using device = {device} - {torch.cuda.get_device_name(device)}")
    gpu_title = str(torch.cuda.get_device_name(device)).split(' ')[-1]

    # Multi-GPU 사용할 거라면, fix the seed for reproducibility 
    seed = args.seed + utils.get_rank() 
    torch.manual_seed(seed)
    np.random.seed(seed)
    random.seed(seed)

    ## 전까지는, args 설정 및 seed 설정이 전부이다.
    ## 여기가 진짜 시작 
    model, criterion, postprocessors = build_model(args)
    model_2, criterion_2, postprocessors_2 = build_model(args)
    ## The models are also available via torch hub, to load DETR R50 with pretrained weights simply do:
    if args.pretraind_model is not None:
        checkpoint = torch.load(f"{args.nas_mount}/yskim/BoundingBox/code/DETR/{args.pretraind_model}", map_location=device)
        del checkpoint["model"]["class_embed.weight"]
        del checkpoint["model"]["class_embed.bias"]
        del checkpoint["model"]["query_embed.weight"]
        model.load_state_dict(checkpoint['model'], strict=False)
        model_2.load_state_dict(checkpoint['model'], strict=False)
        print("Load Pretrained COCO Data model ")
    
    model.to(device)
    model_2.to(device)
    # print(model.backbone)
    # for k in model.backbone:
        # print(k.body)
        # path_model = "Y:/yskim/BoundingBox/code/DETR/res_detr-r101-dc5.pth"
        # torch.save({'model_state_dict':k.body.state_dict()}, path_model)
        # exit()
    ####
    # dtpt = torch.load(args.rocoresnet_dir , map_location=device)
    # list_check = list(dtpt['model_state_dict'].items())
    # del_list = []
    # list_check_body = []
    # for i in range(len(list_check)):
    #     if "fc." not in list_check[i][0]:
    #         # del_list.append(f"0.body.{list_check[i]}") 
    #         # del_list.append(f"{list_check[i]}")
    #         list_check_body.append((f"0.body.{list_check[i][0]}",list_check[i][1])) 
    #     # else:
    #     #     list_check_body.append((f"0.body.{list_check[i][0]}",list_check[i][1]))
    # load_check = list_check_body
    # load_check = dict(load_check)
    # model.backbone.load_state_dict(load_check)
    # print("ROCO Resnet load")

    model_without_ddp = model
    model_without_ddp_2 = model_2
    # if args.distributed:
    #     model = torch.nn.parallel.DistributedDataParallel(model, device_ids=[args.gpu])
    #     model_without_ddp = model.module
    
    ## 총 파라메터 갯수 계산 하는 방법
    n_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad) 
    # print('number of params:', n_parameters)

    ## backbone / Transformer-encoder, decoder / detector head 각각의 learning rate를 다르게 주는 방법
    param_dicts = [
        {"params": [p for n, p in model_without_ddp.named_parameters() if "backbone" not in n and p.requires_grad]},
        {
            "params": [p for n, p in model_without_ddp.named_parameters() if "backbone" in n and p.requires_grad],
            "lr": args.lr_backbone,
        },
    ]
    param_dicts_2 = [
        {"params": [p for n, p in model_without_ddp_2.named_parameters() if "backbone" not in n and p.requires_grad]},
        {
            "params": [p for n, p in model_without_ddp_2.named_parameters() if "backbone" in n and p.requires_grad],
            "lr": args.lr_backbone,
        },
    ]

    ## optimizer와 ir_scheduler 설정
    optimizer = torch.optim.AdamW(param_dicts, lr=args.lr, weight_decay=args.weight_decay)
    lr_scheduler = torch.optim.lr_scheduler.StepLR(optimizer, args.lr_drop)
    optimizer_2 = torch.optim.AdamW(param_dicts_2, lr=args.lr, weight_decay=args.weight_decay)
    lr_scheduler_2 = torch.optim.lr_scheduler.StepLR(optimizer_2, args.lr_drop)

    test_dataset = BoundingBoxDataSet(args.data_test, flag_shuffle=False)
    data_loader_test = DataLoader(test_dataset, batch_size=args.batch_size, shuffle=False, num_workers=0)
    del test_dataset
   
    max_segmark = []
    # max_segmark = [[27,168],[474,401]] ## 2011_0154 085
    
    print(f"{args.output_dir} START")
    os.makedirs(args.output_dir, exist_ok=True)

    ## If args.eval=False 이라면, 바로 위 코드 안하고, Traiining 시작하기
    args.epochs += 1
    print("Start training")

    start_time = time.time()
    train_bbox_s1 = []
    val_bbox_s1 = []
    train_ciou_s1 = []
    val_ciou_s1 = []
    train_score_s1 = []
    val_score_s1 = []
    train_bbox_s2 = []
    val_bbox_s2 = []
    train_ciou_s2 = []
    val_ciou_s2 = []
    train_score_s2 = []
    val_score_s2 = []
    train_iou_s1 = []
    train_iou_s2 = []
    val_iou_s1 = []
    val_iou_s2 = []
 
    Dir_Loss_train = f"{args.output_dir}/Loss_train_{args.output_name}.xlsx"
    Dir_Loss_val = f"{args.output_dir}/Loss_val_{args.output_name}.xlsx"
    Dir_Statics_train = f"{args.output_dir}/Statics_train_{args.output_name}.xlsx"
    Dir_Statics_val = f"{args.output_dir}/Statics_val_{args.output_name}.xlsx"
    if args.model_pretrain is None:
        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Data", "Bbox_S1", "Ciou_S1", "Score_S1", "Data", "Bbox_S2", "Ciou_S2", "Score_S2"])
        wb.save(Dir_Loss_train)

        wb = openpyxl.Workbook()
        worksheet = wb.active
        worksheet.append(["Epoch", "Data", "Bbox_S1", "Ciou_S1", "Score_S1" , "Data", "Bbox_S2", "Ciou_S2", "Score_S2"])
        wb.save(Dir_Loss_val)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(['Epoch', 'Iou_S1', 'Dice', 'Speci', 'Sensi', 'Precision', 't_num', 't_acc','n_num', 'n_acc', ' ', 
                            'Iou_S2', 'Dice', 'Speci', 'Sensi', 'Precision', 't_num', 't_acc','n_num', 'n_acc'])
        wb_result.save(Dir_Statics_train)

        wb_result = openpyxl.Workbook()
        worksheet = wb_result.active
        worksheet.append(['Epoch', 'Iou_S1', 'Dice', 'Speci', 'Sensi', 'Precision', 't_num', 't_acc','n_num', 'n_acc', ' ', 
                            'Iou_S2', 'Dice', 'Speci', 'Sensi', 'Precision', 't_num', 't_acc','n_num', 'n_acc'])
        wb_result.save(Dir_Statics_val)
    else:
        pt_epoch = args.model_pretrain.split('/')[-1].split('_')[1]
        for file_name in os.listdir(args.output_dir):
            if "Loss_train_" in file_name:
                Dir_Loss_train = f"{args.output_dir}/{file_name}"
            if "Loss_val_" in file_name:
                Dir_Loss_val = f"{args.output_dir}/{file_name}"
            if "Statics_train_" in file_name:
                Dir_Statics_train = f"{args.output_dir}/{file_name}"
            if "Statics_val_" in file_name:
                Dir_Statics_val = f"{args.output_dir}/{file_name}"
        args.start_epoch = int(pt_epoch) + 1
  
        ckpt_s1 = torch.load(f"{args.output_dir}/CheckPoint/DETR_{pt_epoch}_S1.pth", map_location=device)
        ckpt_s2 = torch.load(args.model_pretrain, map_location=device)

        model.load_state_dict(ckpt_s1['model_state_dict'])
        optimizer.load_state_dict(ckpt_s1['optimizer'])
        lr_scheduler.load_state_dict(ckpt_s1['lr_scheduler'])
        model_2.load_state_dict(ckpt_s2['model_state_dict'])
        optimizer_2.load_state_dict(ckpt_s2['optimizer'])
        lr_scheduler_2.load_state_dict(ckpt_s2['lr_scheduler'])
        
        pt_loss_train = pd.read_excel(Dir_Loss_train)
        pt_loss_val = pd.read_excel(Dir_Loss_val)
        pt_static_train = pd.read_excel(Dir_Statics_train)
        pt_static_val = pd.read_excel(Dir_Statics_val)

        for num in range(len(pt_loss_train)):
            train_bbox_s1.append(pt_loss_train['Bbox_S1'][num])
            train_ciou_s1.append(pt_loss_train['Ciou_S1'][num])
            train_score_s1.append(pt_loss_train['Score_S1'][num])
            train_bbox_s2.append(pt_loss_train['Bbox_S2'][num])
            train_ciou_s2.append(pt_loss_train['Ciou_S2'][num])
            train_score_s2.append(pt_loss_train['Score_S2'][num])

            val_bbox_s1.append(pt_loss_val['Bbox_S1'][num])
            val_ciou_s1.append(pt_loss_val['Ciou_S1'][num])
            val_score_s1.append(pt_loss_val['Score_S1'][num])
            val_bbox_s2.append(pt_loss_val['Bbox_S2'][num])
            val_ciou_s2.append(pt_loss_val['Ciou_S2'][num])
            val_score_s2.append(pt_loss_val['Score_S2'][num])

            train_iou_s1.append(pt_static_train['Iou_S1'][num])
            train_iou_s2.append(pt_static_train['Iou_S2'][num])
            val_iou_s1.append(pt_static_val['Iou_S1'][num])
            val_iou_s2.append(pt_static_val['Iou_S2'][num])
    
    if "Test" in args.output_dir:
        train_dataset = BoundingBoxDataSet(args.data_train, flag_shuffle=False)
        data_loader_train = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=False, num_workers=0)
    else:
        train_dataset = BoundingBoxDataSet(args.data_train, flag_shuffle=True)
        data_loader_train = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True, num_workers=0)

    del train_dataset

    for epoch in range(args.start_epoch, args.epochs):
        epoch_start = time.time()
        train_stats_s1, epoch_score_1, t_results_s1, train_stats_s2, epoch_score_2, t_results_s2, s2_data_num = train_epoch(args, postprocessors, model, criterion, data_loader_train, 
                                                                    optimizer, postprocessors_2, model_2, criterion_2, optimizer_2, device, epoch, args.clip_max_norm)
        lr_scheduler.step()
        lr_scheduler_2.step()
       
        str_num = len(str(epoch))
        zero_need = 4 - str_num
        z = ''
        for i in range(zero_need):
            z += '0'
        os.makedirs(f"{args.output_dir}/CheckPoint/", exist_ok=True)
        path_model = f"{args.output_dir}/CheckPoint/DETR_{z}{epoch}_S1.pth"
        torch.save({'model_state_dict':model.state_dict(), 'optimizer': optimizer.state_dict(),'lr_scheduler': lr_scheduler.state_dict()}, path_model)
        path_model = f"{args.output_dir}/CheckPoint/DETR_{z}{epoch}_S2.pth"
        torch.save({'model_state_dict':model_2.state_dict(), 'optimizer': optimizer_2.state_dict(),'lr_scheduler': lr_scheduler_2.state_dict(), 'args': args}, path_model)
        # extra checkpoint before LR drop and every 100 epochs
        # if epoch % args.lr_drop == 0 or epoch % 100 == 0:
        #     checkpoint_paths.append(output_dir / f'checkpoint{epoch:04}.pth')
        # for checkpoint_path in checkpoint_paths:
        #     utils.save_on_master({'model': model_without_ddp.state_dict(), 'optimizer': optimizer.state_dict(),'lr_scheduler': lr_scheduler.state_dict(), 'epoch': epoch,'args': args}, checkpoint_path)
        
        bbox_s1_t = train_stats_s1['loss_bbox']
        ciou_s1_t = train_stats_s1['loss_giou']
        score_s1_t = epoch_score_1
        bbox_s2_t = train_stats_s2['loss_bbox']
        ciou_s2_t = train_stats_s2['loss_giou']
        score_s2_t = epoch_score_2
        if s2_data_num < args.batch_size:
            bbox_s2_t = 0
            ciou_s2_t = 0
            score_s2_t = 0
        train_bbox_s1.append(bbox_s1_t)
        train_ciou_s1.append(ciou_s1_t)
        train_score_s1.append(score_s1_t)
        train_bbox_s2.append(bbox_s2_t)
        train_ciou_s2.append(ciou_s2_t)
        train_score_s2.append(score_s2_t)
        train_iou_s1.append(t_results_s1[0])
        train_iou_s2.append(t_results_s2[0])
        epoch_time = (time.time() - epoch_start) / 60
        time_H = int(epoch_time // 60)
        time_M = epoch_time % 60
        print(f"Epoch {epoch} Train Finish --- 1st Data = {len(data_loader_train.dataset)}, 2nd Data = {s2_data_num}, Train_Time = {time_H} H {round(time_M,2)} M")
        ### def evaluate(args, model, postprocessors, criterion, data_loader, model_2, postprocessors_2, criterion_2, device):
        # 학습이 잘 되고 있는지, 확인하기 위해서 위에서 사용하는 함수를 그대로 사용한다. 
        test_start = time.time()
        val_stats_s1, epoch_score_1, v_results_s1, val_stats_s2, epoch_score_2, v_results_s2, s2_data_num_v = evaluate(args, model, postprocessors, criterion, data_loader_test, 
                                                                                                                            model_2, postprocessors_2, criterion_2, epoch, device)
        
        bbox_s1_v = val_stats_s1['loss_bbox']
        ciou_s1_v = val_stats_s1['loss_giou']
        score_s1_v = epoch_score_2
        bbox_s2_v = val_stats_s2['loss_bbox']
        ciou_s2_v = val_stats_s2['loss_giou']
        score_s2_v = epoch_score_2
        if s2_data_num_v < args.batch_size:
            bbox_s2_v = 0
            ciou_s2_v = 0
            score_s2_v = 0
        val_bbox_s1.append(bbox_s1_v)
        val_ciou_s1.append(ciou_s1_v)
        val_score_s1.append(score_s1_v)
        val_bbox_s2.append(bbox_s2_v)
        val_ciou_s2.append(ciou_s2_v)
        val_score_s2.append(score_s2_v)
        val_iou_s1.append(v_results_s1[0])
        val_iou_s2.append(v_results_s2[0])
        epoch_time = (time.time() - test_start) / 60
        time_H = int(epoch_time // 60)
        time_M = epoch_time % 60
        print(f"Epoch {epoch} Val Finish --- 1st Data = {len(data_loader_test.dataset)}, 2nd Data = {s2_data_num_v}, Train_Time = {time_H} H {round(time_M,2)} M / {device}")


        wb = openpyxl.load_workbook(Dir_Loss_train)
        ws = wb.active
        result_list = [epoch, int(len(data_loader_train.dataset)), round(bbox_s1_t,4), round(ciou_s1_t,4), round(score_s1_t,4), s2_data_num, round(bbox_s2_t,4), round(ciou_s2_t,4), round(score_s2_t,4)]
        ws.append(result_list)
        wb.save(Dir_Loss_train)

        wb = openpyxl.load_workbook(Dir_Loss_val)
        ws = wb.active
        result_list = [epoch, int(len(data_loader_test.dataset)), round(bbox_s1_v,4), round(ciou_s1_v,4), round(score_s1_v,4), s2_data_num_v, round(bbox_s2_v,4), round(ciou_s2_v,4), round(score_s2_v,4)]
        ws.append(result_list)
        wb.save(Dir_Loss_val)

        wb = openpyxl.load_workbook(Dir_Statics_train)
        ws = wb.active
        result_list = [epoch, t_results_s1[0], t_results_s1[1], t_results_s1[2], t_results_s1[3], t_results_s1[4], t_results_s1[5], t_results_s1[6], t_results_s1[7], t_results_s1[8], "", 
                        t_results_s2[0], t_results_s2[1], t_results_s2[2], t_results_s2[3], t_results_s2[4], t_results_s2[5], t_results_s2[6], t_results_s2[7], t_results_s2[8]]
        ws.append(result_list)
        wb.save(Dir_Statics_train)

        wb = openpyxl.load_workbook(Dir_Statics_val)
        ws = wb.active
        result_list = [epoch, v_results_s1[0], v_results_s1[1], v_results_s1[2], v_results_s1[3], v_results_s1[4], v_results_s1[5], v_results_s1[6], v_results_s1[7], v_results_s1[8], "", 
                        v_results_s2[0], v_results_s2[1], v_results_s2[2], v_results_s2[3], v_results_s2[4], v_results_s2[5], v_results_s2[6], v_results_s2[7], v_results_s2[8]]
        ws.append(result_list)
        wb.save(Dir_Statics_val)
  

        epoch_x = np.array(range(len(train_bbox_s1)))
        epoch_x += 1
        plt.plot(epoch_x, train_score_s1, 'g', label='Train Score')
        plt.plot(epoch_x, val_score_s1, 'r', label='Val Score')
        plt.title('Score')
        plt.xlabel("Epoch")
        plt.ylabel("Score")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Score_S1.jpg")
        plt.close()

        plt.plot(epoch_x, train_score_s2, 'g', label='Train Score')
        plt.plot(epoch_x, val_score_s2, 'r', label='Val Score')
        plt.title('Score')
        plt.xlabel("Epoch")
        plt.ylabel("Score")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Score_S2.jpg")
        plt.close()

        plt.plot(epoch_x, train_bbox_s1, 'g', label='Train Loss')
        plt.plot(epoch_x, val_bbox_s1, 'r', label='Val Loss')
        plt.title('Train and Val Loss')
        plt.xlabel("Epoch")
        plt.ylabel("Bbox Loss")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Bbox_S1.jpg")
        plt.close()

        plt.plot(epoch_x, train_bbox_s2, 'g', label='Train Loss')
        plt.plot(epoch_x, val_bbox_s2, 'r', label='Val Loss')
        plt.title('Train and Val Loss')
        plt.xlabel("Epoch")
        plt.ylabel("Bbox Loss")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Bbox_S2.jpg")
        plt.close()

        plt.plot(epoch_x, train_ciou_s1, 'g', label='Train Loss')
        plt.plot(epoch_x, val_ciou_s1, 'r', label='Val Loss')
        plt.title('Train and Val Loss')
        plt.xlabel("Epoch")
        plt.ylabel("CIOU Loss")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Ciou_S1.jpg")
        plt.close()

        plt.plot(epoch_x, train_ciou_s2, 'g', label='Train Loss')
        plt.plot(epoch_x, val_ciou_s2, 'r', label='Val Loss')
        plt.title('Train and Val Loss')
        plt.xlabel("Epoch")
        plt.ylabel("CIOU Loss")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_Ciou_S2.jpg")
        plt.close()

        plt.plot(epoch_x, train_iou_s1, 'g', label='Train IOU')
        plt.plot(epoch_x, val_iou_s1, 'r', label='Val IOU')
        plt.title('IOU')
        plt.xlabel("Epoch")
        plt.ylabel("Score")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_IOU_S1.jpg")
        plt.close()

        plt.plot(epoch_x, train_iou_s2, 'g', label='Train IOU')
        plt.plot(epoch_x, val_iou_s2, 'r', label='Val IOU')
        plt.title('IOU')
        plt.xlabel("Epoch")
        plt.ylabel("Score")
        plt.legend(loc='upper left')
        plt.xticks(range(0, len(train_bbox_s1)+1))
        plt.tight_layout()
        plt.savefig(f"{args.output_dir}Result_IOU_S2.jpg")
        plt.close()

    total_time = time.time() - start_time
    total_time_str = str(datetime.timedelta(seconds=int(total_time)))
    print('Training time {}'.format(total_time_str))

    print(f"{args.output_dir} / {args.device} FINSH")
    

if __name__ == '__main__':
    parser = argparse.ArgumentParser('DETR training and evaluation script', parents=[get_args_parser()])
    args = parser.parse_args()
    print(f"Dilation Mode = {args.dilation}")
    
    # if not args.eval and not args.add_hg:
    #     if args.output_dir:
    #         Path(args.output_dir).mkdir(parents=True, exist_ok=True)
    main(args)


