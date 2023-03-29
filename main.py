# Copyright (c) Facebook, Inc. and its affiliates. All Rights Reserved
import argparse
import json
import random
import time
from pathlib import Path

import numpy as np
import torch
from torch.utils.data import DataLoader

import util.misc as utils
from engine import evaluate_val
from models import build_model

from via import via

import random

from BoundingBoxDataSet import BoundingBoxDataSet
import os
os.environ['KMP_DUPLICATE_LIB_OK'] = 'True'
import openpyxl
from calculate import *

def get_args_parser():
    GPU_Name = "cuda:0"
    Box_TH = 0.9
    Box_UseNum = 4

    # Back_bone = 'resnet50'
    NAME_RESULT = f"results"
    DIR_RESULT = f"./{NAME_RESULT}/"
    os.makedirs(DIR_RESULT, exist_ok=True)

    parser = argparse.ArgumentParser('Set transformer detector', add_help=False)
    parser.add_argument('--lr', default=1e-4, type=float)
    parser.add_argument('--lr_backbone', default=1e-5, type=float)
    parser.add_argument('--batch_size', default=4, type=int)
    parser.add_argument('--weight_decay', default=1e-4, type=float)
    parser.add_argument('--epochs', default=20, type=int)
    parser.add_argument('--lr_drop', default=200, type=int)
    parser.add_argument('--clip_max_norm', default=0.1, type=float, help='gradient clipping max norm')

    # Model parameters
    parser.add_argument('--frozen_weights', type=str, default=None, help="Path to the pretrained model. If set, only the mask head will be trained")
    # * Backbone
    parser.add_argument('--backbone', default="resnet50", type=str, help="Name of the convolutional backbone to use")
    parser.add_argument('--dilation', default=True,
                        help="If true, we replace stride with dilation in the last convolutional block (DC5)") # stride를 적용하지 않고, dilation을 추가해서 backbone output resolution을 높힌다.
    parser.add_argument('--position_embedding', default='sine', type=str, choices=('sine', 'learned'), help="Type of positional embedding to use on top of the image features")
    parser.add_argument('--pretraind_model', default="detr-r50-dc5.pth", type=str, help="Name of the pretrain model")

    # * Transformer
    parser.add_argument('--enc_layers', default=6, type=int,
                        help="Number of encoding layers in the transformer")
    parser.add_argument('--dec_layers', default=6, type=int,
                        help="Number of decoding layers in the transformer")
    parser.add_argument('--dim_feedforward', default=2048, type=int,
                        help="Intermediate size of the feedforward layers in the transformer blocks")
    parser.add_argument('--hidden_dim', default=256, type=int,
                        help="Size of the embeddings (dimension of the transformer)")
    parser.add_argument('--dropout', default=0.1, type=float,
                        help="Dropout applied in the transformer")
    parser.add_argument('--nheads', default=8, type=int,
                        help="Number of attention heads inside the transformer's attentions")
    parser.add_argument('--num_queries', default=6, type=int, help="Number of query slots")
    parser.add_argument('--num_classes', default=2, type=int, help="Number of class")
    parser.add_argument('--zero_label', default=0, type=int, help="zero label 0 or 2")
    parser.add_argument('--pre_norm', action='store_true')

    # * Segmentation
    parser.add_argument('--masks', action='store_true',
                        help="Train segmentation head if the flag is provided")

    # Loss
    parser.add_argument('--no_aux_loss', dest='aux_loss', action='store_false',
                        help="Disables auxiliary decoding losses (loss at each layer)")  # args.aux_loss == false
                        
    # * Matcher
    parser.add_argument('--set_cost_class', default=1, type=float, help="Class coefficient in the matching cost")
    parser.add_argument('--set_cost_bbox', default=5, type=float, help="L1 box coefficient in the matching cost")
    parser.add_argument('--set_cost_giou', default=2, type=float, help="giou box coefficient in the matching cost")

    # * Loss coefficients
    parser.add_argument('--mask_loss_coef', default=1, type=float)
    parser.add_argument('--dice_loss_coef', default=1, type=float)
    parser.add_argument('--bbox_loss_coef', default=5, type=float)
    parser.add_argument('--giou_loss_coef', default=2, type=float)
    parser.add_argument('--eos_coef', default=0.1, type=float, help="Relative classification weight of the no-object class")

    # dataset parameters
    parser.add_argument('--dataset_file', default='coco')
    parser.add_argument('--coco_path', type=str)
    parser.add_argument('--coco_panoptic_path', type=str)
    parser.add_argument('--remove_difficult', action='store_true')

    parser.add_argument('--output_dir', default=DIR_RESULT, help='path where to save, empty for no saving')
    parser.add_argument('--output_name', default=NAME_RESULT, help='path where to save name')
    parser.add_argument('--device', default=GPU_Name, help='device to use for training / testing')
    parser.add_argument('--seed', default=42, type=int)
    parser.add_argument('--resume', default='', help='resume from checkpoint')
    parser.add_argument('--start_epoch', default=1, type=int, metavar='N', help='start epoch')
    parser.add_argument('--eval', action='store_true')
    parser.add_argument('--num_workers', default=2, type=int) # 

    # distributed training parameters
    parser.add_argument('--world_size', default=1, type=int, help='number of distributed processes')
    parser.add_argument('--dist_url', default='env://', help='url used to set up distributed training')

    ## parameters
    parser.add_argument('--box_th', default=Box_TH, type=float, help='box threshold')
    parser.add_argument('--box_mode', default="SB", type=str, help='box mode')
    parser.add_argument('--box_usenum', default=Box_UseNum, type=int, help='box usenum')
    return parser


def main(args):
    utils.init_distributed_mode(args) # Multi-GPU 사용할 거라면, args.gpu / args.world_size / args.rank 가 여기서 정의 된다.
    # print("git:\n  {}\n".format(utils.get_sha()))

    if args.frozen_weights is not None:
        assert args.masks, "Frozen training is meant for segmentation only"
    # print(args)
    
    device = torch.device(args.device)
    print(f"Using device = {device} - {torch.cuda.get_device_name(device)}")

    # Multi-GPU 사용할 거라면, fix the seed for reproducibility 
    seed = args.seed + utils.get_rank() 
    torch.manual_seed(seed)
    np.random.seed(seed)
    random.seed(seed)

    ## 전까지는, args 설정 및 seed 설정이 전부이다.
    ## 여기가 진짜 시작 
    model, criterion, postprocessors = build_model(args)
    ## The models are also available via torch hub, to load DETR R50 with pretrained weights simply do:
    checkpoint = torch.load(f"Z:/Backup/Users/kys/BoundingBox/code/DETR/{args.pretraind_model}", map_location=device)
    del checkpoint["model"]["class_embed.weight"]
    del checkpoint["model"]["class_embed.bias"]
    del checkpoint["model"]["query_embed.weight"]
    model.load_state_dict(checkpoint['model'], strict=False)
    
    model.to(device)

    model_without_ddp = model
    
    ## 총 파라메터 갯수 계산 하는 방법
    n_parameters = sum(p.numel() for p in model.parameters() if p.requires_grad) 
    print('number of params:', n_parameters)


    ## Model Parameters load 하기(1) - frozen_weights. panoptic segmentaion 모듈만 학습시키고 싶다면
    if args.frozen_weights is not None:
        checkpoint = torch.load(args.frozen_weights, map_location='cpu')
        model_without_ddp.detr.load_state_dict(checkpoint['model'])

    ## Evaluation
    args.batch_size = 8
    DIR_Val_data = "./samples"

    model_name = f"./model.pth"
    ## val data load
    val_dataset = BoundingBoxDataSet(DIR_Val_data, flag_shuffle=False)
    data_loader_val = DataLoader(val_dataset, batch_size=args.batch_size, shuffle=False, num_workers=0)
    del val_dataset

    ckpt = torch.load(model_name, map_location=device)
    model.load_state_dict(ckpt['model_state_dict'])

    time_epoch_val_start = time.time()
    results = evaluate_val(args, model, criterion, postprocessors, data_loader_val, device, args.output_dir)
    ## remain_result = [remain_box_list, remain_score_list, remain_score_weight_list, remain_tumor_score, remain_tumor_score_weight, remain_zero_score, remain_zero_score_weight]
    print("")
    cal = results[0]
    ## init excel
    path_excel_result = f"{args.output_dir}Statics.xlsx"
    wb_result = openpyxl.Workbook()
    worksheet = wb_result.active
    worksheet.append(['Type', 'IoU', 'Dice', 'Precision', 'Sensitivity', 'Specificity', 'Score_1st', "Score_2st"])
    wb_result.save(path_excel_result)

    path_excel_val = f'{args.output_dir}Statics_Slice.xlsx'
    wb_val = openpyxl.Workbook()
    worksheet = wb_val.active
    worksheet.append(["Patient", "Dx", "GT 1", "Pred 1", "GT 2", "Pred 2", "IoU", "Dice", "Precision", "Sensitivity", "Specificity", "Score_1st", "Score_2st"])
    wb_val.save(path_excel_val)

    path_excel_val_box = f'{args.output_dir}Statics_BoxInfo.xlsx'
    wb_val_box = openpyxl.Workbook()
    worksheet = wb_val_box.active
    worksheet.append(["Patient", "Dx", "GT 1", "Pred 1", "GT 2", "Pred 2", "B1 1", "B1 2", "B1_S", "B1_S_W", "B2 1", "B2 2", "B2_S", "B2_S_W",
                        "B3 1", "B3 2", "B3_S", "B3_S_W", "B4 1", "B4 2", "B4_S", "B4_S_W"])
    wb_val_box.save(path_excel_val_box)
    
    ## Write excel
    wb_result = openpyxl.load_workbook(path_excel_result)
    ws = wb_result.active
    reulst_list = [f"All", round(np.mean(cal[0]), 4), round(np.mean(cal[1]), 4), round(np.mean(cal[2]), 4), round(np.mean(cal[3]), 4), round(np.mean(cal[4]), 4), round(np.mean(results[8]), 4), 
                    round(np.mean(results[9]), 4)]
    ws.append(reulst_list)
    iou = []
    dice = []
    precision = []
    sensitiy = []
    speci = []
    scores = []
    scores_2 = []
    for i in range(len(results[6])):
        iou.append(results[6][i][0])
        dice.append(results[6][i][1])
        precision.append(results[6][i][2])
        sensitiy.append(results[6][i][3])
        speci.append(results[6][i][4])
        scores.append(results[6][i][5])
        scores_2.append(results[6][i][6])
    reulst_list = ["Tumor", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(precision), 4), round(np.mean(sensitiy), 4), round(np.mean(speci), 4),
                    round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
    ws.append(reulst_list)
    iou = []
    dice = []
    precision = []
    sensitiy = []
    speci = []
    scores = []
    scores_2 = []
    for i in range(len(results[7])):
        iou.append(results[7][i][0])
        dice.append(results[7][i][1])
        precision.append(results[7][i][2])
        sensitiy.append(results[7][i][3])
        speci.append(results[7][i][4])
        scores.append(results[7][i][5])
        scores_2.append(results[7][i][6])
    reulst_list = ["Normal", round(np.mean(iou), 4), round(np.mean(dice), 4), round(np.mean(precision), 4), round(np.mean(sensitiy), 4), round(np.mean(speci), 4), 
                    round(np.mean(scores), 4), round(np.mean(scores_2), 4)]
    ws.append(reulst_list)
    wb_result.save(path_excel_result)

    wb_eval = openpyxl.load_workbook(path_excel_val)
    ws = wb_eval.active
    for s in range(len(results[1])):
        print(f"Write Slice Excel {s+1} / {len(results[1])}", end="\r")
        reulst_list = [results[1][s], results[2][s], str(results[4][s][0])[1:-1], str(results[5][s][0])[1:-1], str(results[4][s][1])[1:-1], str(results[5][s][1])[1:-1], 
                        np.round(cal[0][s],4), np.round(cal[1][s],4), np.round(cal[2][s],4), np.round(cal[3][s],4),  np.round(cal[4][s],4), 
                        np.round(results[8][s],4), np.round(results[9][s],4)]
        ws.append(reulst_list)
    wb_eval.save(path_excel_val)
    
    wb_eval_box = openpyxl.load_workbook(path_excel_val_box)
    ws = wb_eval_box.active
    for s in range(len(results[1])):
        print(f"Write BoxInfo Excel {s+1} / {len(results[1])}", end="\r")
        reulst_list = [results[1][s], results[2][s], str(results[4][s][0])[1:-1], str(results[5][s][0])[1:-1], str(results[4][s][1])[1:-1], str(results[5][s][1])[1:-1], 
                        str(results[11][0][s][0][0])[1:-1], str(results[11][0][s][0][1])[1:-1], round(results[11][1][s][0],4), round(results[11][2][s][0],4),
                        str(results[11][0][s][1][0])[1:-1], str(results[11][0][s][1][1])[1:-1], round(results[11][1][s][1],4), round(results[11][2][s][1],4),
                        str(results[11][0][s][2][0])[1:-1], str(results[11][0][s][2][1])[1:-1], round(results[11][1][s][2],4), round(results[11][2][s][2],4),
                        str(results[11][0][s][3][0])[1:-1], str(results[11][0][s][3][1])[1:-1], round(results[11][1][s][3],4), round(results[11][2][s][3],4)]
        ws.append(reulst_list)
    wb_eval_box.save(path_excel_val_box)
    print("")

    time_epoch_val_end = time.time()
    time_epoch_val = (time_epoch_val_end - time_epoch_val_start) / 60
    time_eval_H = int(time_epoch_val // 60)
    time_eval_M = round((time_epoch_val % 60), 2)
    print(f"Eval --- {model_name} Finish ---- Time = {time_eval_H} h {time_eval_M} m")
   

if __name__ == '__main__':
    parser = argparse.ArgumentParser('DETR training and evaluation script', parents=[get_args_parser()])
    args = parser.parse_args()

    main(args)


